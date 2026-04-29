from __future__ import annotations

import csv
from io import BytesIO
from io import StringIO
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

import sqlalchemy as sa
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app import db
from app.models import (
    MapIcdCodBucket,
    MapIcd10LegacyReportingAlias,
    MasCodBucketNode,
    MasCodBucketScheme,
    MasCodBucketSchemeAgeBand,
    MasIcd1020192,
    VaSmartvaResults,
    VaForms,
    VaSubmissions,
)
from app.services.submission_analytics_mv import (
    CORE_MV_NAME,
    COD_MV_NAME,
    DEMOGRAPHICS_MV_NAME,
)

SCHEME_CODE_SRS_INDIA = "SRS_INDIA"
SCHEME_CODE_CMEA10 = "CMEA10"
SCHEME_CODE_WHO_2022_VA = "WHO_2022_VA"

AGE_SCOPE_ADULT_OVER5Y = "adult_over5y"
AGE_SCOPE_CHILD_1_59M = "child_1_59m"
AGE_SCOPE_NEONATE = "neonate"

REPORTING_AGE_BAND_0_27_DAYS = "0-<28 days"
REPORTING_AGE_BAND_28_364_DAYS = "28 days-<365 days"
REPORTING_AGE_BAND_365_DAYS_TO_11_YEARS = "365 days-<12 years"
REPORTING_AGE_BAND_12_49_YEARS = "12 years-<50 years"
REPORTING_AGE_BAND_50_PLUS_YEARS = ">=50 years"

AGE_UNIT_DAYS = "days"
AGE_UNIT_MONTHS = "months"
AGE_UNIT_YEARS = "years"

SMARTVA_PSEUDO_ICD_ROWS = (
    {
        "code": "UU1",
        "title": "Other Non-communicable Diseases",
        "semantic_level": "three_character",
        "is_coding_selectable": False,
        "chapter_code": "",
        "chapter_title": "SmartVA pseudo-codes",
        "three_character_code": "UU1",
        "three_character_title": "Other Non-communicable Diseases",
    },
    {
        "code": "UU2",
        "title": "Other Defined Causes of Child Deaths",
        "semantic_level": "three_character",
        "is_coding_selectable": False,
        "chapter_code": "",
        "chapter_title": "SmartVA pseudo-codes",
        "three_character_code": "UU2",
        "three_character_title": "Other Defined Causes of Child Deaths",
    },
)
AGE_UNITS = (AGE_UNIT_DAYS, AGE_UNIT_MONTHS, AGE_UNIT_YEARS)
DEFAULT_MIN_AGE_VALUE = 0
DEFAULT_MIN_AGE_UNIT = AGE_UNIT_DAYS
DEFAULT_MAX_AGE_VALUE = 120
DEFAULT_MAX_AGE_UNIT = AGE_UNIT_YEARS

NODE_TYPE_CATEGORY = "category"
NODE_TYPE_SUBCATEGORY = "subcategory"
NODE_TYPE_FIELD = "field"

NODE_DELETE_DISPOSITION_UNMAP = "unmap"
NODE_DELETE_DISPOSITION_MOVE_TO_UNMAPPED = "move_to_unmapped"

MIGRATION_ARTIFACT_SRS_WORKBOOK_PATH = (
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "srs-india-cod-2026-04-27/icd-10-CODES_SRS_India.xlsx"
)
MIGRATION_ARTIFACT_CMEA10_WORKBOOK_PATH = (
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "cmea10-cod-2026-04-27/icd-10-CODES_CMEA10_mapped.xlsx"
)
MIGRATION_ARTIFACT_WHO_2022_VA_WORKBOOK_PATH = (
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "who-2022-va-icd-cod-2026-04-27/WHO_2022_VA_Bucket_Mapping_document_derived.xlsx"
)
DEFAULT_SRS_WORKBOOK_PATH = MIGRATION_ARTIFACT_SRS_WORKBOOK_PATH
DEFAULT_CMEA10_WORKBOOK_PATH = MIGRATION_ARTIFACT_CMEA10_WORKBOOK_PATH
DEFAULT_WHO_2022_VA_WORKBOOK_PATH = MIGRATION_ARTIFACT_WHO_2022_VA_WORKBOOK_PATH
SOURCE_RESETTABLE_SCHEME_CODES = {
    SCHEME_CODE_SRS_INDIA,
    SCHEME_CODE_CMEA10,
    SCHEME_CODE_WHO_2022_VA,
}
MANUAL_OVERRIDE_SOURCE_SHEET = "admin_cod_bucket_editor"
MANUAL_OVERRIDE_MATCH_TYPE = "manual_override"
MANUAL_OVERRIDE_NOTE = "Manual override to default COD bucket scheme mapping."

_LIKE_ESCAPE = "\\"
_ICD_SEARCH_LIMIT = 30
_DAYS_PER_MONTH = 365 / 12

_SRS_SCOPE_CONFIG = (
    (
        AGE_SCOPE_ADULT_OVER5Y,
        "SRS_India_over5y_main_group",
        "SRS_India_over5y_sub_group",
        "SRS_India_over5y_disease",
        "SRS_India_over5y_match_type",
        "SRS_India_over5y_note",
    ),
    (
        AGE_SCOPE_NEONATE,
        "SRS_India_neonate_main_group",
        "SRS_India_neonate_sub_group",
        "SRS_India_neonate_disease",
        "SRS_India_neonate_match_type",
        "SRS_India_neonate_note",
    ),
    (
        AGE_SCOPE_CHILD_1_59M,
        "SRS_India_1_59mth_main_group",
        "SRS_India_1_59mth_sub_group",
        "SRS_India_1_59mth_disease",
        "SRS_India_1_59mth_type",
        "SRS_India_1_59mth_note",
    ),
)


def _normalize_icd_code(value) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip().upper()
    return normalized or None


def _normalize_label(value) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value).strip())
    return normalized or None


def _normalize_search_query(value: str | None) -> str:
    return " ".join((value or "").strip().lower().split())


def _normalize_reporting_icd_code(value) -> str | None:
    normalized = _normalize_icd_code(value)
    if normalized is None:
        return None
    alias = db.session.get(MapIcd10LegacyReportingAlias, normalized)
    return alias.reporting_code if alias else normalized


def _reporting_icd_sql(column, alias_column):
    return sa.func.coalesce(alias_column, column)


def _normalize_gender_label(value: str | None) -> str:
    normalized = _normalize_label(value)
    if normalized is None:
        return "Unknown"
    lowered = normalized.lower()
    if lowered == "male":
        return "Male"
    if lowered == "female":
        return "Female"
    if lowered in {"other", "non-binary", "non binary"}:
        return "Other"
    return "Unknown"


def _normalize_gender_filter(value: str | None) -> str | None:
    normalized = _normalize_label(value)
    if normalized is None:
        return None
    lowered = normalized.lower()
    if lowered in {"male", "female", "unknown"}:
        return lowered
    return None


def _gender_normalized_sql(column):
    return sa.func.lower(sa.func.trim(sa.func.coalesce(column, "unknown")))


def _gender_filter_clause(column, gender_filter: str | None):
    normalized = _gender_normalized_sql(column)
    if gender_filter == "male":
        return normalized == "male"
    if gender_filter == "female":
        return normalized == "female"
    if gender_filter == "unknown":
        return sa.and_(normalized != "male", normalized != "female")
    return sa.true()


def _reporting_age_band_sql(column):
    return sa.case(
        (
            column < sa.literal(28),
            REPORTING_AGE_BAND_0_27_DAYS,
        ),
        (
            column < sa.literal(365),
            REPORTING_AGE_BAND_28_364_DAYS,
        ),
        (
            column < sa.literal(12 * 365),
            REPORTING_AGE_BAND_365_DAYS_TO_11_YEARS,
        ),
        (
            column < sa.literal(50 * 365),
            REPORTING_AGE_BAND_12_49_YEARS,
        ),
        else_=REPORTING_AGE_BAND_50_PLUS_YEARS,
    )


def _reporting_age_band_sort_sql(column):
    return sa.case(
        (
            column < sa.literal(28),
            1,
        ),
        (
            column < sa.literal(365),
            2,
        ),
        (
            column < sa.literal(12 * 365),
            3,
        ),
        (
            column < sa.literal(50 * 365),
            4,
        ),
        else_=5,
    )


def _reporting_age_band_order():
    return [
        REPORTING_AGE_BAND_0_27_DAYS,
        REPORTING_AGE_BAND_28_364_DAYS,
        REPORTING_AGE_BAND_365_DAYS_TO_11_YEARS,
        REPORTING_AGE_BAND_12_49_YEARS,
        REPORTING_AGE_BAND_50_PLUS_YEARS,
    ]


def _uses_reporting_age_band_detail_sections(scheme: MasCodBucketScheme) -> bool:
    if scheme.scheme_code in {SCHEME_CODE_CMEA10, SCHEME_CODE_WHO_2022_VA}:
        return True
    age_scopes = db.session.scalars(
        sa.select(MasCodBucketSchemeAgeBand.age_scope).where(
            MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id,
            MasCodBucketSchemeAgeBand.is_active.is_(True),
        )
    ).all()
    return bool(age_scopes) and all(age_scope is None for age_scope in age_scopes)


def _detailed_reporting_age_scope_sort_sql(column):
    return _reporting_age_band_sort_sql(column) + sa.literal(1)


def _reporting_age_band_label_for_days(age_normalized_days) -> str:
    if age_normalized_days is None:
        return REPORTING_AGE_BAND_50_PLUS_YEARS
    if age_normalized_days < 28:
        return REPORTING_AGE_BAND_0_27_DAYS
    if age_normalized_days < 365:
        return REPORTING_AGE_BAND_28_364_DAYS
    if age_normalized_days < (12 * 365):
        return REPORTING_AGE_BAND_365_DAYS_TO_11_YEARS
    if age_normalized_days < (50 * 365):
        return REPORTING_AGE_BAND_12_49_YEARS
    return REPORTING_AGE_BAND_50_PLUS_YEARS


def get_reporting_icd_alias_rows() -> list[dict]:
    rows = db.session.execute(
        sa.select(
            MapIcd10LegacyReportingAlias.legacy_code,
            MapIcd10LegacyReportingAlias.reporting_code,
            MapIcd10LegacyReportingAlias.note,
            MasIcd1020192.title.label("reporting_title"),
        )
        .select_from(MapIcd10LegacyReportingAlias)
        .join(MasIcd1020192, MasIcd1020192.code == MapIcd10LegacyReportingAlias.reporting_code)
        .order_by(MapIcd10LegacyReportingAlias.legacy_code.asc())
    ).mappings().all()
    return [dict(row) for row in rows]


def create_reporting_icd_alias(*, legacy_code: str, reporting_code: str, note: str | None = None) -> dict:
    normalized_legacy_code = _normalize_icd_code(legacy_code)
    normalized_reporting_code = _normalize_icd_code(reporting_code)
    normalized_note = _normalize_label(note)

    if not normalized_legacy_code or not normalized_reporting_code:
        raise ValueError("legacy_code and reporting_code are required.")
    if normalized_legacy_code == normalized_reporting_code:
        raise ValueError("legacy_code and reporting_code must differ.")
    if db.session.get(MapIcd10LegacyReportingAlias, normalized_legacy_code):
        raise ValueError(f"Legacy ICD alias '{normalized_legacy_code}' already exists.")
    if db.session.get(MasIcd1020192, normalized_legacy_code):
        raise ValueError(
            f"Legacy ICD code '{normalized_legacy_code}' already exists in the ICD-10 2019 master."
        )

    master_row = db.session.get(MasIcd1020192, normalized_reporting_code)
    if (
        master_row is None
        or not master_row.is_active
        or master_row.semantic_level not in {"three_character", "detailed_code"}
    ):
        raise ValueError(
            f"Reporting ICD code '{normalized_reporting_code}' must be an active ICD-10 2019 three-character or detailed code."
        )

    alias = MapIcd10LegacyReportingAlias(
        legacy_code=normalized_legacy_code,
        reporting_code=normalized_reporting_code,
        note=normalized_note,
    )
    db.session.add(alias)
    db.session.commit()
    return next(
        row
        for row in get_reporting_icd_alias_rows()
        if row["legacy_code"] == normalized_legacy_code
    )


def delete_reporting_icd_alias(*, legacy_code: str) -> dict:
    normalized_legacy_code = _normalize_icd_code(legacy_code)
    alias = db.session.get(MapIcd10LegacyReportingAlias, normalized_legacy_code)
    if alias is None:
        raise LookupError(f"Legacy ICD alias '{normalized_legacy_code}' not found.")

    db.session.delete(alias)
    db.session.commit()
    return {"deleted": normalized_legacy_code}


def _escape_like(value: str) -> str:
    return (
        value.replace(_LIKE_ESCAPE, _LIKE_ESCAPE * 2)
        .replace("%", f"{_LIKE_ESCAPE}%")
        .replace("_", f"{_LIKE_ESCAPE}_")
    )


def _reporting_icd_catalog_subquery():
    master_rows = (
        sa.select(
            MasIcd1020192.code.label("code"),
            MasIcd1020192.title.label("title"),
            MasIcd1020192.semantic_level.label("semantic_level"),
            MasIcd1020192.is_coding_selectable.label("is_coding_selectable"),
            MasIcd1020192.chapter_code.label("chapter_code"),
            MasIcd1020192.chapter_title.label("chapter_title"),
            MasIcd1020192.three_character_code.label("three_character_code"),
            MasIcd1020192.three_character_title.label("three_character_title"),
        )
        .where(
            MasIcd1020192.is_active.is_(True),
            MasIcd1020192.semantic_level.in_(("three_character", "detailed_code")),
        )
    )
    pseudo_rows = [
        sa.select(
            sa.literal(row["code"]).label("code"),
            sa.literal(row["title"]).label("title"),
            sa.literal(row["semantic_level"]).label("semantic_level"),
            sa.literal(row["is_coding_selectable"]).label("is_coding_selectable"),
            sa.literal(row["chapter_code"]).label("chapter_code"),
            sa.literal(row["chapter_title"]).label("chapter_title"),
            sa.literal(row["three_character_code"]).label("three_character_code"),
            sa.literal(row["three_character_title"]).label("three_character_title"),
        )
        for row in SMARTVA_PSEUDO_ICD_ROWS
    ]
    return sa.union_all(master_rows, *pseudo_rows).subquery()


def _icd_master_display_subquery():
    catalog_sq = _reporting_icd_catalog_subquery()
    return (
        sa.select(
            catalog_sq.c.code.label("icd_code"),
            sa.func.min(
                sa.func.concat(catalog_sq.c.code, sa.literal("-"), catalog_sq.c.title)
            ).label("icd_to_display"),
        )
        .select_from(catalog_sq)
        .group_by(catalog_sq.c.code)
        .subquery()
    )


def _slugify(value: str, *, fallback_prefix: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or fallback_prefix


def _approx_age_days(value: int | None, unit: str | None) -> float | None:
    if value is None:
        return None
    if unit == AGE_UNIT_DAYS:
        return float(value)
    if unit == AGE_UNIT_MONTHS:
        return float(value) * _DAYS_PER_MONTH
    if unit == AGE_UNIT_YEARS:
        return float(value) * 365
    return None
def _scope_clause(column, age_scope: str | None):
    return column.is_(None) if age_scope is None else column == age_scope


def _scheme_default_source_path(scheme_code: str) -> str | None:
    if scheme_code == SCHEME_CODE_SRS_INDIA:
        return DEFAULT_SRS_WORKBOOK_PATH
    if scheme_code == SCHEME_CODE_CMEA10:
        return DEFAULT_CMEA10_WORKBOOK_PATH
    if scheme_code == SCHEME_CODE_WHO_2022_VA:
        return DEFAULT_WHO_2022_VA_WORKBOOK_PATH
    return None


def _scheme_source_path(scheme: MasCodBucketScheme) -> Path | None:
    candidate = scheme.source_path or _scheme_default_source_path(scheme.scheme_code)
    if not candidate:
        return None
    return Path(candidate)


def _scheme_reset_source_path(scheme: MasCodBucketScheme) -> Path | None:
    if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
        return Path(MIGRATION_ARTIFACT_SRS_WORKBOOK_PATH)
    if scheme.scheme_code == SCHEME_CODE_CMEA10:
        return Path(MIGRATION_ARTIFACT_CMEA10_WORKBOOK_PATH)
    if scheme.scheme_code == SCHEME_CODE_WHO_2022_VA:
        return Path(MIGRATION_ARTIFACT_WHO_2022_VA_WORKBOOK_PATH)
    return None


def scheme_can_reset_from_source(scheme: MasCodBucketScheme) -> bool:
    path = _scheme_reset_source_path(scheme)
    return scheme.scheme_code in SOURCE_RESETTABLE_SCHEME_CODES and path is not None and path.exists()


def _age_band_can_reset_from_source(
    scheme: MasCodBucketScheme,
    age_band: MasCodBucketSchemeAgeBand,
) -> bool:
    if not scheme_can_reset_from_source(scheme):
        return False
    if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
        return age_band.age_scope in {
            AGE_SCOPE_ADULT_OVER5Y,
            AGE_SCOPE_CHILD_1_59M,
            AGE_SCOPE_NEONATE,
        }
    if scheme.scheme_code == SCHEME_CODE_CMEA10:
        return age_band.age_scope is None
    if scheme.scheme_code == SCHEME_CODE_WHO_2022_VA:
        return age_band.age_scope is None
    return False


def _serialize_age_band(age_band: MasCodBucketSchemeAgeBand) -> dict:
    return {
        "age_band_id": str(age_band.age_band_id),
        "value": age_band.age_scope or "",
        "label": age_band.age_label,
        "min_age_value": age_band.min_age_value,
        "min_age_unit": age_band.min_age_unit,
        "max_age_value": age_band.max_age_value,
        "max_age_unit": age_band.max_age_unit,
        "level_count": age_band.level_count,
        "sort_order": age_band.sort_order,
    }


def _age_scope_warning_messages(age_bands: list[MasCodBucketSchemeAgeBand]) -> dict[str, list[str]]:
    warnings_by_scope: dict[str, list[str]] = {}
    ordered = sorted(age_bands, key=lambda band: (band.sort_order, band.age_label.lower()))
    previous_band = None
    previous_max_days = None
    for band in ordered:
        scope_key = band.age_scope or ""
        band_warnings: list[str] = []
        min_days = _approx_age_days(band.min_age_value, band.min_age_unit)
        max_days = _approx_age_days(band.max_age_value, band.max_age_unit)
        if (
            band.min_age_value is not None
            and band.max_age_value is not None
            and min_days is not None
            and max_days is not None
            and min_days >= max_days
        ):
            band_warnings.append(
                "Upper bound must be greater than lower bound because age bands use >= lower and < upper."
            )
        if previous_band is not None and previous_max_days is not None and min_days is not None:
            if min_days < previous_max_days:
                band_warnings.append(
                    f"Overlaps with {previous_band.age_label} using approximate unit conversion."
                )
            elif min_days > previous_max_days:
                band_warnings.append(
                    f"Gap after {previous_band.age_label} using approximate unit conversion."
                )
        if band_warnings:
            warnings_by_scope[scope_key] = band_warnings
        previous_band = band
        previous_max_days = max_days
    return warnings_by_scope


def _load_sheet_rows(workbook_path: str | Path, sheet_name: str):
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    for row_number, row in enumerate(rows[1:], start=2):
        payload = {headers[idx]: row[idx] for idx in range(len(headers))}
        yield row_number, payload


def _who_2022_default_mapping_by_code(source_path: str | None = None) -> dict[str, dict]:
    workbook_path = source_path or DEFAULT_WHO_2022_VA_WORKBOOK_PATH
    defaults: dict[str, dict] = {}
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD_Mapped"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        va_code = _normalize_label(payload.get("WHO_2022_VA_code"))
        if not icd_code or not va_code or icd_code in defaults:
            continue
        defaults[icd_code] = {
            "node_code": _slugify(va_code, fallback_prefix=NODE_TYPE_FIELD),
            "source_sheet": "ICD_Mapped",
            "source_row_number": row_number,
            "source_category": _normalize_label(payload.get("category")),
            "match_type": _normalize_label(payload.get("WHO_2022_VA_match_type")),
            "mapping_note": _normalize_label(payload.get("WHO_2022_VA_note")),
        }
    return defaults


def apply_admin_cod_bucket_mapping_metadata(
    *,
    scheme: MasCodBucketScheme,
    mapping: MapIcdCodBucket,
    target_node: MasCodBucketNode,
) -> None:
    """Stamp mapping provenance based on whether the target matches source defaults."""
    default_mapping = None
    if scheme.scheme_code == SCHEME_CODE_WHO_2022_VA:
        source_path = _scheme_reset_source_path(scheme)
        default_mapping = _who_2022_default_mapping_by_code(
            str(source_path) if source_path else None
        ).get(mapping.icd_code)

    if default_mapping and target_node.node_code == default_mapping["node_code"]:
        mapping.source_sheet = default_mapping["source_sheet"]
        mapping.source_row_number = default_mapping["source_row_number"]
        mapping.source_category = default_mapping["source_category"]
        mapping.match_type = default_mapping["match_type"]
        mapping.mapping_note = default_mapping["mapping_note"]
        return

    mapping.source_sheet = MANUAL_OVERRIDE_SOURCE_SHEET
    mapping.source_row_number = None
    mapping.source_category = None
    mapping.match_type = MANUAL_OVERRIDE_MATCH_TYPE
    mapping.mapping_note = MANUAL_OVERRIDE_NOTE


def _get_or_create_scheme(
    *,
    scheme_code: str,
    scheme_name: str,
    scheme_description: str,
    source_path: str,
) -> MasCodBucketScheme:
    scheme = db.session.scalar(
        sa.select(MasCodBucketScheme).where(MasCodBucketScheme.scheme_code == scheme_code)
    )
    if scheme is None:
        scheme = MasCodBucketScheme(
            scheme_code=scheme_code,
            scheme_name=scheme_name,
            scheme_description=scheme_description,
            source_path=source_path,
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        return scheme

    scheme.scheme_name = scheme_name
    scheme.scheme_description = scheme_description
    scheme.source_path = source_path
    scheme.mapping_version = (scheme.mapping_version or 0) + 1
    scheme.is_active = True
    db.session.flush()
    return scheme


def _replace_scheme_contents(scheme: MasCodBucketScheme) -> None:
    db.session.execute(
        sa.delete(MapIcdCodBucket).where(MapIcdCodBucket.scheme_id == scheme.scheme_id)
    )
    db.session.execute(
        sa.delete(MasCodBucketNode).where(MasCodBucketNode.scheme_id == scheme.scheme_id)
    )
    db.session.execute(
        sa.delete(MasCodBucketSchemeAgeBand).where(
            MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id
        )
    )
    db.session.flush()


def _create_age_band(
    *,
    scheme: MasCodBucketScheme,
    age_scope: str | None,
    age_label: str,
    min_age_value: int | None,
    min_age_unit: str | None,
    max_age_value: int | None,
    max_age_unit: str | None,
    level_count: int,
    sort_order: int,
) -> MasCodBucketSchemeAgeBand:
    age_band = MasCodBucketSchemeAgeBand(
        scheme_id=scheme.scheme_id,
        age_scope=age_scope,
        age_label=age_label,
        min_age_value=min_age_value,
        min_age_unit=min_age_unit,
        max_age_value=max_age_value,
        max_age_unit=max_age_unit,
        level_count=level_count,
        sort_order=sort_order,
        is_active=True,
    )
    db.session.add(age_band)
    db.session.flush()
    return age_band


def _create_node(
    *,
    scheme: MasCodBucketScheme,
    age_scope: str | None,
    node_type: str,
    node_label: str,
    sort_order: int,
    parent: MasCodBucketNode | None = None,
    node_code_suffix: str | None = None,
) -> MasCodBucketNode:
    code_source = node_code_suffix or node_label
    node = MasCodBucketNode(
        scheme_id=scheme.scheme_id,
        age_scope=age_scope,
        node_type=node_type,
        parent=parent,
        node_code=_slugify(code_source, fallback_prefix=node_type),
        node_label=node_label,
        sort_order=sort_order,
        is_active=True,
    )
    db.session.add(node)
    db.session.flush()
    return node


def _builtin_age_band_metadata(scheme_code: str, age_scope: str | None) -> dict:
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_ADULT_OVER5Y:
        return {
            "age_label": "Adult / Over 5 Years",
            "min_age_value": 5,
            "min_age_unit": AGE_UNIT_YEARS,
            "max_age_value": DEFAULT_MAX_AGE_VALUE,
            "max_age_unit": DEFAULT_MAX_AGE_UNIT,
            "level_count": 3,
            "sort_order": 1,
        }
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_CHILD_1_59M:
        return {
            "age_label": "Child / 1-59 Months",
            "min_age_value": 1,
            "min_age_unit": AGE_UNIT_MONTHS,
            "max_age_value": 60,
            "max_age_unit": AGE_UNIT_MONTHS,
            "level_count": 3,
            "sort_order": 2,
        }
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_NEONATE:
        return {
            "age_label": "Neonate",
            "min_age_value": 0,
            "min_age_unit": AGE_UNIT_DAYS,
            "max_age_value": 29,
            "max_age_unit": AGE_UNIT_DAYS,
            "level_count": 3,
            "sort_order": 3,
        }
    if scheme_code == SCHEME_CODE_CMEA10 and age_scope is None:
        return {
            "age_label": "All Ages",
            "min_age_value": DEFAULT_MIN_AGE_VALUE,
            "min_age_unit": DEFAULT_MIN_AGE_UNIT,
            "max_age_value": DEFAULT_MAX_AGE_VALUE,
            "max_age_unit": DEFAULT_MAX_AGE_UNIT,
            "level_count": 1,
            "sort_order": 1,
        }
    if scheme_code == SCHEME_CODE_WHO_2022_VA and age_scope is None:
        return {
            "age_label": "All Ages",
            "min_age_value": DEFAULT_MIN_AGE_VALUE,
            "min_age_unit": DEFAULT_MIN_AGE_UNIT,
            "max_age_value": DEFAULT_MAX_AGE_VALUE,
            "max_age_unit": DEFAULT_MAX_AGE_UNIT,
            "level_count": 2,
            "sort_order": 1,
        }
    raise ValueError(f"Unsupported built-in scheme/age scope: {scheme_code} / {age_scope}")


def _load_srs_rows_by_scope(workbook_path: str | Path) -> dict[str, list[dict]]:
    rows_by_scope: dict[str, list[dict]] = {
        AGE_SCOPE_ADULT_OVER5Y: [],
        AGE_SCOPE_CHILD_1_59M: [],
        AGE_SCOPE_NEONATE: [],
    }
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD_Mapped"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        if not icd_code:
            continue
        for (
            age_scope,
            main_group_key,
            sub_group_key,
            disease_key,
            match_type_key,
            note_key,
        ) in _SRS_SCOPE_CONFIG:
            disease = _normalize_label(payload.get(disease_key))
            if not disease:
                continue
            rows_by_scope[age_scope].append(
                {
                    "row_number": row_number,
                    "icd_code": icd_code,
                    "source_category": _normalize_label(payload.get("category")),
                    "main_group": _normalize_label(payload.get(main_group_key)),
                    "sub_group": _normalize_label(payload.get(sub_group_key)),
                    "disease": disease,
                    "match_type": _normalize_label(payload.get(match_type_key)),
                    "mapping_note": _normalize_label(payload.get(note_key)),
                }
            )
    return rows_by_scope


def _populate_srs_scope(
    *,
    scheme: MasCodBucketScheme,
    age_scope: str,
    scope_rows: list[dict],
) -> None:
    category_nodes: dict[str, MasCodBucketNode] = {}
    subcategory_nodes: dict[tuple[str, str], MasCodBucketNode] = {}
    field_nodes: dict[tuple[str, str | None, str], MasCodBucketNode] = {}
    category_order = 0
    subcategory_order = 0
    field_order = 0

    for row in scope_rows:
        category_label = row["main_group"] or "Unspecified"
        subcategory_label = row["sub_group"]
        field_label = row["disease"]

        category_node = category_nodes.get(category_label)
        if category_node is None:
            category_order += 1
            category_node = _create_node(
                scheme=scheme,
                age_scope=age_scope,
                node_type=NODE_TYPE_CATEGORY,
                node_label=category_label,
                sort_order=category_order,
            )
            category_nodes[category_label] = category_node

        parent_node = category_node
        if subcategory_label:
            subcategory_key = (category_label, subcategory_label)
            subcategory_node = subcategory_nodes.get(subcategory_key)
            if subcategory_node is None:
                subcategory_order += 1
                subcategory_node = _create_node(
                    scheme=scheme,
                    age_scope=age_scope,
                    node_type=NODE_TYPE_SUBCATEGORY,
                    node_label=subcategory_label,
                    sort_order=subcategory_order,
                    parent=category_node,
                )
                subcategory_nodes[subcategory_key] = subcategory_node
            parent_node = subcategory_node

        field_key = (category_label, subcategory_label, field_label)
        field_node = field_nodes.get(field_key)
        if field_node is None:
            field_order += 1
            field_node = _create_node(
                scheme=scheme,
                age_scope=age_scope,
                node_type=NODE_TYPE_FIELD,
                node_label=field_label,
                sort_order=field_order,
                parent=parent_node,
            )
            field_nodes[field_key] = field_node

        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope=age_scope,
                icd_code=row["icd_code"],
                node_id=field_node.node_id,
                source_sheet="ICD_Mapped",
                source_row_number=row["row_number"],
                source_category=row["source_category"],
                match_type=row["match_type"],
                mapping_note=row["mapping_note"],
                is_active=True,
            )
        )


def _populate_cmea10_scheme(
    *,
    scheme: MasCodBucketScheme,
    workbook_path: str | Path,
) -> None:
    field_nodes: dict[str, MasCodBucketNode] = {}
    field_order = 0
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD10_CMEA10"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        field_label = _normalize_label(payload.get("CMEA10"))
        if not icd_code or not field_label:
            continue
        field_node = field_nodes.get(field_label)
        if field_node is None:
            field_order += 1
            field_node = _create_node(
                scheme=scheme,
                age_scope=None,
                node_type=NODE_TYPE_FIELD,
                node_label=field_label,
                sort_order=field_order,
            )
            field_nodes[field_label] = field_node

        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope=None,
                icd_code=icd_code,
                node_id=field_node.node_id,
                source_sheet="ICD10_CMEA10",
                source_row_number=row_number,
                source_category=_normalize_label(payload.get("category")),
                is_active=True,
            )
        )


def _populate_who_2022_va_scheme(
    *,
    scheme: MasCodBucketScheme,
    workbook_path: str | Path,
) -> None:
    category_nodes: dict[str, MasCodBucketNode] = {}
    field_nodes: dict[tuple[str, str], MasCodBucketNode] = {}
    category_order = 0
    field_order = 0

    for row_number, payload in _load_sheet_rows(workbook_path, "ICD_Mapped"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        section_label = _normalize_label(payload.get("WHO_2022_VA_section"))
        va_code = _normalize_label(payload.get("WHO_2022_VA_code"))
        va_title = _normalize_label(payload.get("WHO_2022_VA_cause"))
        if not icd_code or not section_label or not va_code or not va_title:
            continue

        category_node = category_nodes.get(section_label)
        if category_node is None:
            category_order += 1
            category_node = _create_node(
                scheme=scheme,
                age_scope=None,
                node_type=NODE_TYPE_CATEGORY,
                node_label=section_label,
                sort_order=category_order,
            )
            category_nodes[section_label] = category_node

        field_key = (section_label, va_code)
        field_node = field_nodes.get(field_key)
        if field_node is None:
            field_order += 1
            field_node = _create_node(
                scheme=scheme,
                age_scope=None,
                node_type=NODE_TYPE_FIELD,
                node_label=va_title,
                sort_order=field_order,
                parent=category_node,
                node_code_suffix=va_code,
            )
            field_nodes[field_key] = field_node

        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope=None,
                icd_code=icd_code,
                node_id=field_node.node_id,
                source_sheet="ICD_Mapped",
                source_row_number=row_number,
                source_category=_normalize_label(payload.get("category")),
                match_type=_normalize_label(payload.get("WHO_2022_VA_match_type")),
                mapping_note=_normalize_label(payload.get("WHO_2022_VA_note")),
                is_active=True,
            )
        )


def _replace_scheme_scope_contents(scheme: MasCodBucketScheme, age_scope: str | None) -> None:
    db.session.execute(
        sa.delete(MapIcdCodBucket).where(
            MapIcdCodBucket.scheme_id == scheme.scheme_id,
            _scope_clause(MapIcdCodBucket.age_scope, age_scope),
        )
    )
    db.session.execute(
        sa.delete(MasCodBucketNode).where(
            MasCodBucketNode.scheme_id == scheme.scheme_id,
            _scope_clause(MasCodBucketNode.age_scope, age_scope),
        )
    )
    db.session.execute(
        sa.delete(MasCodBucketSchemeAgeBand).where(
            MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id,
            _scope_clause(MasCodBucketSchemeAgeBand.age_scope, age_scope),
        )
    )
    db.session.flush()


def import_srs_india_scheme(workbook_path: str | Path = DEFAULT_SRS_WORKBOOK_PATH) -> MasCodBucketScheme:
    """Replace the SRS India bucket scheme from the workbook source."""
    workbook_path = str(workbook_path)
    scheme = _get_or_create_scheme(
        scheme_code=SCHEME_CODE_SRS_INDIA,
        scheme_name="SRS India",
        scheme_description=(
            "Age-scoped SRS India cause-of-death reporting hierarchy imported from "
            "the ICD_Mapped workbook."
        ),
        source_path=workbook_path,
    )
    _replace_scheme_contents(scheme)
    rows_by_scope = _load_srs_rows_by_scope(workbook_path)
    for age_scope in (
        AGE_SCOPE_ADULT_OVER5Y,
        AGE_SCOPE_CHILD_1_59M,
        AGE_SCOPE_NEONATE,
    ):
        meta = _builtin_age_band_metadata(SCHEME_CODE_SRS_INDIA, age_scope)
        _create_age_band(
            scheme=scheme,
            age_scope=age_scope,
            age_label=meta["age_label"],
            min_age_value=meta["min_age_value"],
            min_age_unit=meta["min_age_unit"],
            max_age_value=meta["max_age_value"],
            max_age_unit=meta["max_age_unit"],
            level_count=meta["level_count"],
            sort_order=meta["sort_order"],
        )
        _populate_srs_scope(
            scheme=scheme,
            age_scope=age_scope,
            scope_rows=rows_by_scope[age_scope],
        )

    db.session.commit()
    return scheme


def import_cmea10_scheme(workbook_path: str | Path = DEFAULT_CMEA10_WORKBOOK_PATH) -> MasCodBucketScheme:
    """Replace the flat CMEA10 bucket scheme from the workbook source."""
    workbook_path = str(workbook_path)
    scheme = _get_or_create_scheme(
        scheme_code=SCHEME_CODE_CMEA10,
        scheme_name="CMEA10",
        scheme_description="Flat ICD-10 to CMEA10 cause bucket mapping.",
        source_path=workbook_path,
    )
    _replace_scheme_contents(scheme)
    meta = _builtin_age_band_metadata(SCHEME_CODE_CMEA10, None)
    _create_age_band(
        scheme=scheme,
        age_scope=None,
        age_label=meta["age_label"],
        min_age_value=meta["min_age_value"],
        min_age_unit=meta["min_age_unit"],
        max_age_value=meta["max_age_value"],
        max_age_unit=meta["max_age_unit"],
        level_count=meta["level_count"],
        sort_order=meta["sort_order"],
    )
    _populate_cmea10_scheme(scheme=scheme, workbook_path=workbook_path)

    db.session.commit()
    return scheme


def import_who_2022_va_scheme(
    workbook_path: str | Path = DEFAULT_WHO_2022_VA_WORKBOOK_PATH,
) -> MasCodBucketScheme:
    """Replace the WHO 2022 VA bucket scheme from the generated workbook source."""
    workbook_path = str(workbook_path)
    scheme = _get_or_create_scheme(
        scheme_code=SCHEME_CODE_WHO_2022_VA,
        scheme_name="WHO 2022 VA",
        scheme_description=(
            "WHO 2022 verbal autopsy cause-of-death bucket mapping imported from "
            "the generated ICD_Mapped workbook."
        ),
        source_path=workbook_path,
    )
    _replace_scheme_contents(scheme)
    meta = _builtin_age_band_metadata(SCHEME_CODE_WHO_2022_VA, None)
    _create_age_band(
        scheme=scheme,
        age_scope=None,
        age_label=meta["age_label"],
        min_age_value=meta["min_age_value"],
        min_age_unit=meta["min_age_unit"],
        max_age_value=meta["max_age_value"],
        max_age_unit=meta["max_age_unit"],
        level_count=meta["level_count"],
        sort_order=meta["sort_order"],
    )
    _populate_who_2022_va_scheme(scheme=scheme, workbook_path=workbook_path)

    db.session.commit()
    return scheme


def list_cod_bucket_schemes() -> list[MasCodBucketScheme]:
    return list(
        db.session.scalars(
            sa.select(MasCodBucketScheme).order_by(MasCodBucketScheme.scheme_code.asc())
        )
    )


def list_cod_bucket_scheme_cards() -> list[dict]:
    schemes = list_cod_bucket_schemes()
    if not schemes:
        return []

    age_bands = list(
        db.session.scalars(
            sa.select(MasCodBucketSchemeAgeBand)
            .where(
                MasCodBucketSchemeAgeBand.scheme_id.in_(
                    [scheme.scheme_id for scheme in schemes]
                )
            )
            .order_by(
                MasCodBucketSchemeAgeBand.scheme_id.asc(),
                MasCodBucketSchemeAgeBand.sort_order.asc(),
                MasCodBucketSchemeAgeBand.age_label.asc(),
            )
        )
    )
    age_bands_by_scheme: dict[uuid.UUID, list[MasCodBucketSchemeAgeBand]] = {}
    for age_band in age_bands:
        age_bands_by_scheme.setdefault(age_band.scheme_id, []).append(age_band)

    warnings_by_scheme = {
        scheme_id: _age_scope_warning_messages(rows)
        for scheme_id, rows in age_bands_by_scheme.items()
    }
    return [
        {
            "scheme_code": scheme.scheme_code,
            "scheme_name": scheme.scheme_name,
            "mapping_version": scheme.mapping_version,
            "is_active": scheme.is_active,
            "can_reset_from_source": scheme_can_reset_from_source(scheme),
            "age_bands": [
                {
                    **_serialize_age_band(age_band),
                    "can_reset_from_source": _age_band_can_reset_from_source(
                        scheme, age_band
                    ),
                    "warnings": warnings_by_scheme.get(scheme.scheme_id, {}).get(
                        age_band.age_scope or "",
                        [],
                    ),
                }
                for age_band in age_bands_by_scheme.get(scheme.scheme_id, [])
            ],
        }
        for scheme in schemes
    ]


def reset_cod_bucket_scheme_age_band_to_source(
    *,
    scheme_code: str,
    age_scope: str | None,
    reset_entire_scheme: bool = False,
) -> MasCodBucketScheme:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")
    if not scheme_can_reset_from_source(scheme):
        raise ValueError("This scheme cannot be reset from source.")

    workbook_path = _scheme_reset_source_path(scheme)
    if workbook_path is None:
        raise ValueError("No source workbook is configured for this scheme.")

    if reset_entire_scheme:
        if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
            return import_srs_india_scheme(workbook_path)
        if scheme.scheme_code == SCHEME_CODE_CMEA10:
            return import_cmea10_scheme(workbook_path)
        if scheme.scheme_code == SCHEME_CODE_WHO_2022_VA:
            return import_who_2022_va_scheme(workbook_path)
        raise ValueError("This scheme does not support reset from source.")

    if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
        if age_scope not in {
            AGE_SCOPE_ADULT_OVER5Y,
            AGE_SCOPE_CHILD_1_59M,
            AGE_SCOPE_NEONATE,
        }:
            raise ValueError("A valid SRS India age band is required for reset.")
        rows_by_scope = _load_srs_rows_by_scope(workbook_path)
        meta = _builtin_age_band_metadata(SCHEME_CODE_SRS_INDIA, age_scope)
        _replace_scheme_scope_contents(scheme, age_scope)
        _create_age_band(
            scheme=scheme,
            age_scope=age_scope,
            age_label=meta["age_label"],
            min_age_value=meta["min_age_value"],
            min_age_unit=meta["min_age_unit"],
            max_age_value=meta["max_age_value"],
            max_age_unit=meta["max_age_unit"],
            level_count=meta["level_count"],
            sort_order=meta["sort_order"],
        )
        _populate_srs_scope(
            scheme=scheme,
            age_scope=age_scope,
            scope_rows=rows_by_scope[age_scope],
        )
    elif scheme.scheme_code == SCHEME_CODE_CMEA10:
        _replace_scheme_contents(scheme)
        meta = _builtin_age_band_metadata(SCHEME_CODE_CMEA10, None)
        _create_age_band(
            scheme=scheme,
            age_scope=None,
            age_label=meta["age_label"],
            min_age_value=meta["min_age_value"],
            min_age_unit=meta["min_age_unit"],
            max_age_value=meta["max_age_value"],
            max_age_unit=meta["max_age_unit"],
            level_count=meta["level_count"],
            sort_order=meta["sort_order"],
        )
        _populate_cmea10_scheme(scheme=scheme, workbook_path=workbook_path)
    elif scheme.scheme_code == SCHEME_CODE_WHO_2022_VA:
        _replace_scheme_contents(scheme)
        meta = _builtin_age_band_metadata(SCHEME_CODE_WHO_2022_VA, None)
        _create_age_band(
            scheme=scheme,
            age_scope=None,
            age_label=meta["age_label"],
            min_age_value=meta["min_age_value"],
            min_age_unit=meta["min_age_unit"],
            max_age_value=meta["max_age_value"],
            max_age_unit=meta["max_age_unit"],
            level_count=meta["level_count"],
            sort_order=meta["sort_order"],
        )
        _populate_who_2022_va_scheme(scheme=scheme, workbook_path=workbook_path)
    else:
        raise ValueError("This scheme does not support reset from source.")

    scheme.mapping_version = (scheme.mapping_version or 0) + 1
    scheme.source_path = str(workbook_path)
    scheme.is_active = True
    db.session.commit()
    return scheme


def age_scope_label(age_scope: str | None) -> str:
    if age_scope == AGE_SCOPE_ADULT_OVER5Y:
        return "Adult / Over 5 Years"
    if age_scope == AGE_SCOPE_CHILD_1_59M:
        return "Child / 1–59 Months"
    if age_scope == AGE_SCOPE_NEONATE:
        return "Neonate"
    return "All Ages"


def _normalize_scheme_age_bands(
    *,
    age_bands: list[dict],
    scheme: MasCodBucketScheme | None = None,
) -> tuple[list[dict], list[str], dict[uuid.UUID, MasCodBucketSchemeAgeBand]]:
    if not age_bands:
        raise ValueError("At least one age band is required.")

    existing_by_id: dict[uuid.UUID, MasCodBucketSchemeAgeBand] = {}
    if scheme is not None:
        existing_rows = list(
            db.session.scalars(
                sa.select(MasCodBucketSchemeAgeBand).where(
                    MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id
                )
            )
        )
        existing_by_id = {row.age_band_id: row for row in existing_rows}

    cleaned_rows: list[dict] = []
    warnings: list[str] = []
    seen_scopes: set[str] = set()

    for index, raw_band in enumerate(age_bands, start=1):
        age_label = _normalize_label(raw_band.get("age_label"))
        if not age_label:
            raise ValueError(f"Age band {index} requires a name.")

        raw_age_band_id = raw_band.get("age_band_id")
        age_band_id = None
        existing_age_band = None
        if raw_age_band_id not in {None, ""}:
            try:
                age_band_id = uuid.UUID(str(raw_age_band_id))
            except (TypeError, ValueError):
                raise ValueError(f"Age band '{age_label}' has an invalid age band id.")
            existing_age_band = existing_by_id.get(age_band_id)
            if existing_age_band is None:
                raise ValueError(f"Age band '{age_label}' does not belong to this scheme.")

        age_scope = (
            existing_age_band.age_scope
            if existing_age_band is not None
            else _slugify(age_label, fallback_prefix="age")
        )
        if age_scope in seen_scopes:
            raise ValueError(f"Age band '{age_label}' duplicates another band name.")
        seen_scopes.add(age_scope)

        try:
            level_count = int(raw_band.get("level_count"))
        except (TypeError, ValueError):
            raise ValueError(f"Age band '{age_label}' requires a numeric level count.")
        if level_count not in {1, 2, 3}:
            raise ValueError(f"Age band '{age_label}' must use 1, 2, or 3 levels.")

        min_age_value = raw_band.get("min_age_value")
        if min_age_value in {None, ""}:
            raise ValueError(f"Age band '{age_label}' minimum age is required.")
        try:
            min_age_value = int(min_age_value)
        except (TypeError, ValueError):
            raise ValueError(f"Age band '{age_label}' minimum age must be an integer.")
        min_age_unit = (raw_band.get("min_age_unit") or "").strip().lower()
        max_age_value = raw_band.get("max_age_value")
        if max_age_value in {None, ""}:
            raise ValueError(f"Age band '{age_label}' maximum age is required.")
        try:
            max_age_value = int(max_age_value)
        except (TypeError, ValueError):
            raise ValueError(f"Age band '{age_label}' maximum age must be an integer.")
        max_age_unit = (raw_band.get("max_age_unit") or "").strip().lower()

        if min_age_value < 0:
            raise ValueError(f"Age band '{age_label}' minimum age cannot be negative.")
        if max_age_value < 0:
            raise ValueError(f"Age band '{age_label}' maximum age cannot be negative.")
        if min_age_unit not in AGE_UNITS:
            raise ValueError(f"Age band '{age_label}' minimum age unit is invalid.")
        if max_age_unit not in AGE_UNITS:
            raise ValueError(f"Age band '{age_label}' maximum age unit is invalid.")

        min_days = _approx_age_days(min_age_value, min_age_unit)
        max_days = _approx_age_days(max_age_value, max_age_unit)
        cleaned_rows.append(
            {
                "age_band_id": age_band_id,
                "age_scope": age_scope,
                "age_label": age_label,
                "min_age_value": min_age_value,
                "min_age_unit": min_age_unit,
                "max_age_value": max_age_value,
                "max_age_unit": max_age_unit,
                "level_count": level_count,
                "min_days": min_days,
                "max_days": max_days,
                "input_index": index,
            }
        )

    cleaned_rows.sort(
        key=lambda row: (
            row["min_days"],
            row["input_index"],
            row["age_label"].lower(),
        )
    )
    previous_band = None
    previous_max_days = None
    for index, row in enumerate(cleaned_rows, start=1):
        row["sort_order"] = index
        min_days = row["min_days"]
        max_days = row["max_days"]
        if min_days is not None and max_days is not None and min_days >= max_days:
            warnings.append(
                f"{row['age_label']}: upper bound must be greater than lower bound because bands use >= lower and < upper."
            )
        if previous_band is not None and previous_max_days is not None and min_days is not None:
            if min_days < previous_max_days:
                warnings.append(
                    f"{row['age_label']}: overlaps with {previous_band} using approximate unit conversion."
                )
            elif min_days > previous_max_days:
                warnings.append(
                    f"{row['age_label']}: gap after {previous_band} using approximate unit conversion."
                )
        previous_band = row["age_label"]
        previous_max_days = max_days

    return cleaned_rows, warnings, existing_by_id


def _max_node_depth_for_scope(*, scheme_id: uuid.UUID, age_scope: str | None) -> int:
    nodes = list(
        db.session.scalars(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme_id,
                MasCodBucketNode.age_scope == age_scope,
            )
        )
    )
    if not nodes:
        return 0

    parent_by_id = {node.node_id: node.parent_node_id for node in nodes}
    max_depth = 1
    for node_id in parent_by_id:
        depth = 1
        parent_id = parent_by_id[node_id]
        while parent_id is not None:
            depth += 1
            parent_id = parent_by_id.get(parent_id)
        max_depth = max(max_depth, depth)
    return max_depth


def create_cod_bucket_scheme(
    *,
    scheme_name: str,
    scheme_code: str,
    age_bands: list[dict],
) -> tuple[MasCodBucketScheme, list[str]]:
    normalized_name = _normalize_label(scheme_name)
    normalized_code = _normalize_icd_code(scheme_code)
    if not normalized_name:
        raise ValueError("Scheme name is required.")
    if not normalized_code:
        raise ValueError("Scheme code is required.")
    if len(normalized_code) > 32 or not re.fullmatch(r"[A-Z0-9_]+", normalized_code):
        raise ValueError("Scheme code must use only A-Z, 0-9, and underscore.")
    if db.session.scalar(
        sa.select(MasCodBucketScheme.scheme_id).where(
            MasCodBucketScheme.scheme_code == normalized_code
        )
    ):
        raise ValueError("Scheme code already exists.")
    cleaned_rows, warnings, _existing_by_id = _normalize_scheme_age_bands(
        age_bands=age_bands,
    )

    scheme = MasCodBucketScheme(
        scheme_code=normalized_code,
        scheme_name=normalized_name,
        scheme_description=f"{normalized_name} COD bucket scheme",
        mapping_version=1,
        is_active=True,
    )
    db.session.add(scheme)
    db.session.flush()

    for row in cleaned_rows:
        _create_age_band(
            scheme=scheme,
            age_scope=row["age_scope"],
            age_label=row["age_label"],
            min_age_value=row["min_age_value"],
            min_age_unit=row["min_age_unit"],
            max_age_value=row["max_age_value"],
            max_age_unit=row["max_age_unit"],
            level_count=row["level_count"],
            sort_order=row["sort_order"],
        )

    db.session.commit()
    return scheme, warnings


def update_cod_bucket_scheme(
    *,
    scheme_code: str,
    scheme_name: str,
    age_bands: list[dict],
) -> tuple[MasCodBucketScheme, list[str]]:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    normalized_name = _normalize_label(scheme_name)
    if not normalized_name:
        raise ValueError("Scheme name is required.")

    cleaned_rows, warnings, existing_by_id = _normalize_scheme_age_bands(
        age_bands=age_bands,
        scheme=scheme,
    )

    retained_band_ids = {
        row["age_band_id"]
        for row in cleaned_rows
        if row["age_band_id"] is not None
    }
    existing_age_bands = list(existing_by_id.values())

    for existing_age_band in existing_age_bands:
        if existing_age_band.age_band_id in retained_band_ids:
            continue
        has_nodes = db.session.scalar(
            sa.select(sa.func.count())
            .select_from(MasCodBucketNode)
            .where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == existing_age_band.age_scope,
            )
        )
        has_mappings = db.session.scalar(
            sa.select(sa.func.count())
            .select_from(MapIcdCodBucket)
            .where(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.age_scope == existing_age_band.age_scope,
            )
        )
        if has_nodes or has_mappings:
            raise ValueError(
                f"Cannot remove age band '{existing_age_band.age_label}' while headings or ICD mappings still exist."
            )
        db.session.delete(existing_age_band)

    scheme.scheme_name = normalized_name
    scheme.scheme_description = f"{normalized_name} COD bucket scheme"
    scheme.mapping_version = (scheme.mapping_version or 0) + 1

    for row in cleaned_rows:
        current_depth = _max_node_depth_for_scope(
            scheme_id=scheme.scheme_id,
            age_scope=row["age_scope"],
        )
        if current_depth > row["level_count"]:
            raise ValueError(
                f"{row['age_label']}: level count cannot be reduced below existing hierarchy depth {current_depth}."
            )
        if row["age_band_id"] is not None:
            age_band = existing_by_id[row["age_band_id"]]
            age_band.age_label = row["age_label"]
            age_band.min_age_value = row["min_age_value"]
            age_band.min_age_unit = row["min_age_unit"]
            age_band.max_age_value = row["max_age_value"]
            age_band.max_age_unit = row["max_age_unit"]
            age_band.level_count = row["level_count"]
            age_band.sort_order = row["sort_order"]
            age_band.is_active = True
        else:
            _create_age_band(
                scheme=scheme,
                age_scope=row["age_scope"],
                age_label=row["age_label"],
                min_age_value=row["min_age_value"],
                min_age_unit=row["min_age_unit"],
                max_age_value=row["max_age_value"],
                max_age_unit=row["max_age_unit"],
                level_count=row["level_count"],
                sort_order=row["sort_order"],
            )

    db.session.commit()
    return scheme, warnings


def get_cod_bucket_scheme(scheme_code: str) -> MasCodBucketScheme | None:
    return db.session.scalar(
        sa.select(MasCodBucketScheme).where(MasCodBucketScheme.scheme_code == scheme_code)
    )


def _node_path_label(node: MasCodBucketNode) -> str:
    labels = [node.node_label]
    parent = node.parent
    while parent is not None:
        labels.append(parent.node_label)
        parent = parent.parent
    return " > ".join(reversed(labels))


def _node_hierarchy_sort_key(node: MasCodBucketNode) -> tuple:
    lineage = []
    current = node
    while current is not None:
        lineage.append(current)
        current = current.parent
    lineage.reverse()
    sort_orders = tuple(part.sort_order for part in lineage)
    labels = tuple(part.node_label.lower() for part in lineage)
    return (
        sort_orders,
        labels,
        len(lineage),
    )


def get_cod_bucket_scheme_editor_payload(
    *,
    scheme_code: str,
    age_scope: str | None = None,
) -> dict:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    age_bands = list(
        db.session.scalars(
            sa.select(MasCodBucketSchemeAgeBand)
            .where(MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id)
            .order_by(
                MasCodBucketSchemeAgeBand.sort_order.asc(),
                MasCodBucketSchemeAgeBand.age_label.asc(),
            )
        )
    )
    available_age_scopes = [age_band.age_scope for age_band in age_bands]
    selected_age_scope = (
        age_scope
        if age_scope in available_age_scopes
        else (available_age_scopes[0] if available_age_scopes else None)
    )
    selected_age_band = next(
        (age_band for age_band in age_bands if age_band.age_scope == selected_age_scope),
        None,
    )

    nodes = list(
        db.session.scalars(
            sa.select(MasCodBucketNode)
            .where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == selected_age_scope,
            )
        )
    )
    nodes.sort(key=_node_hierarchy_sort_key)
    warnings_by_scope = _age_scope_warning_messages(age_bands)

    return {
        "scheme": {
            "scheme_code": scheme.scheme_code,
            "scheme_name": scheme.scheme_name,
            "mapping_version": scheme.mapping_version,
            "is_active": scheme.is_active,
        },
        "age_bands": [
            {
                **_serialize_age_band(age_band),
                "can_reset_from_source": _age_band_can_reset_from_source(
                    scheme, age_band
                ),
                "warnings": warnings_by_scope.get(age_band.age_scope or "", []),
            }
            for age_band in age_bands
        ],
        "selected_age_scope": selected_age_scope,
        "selected_age_band": (
            {
                **_serialize_age_band(selected_age_band),
                "can_reset_from_source": _age_band_can_reset_from_source(
                    scheme, selected_age_band
                ),
                "warnings": warnings_by_scope.get(selected_age_band.age_scope or "", []),
            }
            if selected_age_band is not None
            else None
        ),
        "nodes": [
            {
                "node_id": str(node.node_id),
                "age_scope": node.age_scope,
                "node_type": node.node_type,
                "node_label": node.node_label,
                "sort_order": node.sort_order,
                "parent_node_id": str(node.parent_node_id) if node.parent_node_id else None,
                "path_label": _node_path_label(node),
            }
            for node in nodes
        ],
    }


def export_cod_bucket_scheme_json(*, scheme_code: str) -> dict:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    age_bands = list(
        db.session.scalars(
            sa.select(MasCodBucketSchemeAgeBand)
            .where(MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id)
            .order_by(
                MasCodBucketSchemeAgeBand.sort_order.asc(),
                MasCodBucketSchemeAgeBand.age_label.asc(),
            )
        )
    )
    warnings_by_scope = _age_scope_warning_messages(age_bands)

    nodes = list(
        db.session.scalars(
            sa.select(MasCodBucketNode).where(MasCodBucketNode.scheme_id == scheme.scheme_id)
        )
    )
    nodes.sort(key=_node_hierarchy_sort_key)
    nodes_by_id = {node.node_id: node for node in nodes}

    def _path_label_for_node(node: MasCodBucketNode) -> str:
        labels = [node.node_label]
        parent = nodes_by_id.get(node.parent_node_id)
        while parent is not None:
            labels.append(parent.node_label)
            parent = nodes_by_id.get(parent.parent_node_id)
        return " > ".join(reversed(labels))

    icd_master_sq = _icd_master_display_subquery()
    mapping_rows = db.session.execute(
        sa.select(
            MapIcdCodBucket.mapping_id,
            MapIcdCodBucket.age_scope,
            MapIcdCodBucket.icd_code,
            MapIcdCodBucket.node_id,
            MapIcdCodBucket.source_sheet,
            MapIcdCodBucket.source_row_number,
            MapIcdCodBucket.source_category,
            MapIcdCodBucket.match_type,
            MapIcdCodBucket.mapping_note,
            MapIcdCodBucket.is_active,
            icd_master_sq.c.icd_to_display,
        )
        .select_from(MapIcdCodBucket)
        .outerjoin(
            icd_master_sq,
            icd_master_sq.c.icd_code == MapIcdCodBucket.icd_code,
        )
        .where(MapIcdCodBucket.scheme_id == scheme.scheme_id)
        .order_by(
            sa.func.coalesce(MapIcdCodBucket.age_scope, "").asc(),
            MapIcdCodBucket.icd_code.asc(),
        )
    ).mappings().all()

    return {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "scheme": {
            "scheme_id": str(scheme.scheme_id),
            "scheme_code": scheme.scheme_code,
            "scheme_name": scheme.scheme_name,
            "scheme_description": scheme.scheme_description,
            "mapping_version": scheme.mapping_version,
            "is_active": scheme.is_active,
            "can_reset_from_source": scheme_can_reset_from_source(scheme),
        },
        "age_bands": [
            {
                **_serialize_age_band(age_band),
                "is_active": age_band.is_active,
                "can_reset_from_source": _age_band_can_reset_from_source(
                    scheme, age_band
                ),
                "warnings": warnings_by_scope.get(age_band.age_scope or "", []),
            }
            for age_band in age_bands
        ],
        "nodes": [
            {
                "node_id": str(node.node_id),
                "age_scope": node.age_scope,
                "node_type": node.node_type,
                "node_code": node.node_code,
                "node_label": node.node_label,
                "sort_order": node.sort_order,
                "is_active": node.is_active,
                "parent_node_id": str(node.parent_node_id) if node.parent_node_id else None,
                "path_label": _path_label_for_node(node),
            }
            for node in nodes
        ],
        "mappings": [
            {
                "mapping_id": str(row["mapping_id"]),
                "age_scope": row["age_scope"],
                "icd_code": row["icd_code"],
                "icd_to_display": row["icd_to_display"] or row["icd_code"],
                "node_id": str(row["node_id"]),
                "node_path_label": _path_label_for_node(nodes_by_id[row["node_id"]]),
                "source_sheet": row["source_sheet"],
                "source_row_number": row["source_row_number"],
                "source_category": row["source_category"],
                "match_type": row["match_type"],
                "mapping_note": row["mapping_note"],
                "is_active": row["is_active"],
            }
            for row in mapping_rows
        ],
    }


def import_cod_bucket_scheme_json(*, scheme_code: str, payload: dict) -> MasCodBucketScheme:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")
    if not isinstance(payload, dict):
        raise ValueError("COD bucket import payload must be a JSON object.")
    if {
        "items",
        "source_version",
        "assignability_decision_source",
    }.issubset(payload.keys()):
        raise ValueError(
            "This file looks like an ICD policy JSON, not a COD bucket scheme export. "
            "Use Export JSON from the COD bucket scheme card, or use the COD bucket workbook import/reset flow."
        )

    scheme_payload = payload.get("scheme")
    age_band_payload = payload.get("age_bands")
    node_payload = payload.get("nodes")
    mapping_payload = payload.get("mappings")
    if not isinstance(scheme_payload, dict):
        raise ValueError("COD bucket import payload must include a scheme object.")
    if not isinstance(age_band_payload, list) or not age_band_payload:
        raise ValueError("COD bucket import payload must include at least one age band.")
    if not isinstance(node_payload, list):
        raise ValueError("COD bucket import payload must include a nodes list.")
    if not isinstance(mapping_payload, list):
        raise ValueError("COD bucket import payload must include a mappings list.")

    normalized_name = _normalize_label(scheme_payload.get("scheme_name"))
    if not normalized_name:
        raise ValueError("Imported scheme name is required.")
    normalized_description = _normalize_label(scheme_payload.get("scheme_description"))

    cleaned_age_bands: list[dict] = []
    seen_age_scopes: set[str | None] = set()
    for index, raw_age_band in enumerate(age_band_payload, start=1):
        age_label = _normalize_label(raw_age_band.get("label"))
        if not age_label:
            raise ValueError(f"Age band {index} requires a label.")
        raw_age_scope = raw_age_band.get("value")
        age_scope = _normalize_label(raw_age_scope)
        age_scope = age_scope.lower() if age_scope else None
        if age_scope in seen_age_scopes:
            raise ValueError(f"Age band '{age_label}' duplicates an existing age scope.")
        seen_age_scopes.add(age_scope)

        try:
            level_count = int(raw_age_band.get("level_count"))
        except (TypeError, ValueError):
            raise ValueError(f"Age band '{age_label}' requires a numeric level count.")
        if level_count not in {1, 2, 3}:
            raise ValueError(f"Age band '{age_label}' must use 1, 2, or 3 levels.")

        min_age_value = raw_age_band.get("min_age_value")
        max_age_value = raw_age_band.get("max_age_value")
        try:
            min_age_value = int(min_age_value)
            max_age_value = int(max_age_value)
        except (TypeError, ValueError):
            raise ValueError(
                f"Age band '{age_label}' minimum and maximum ages must be integers."
            )
        min_age_unit = (raw_age_band.get("min_age_unit") or "").strip().lower()
        max_age_unit = (raw_age_band.get("max_age_unit") or "").strip().lower()
        if min_age_value < 0 or max_age_value < 0:
            raise ValueError(f"Age band '{age_label}' ages cannot be negative.")
        if min_age_unit not in AGE_UNITS or max_age_unit not in AGE_UNITS:
            raise ValueError(f"Age band '{age_label}' contains an invalid age unit.")

        sort_order_raw = raw_age_band.get("sort_order", index)
        try:
            sort_order = int(sort_order_raw)
        except (TypeError, ValueError):
            raise ValueError(f"Age band '{age_label}' sort order must be an integer.")

        cleaned_age_bands.append(
            {
                "age_scope": age_scope,
                "age_label": age_label,
                "min_age_value": min_age_value,
                "min_age_unit": min_age_unit,
                "max_age_value": max_age_value,
                "max_age_unit": max_age_unit,
                "level_count": level_count,
                "sort_order": sort_order,
            }
        )
    cleaned_age_bands.sort(key=lambda item: (item["sort_order"], item["age_label"].lower()))
    age_scopes = {item["age_scope"] for item in cleaned_age_bands}

    cleaned_nodes: list[dict] = []
    node_ids_by_import_id: dict[str, dict] = {}
    for index, raw_node in enumerate(node_payload, start=1):
        import_node_id = str(raw_node.get("node_id") or "").strip()
        if not import_node_id:
            raise ValueError(f"Node {index} is missing node_id.")
        if import_node_id in node_ids_by_import_id:
            raise ValueError(f"Node '{import_node_id}' appears more than once in the import.")
        age_scope = _normalize_label(raw_node.get("age_scope"))
        age_scope = age_scope.lower() if age_scope else None
        if age_scope not in age_scopes:
            raise ValueError(
                f"Node '{import_node_id}' references unknown age scope '{age_scope or ''}'."
            )
        node_type = (raw_node.get("node_type") or "").strip().lower()
        if node_type not in {NODE_TYPE_CATEGORY, NODE_TYPE_SUBCATEGORY, NODE_TYPE_FIELD}:
            raise ValueError(f"Node '{import_node_id}' has an invalid node type.")
        node_label = _normalize_label(raw_node.get("node_label"))
        if not node_label:
            raise ValueError(f"Node '{import_node_id}' requires a node label.")
        node_code = _normalize_label(raw_node.get("node_code"))
        if not node_code:
            raise ValueError(f"Node '{import_node_id}' requires a node code.")
        try:
            sort_order = int(raw_node.get("sort_order"))
        except (TypeError, ValueError):
            raise ValueError(f"Node '{import_node_id}' sort order must be an integer.")
        parent_node_id = str(raw_node.get("parent_node_id") or "").strip() or None
        cleaned = {
            "import_node_id": import_node_id,
            "age_scope": age_scope,
            "node_type": node_type,
            "node_code": node_code,
            "node_label": node_label,
            "sort_order": sort_order,
            "is_active": bool(raw_node.get("is_active", True)),
            "parent_node_id": parent_node_id,
        }
        cleaned_nodes.append(cleaned)
        node_ids_by_import_id[import_node_id] = cleaned

    type_rank = {NODE_TYPE_CATEGORY: 1, NODE_TYPE_SUBCATEGORY: 2, NODE_TYPE_FIELD: 3}
    for node in cleaned_nodes:
        parent_import_id = node["parent_node_id"]
        if not parent_import_id:
            if node["node_type"] != NODE_TYPE_CATEGORY:
                raise ValueError(
                    f"Node '{node['node_label']}' must be a category when it has no parent."
                )
            continue
        parent = node_ids_by_import_id.get(parent_import_id)
        if parent is None:
            raise ValueError(
                f"Node '{node['node_label']}' references an unknown parent node."
            )
        if parent["age_scope"] != node["age_scope"]:
            raise ValueError(
                f"Node '{node['node_label']}' must stay within a single age scope hierarchy."
            )
        if type_rank[parent["node_type"]] >= type_rank[node["node_type"]]:
            raise ValueError(
                f"Node '{node['node_label']}' has an invalid parent hierarchy."
            )

    cleaned_mappings_by_key: dict[tuple[str | None, str], dict] = {}
    for index, raw_mapping in enumerate(mapping_payload, start=1):
        icd_code = _normalize_icd_code(raw_mapping.get("icd_code"))
        if not icd_code:
            raise ValueError(f"Mapping {index} requires an ICD code.")
        node_import_id = str(raw_mapping.get("node_id") or "").strip()
        target_node = node_ids_by_import_id.get(node_import_id)
        if target_node is None:
            raise ValueError(f"Mapping '{icd_code}' references an unknown node.")
        if target_node["node_type"] != NODE_TYPE_FIELD:
            raise ValueError(f"Mapping '{icd_code}' must target a disease field node.")
        age_scope = _normalize_label(raw_mapping.get("age_scope"))
        age_scope = age_scope.lower() if age_scope else None
        if age_scope != target_node["age_scope"]:
            raise ValueError(f"Mapping '{icd_code}' age scope does not match its target node.")
        cleaned_mapping = {
            "age_scope": age_scope,
            "icd_code": icd_code,
            "node_import_id": node_import_id,
            "source_sheet": _normalize_label(raw_mapping.get("source_sheet")),
            "source_row_number": raw_mapping.get("source_row_number"),
            "source_category": _normalize_label(raw_mapping.get("source_category")),
            "match_type": _normalize_label(raw_mapping.get("match_type")),
            "mapping_note": _normalize_label(raw_mapping.get("mapping_note")),
            "is_active": bool(raw_mapping.get("is_active", True)),
        }
        mapping_key = (age_scope, icd_code)
        existing = cleaned_mappings_by_key.get(mapping_key)
        if existing is None:
            cleaned_mappings_by_key[mapping_key] = cleaned_mapping
            continue
        if existing["node_import_id"] != node_import_id:
            scope_label = age_scope or "all_ages"
            raise ValueError(
                f"ICD '{icd_code}' is mapped more than once for age scope "
                f"'{scope_label}' in the import payload."
            )

    scheme.scheme_name = normalized_name
    scheme.scheme_description = normalized_description or f"{normalized_name} COD bucket scheme"
    scheme.mapping_version = (scheme.mapping_version or 0) + 1
    scheme.is_active = bool(scheme_payload.get("is_active", True))

    _replace_scheme_contents(scheme)

    for age_band in cleaned_age_bands:
        _create_age_band(scheme=scheme, **age_band)

    created_nodes_by_import_id: dict[str, MasCodBucketNode] = {}
    for node in sorted(
        cleaned_nodes,
        key=lambda item: (
            item["age_scope"] or "",
            type_rank[item["node_type"]],
            item["sort_order"],
            item["node_label"].lower(),
        ),
    ):
        parent_node = (
            created_nodes_by_import_id.get(node["parent_node_id"])
            if node["parent_node_id"]
            else None
        )
        created_node = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=node["age_scope"],
            node_type=node["node_type"],
            parent=parent_node,
            node_code=node["node_code"],
            node_label=node["node_label"],
            sort_order=node["sort_order"],
            is_active=node["is_active"],
        )
        db.session.add(created_node)
        db.session.flush()
        created_nodes_by_import_id[node["import_node_id"]] = created_node

    for mapping in cleaned_mappings_by_key.values():
        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope=mapping["age_scope"],
                icd_code=mapping["icd_code"],
                node_id=created_nodes_by_import_id[mapping["node_import_id"]].node_id,
                source_sheet=mapping["source_sheet"],
                source_row_number=mapping["source_row_number"],
                source_category=mapping["source_category"],
                match_type=mapping["match_type"],
                mapping_note=mapping["mapping_note"],
                is_active=mapping["is_active"],
            )
        )

    db.session.commit()
    return scheme


def _format_xlsx_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells[:100]]
        width = min(max(len(value) for value in values) + 2, 48)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def export_cod_bucket_scheme_xlsx(*, scheme_code: str) -> bytes:
    payload = export_cod_bucket_scheme_json(scheme_code=scheme_code)

    workbook = Workbook()
    nodes_sheet = workbook.active
    nodes_sheet.title = "Bucket Nodes"
    nodes_headers = [
        "Scheme Code",
        "Scheme Name",
        "Age Scope",
        "Node Type",
        "Node Code",
        "Node Label",
        "Bucket Path",
        "Sort Order",
        "Active",
        "Parent Node ID",
        "Node ID",
    ]
    nodes_sheet.append(nodes_headers)
    for node in payload["nodes"]:
        nodes_sheet.append(
            [
                payload["scheme"]["scheme_code"],
                payload["scheme"]["scheme_name"],
                node["age_scope"],
                node["node_type"],
                node["node_code"],
                node["node_label"],
                node["path_label"],
                node["sort_order"],
                "Yes" if node["is_active"] else "No",
                node["parent_node_id"],
                node["node_id"],
            ]
        )

    mappings_sheet = workbook.create_sheet("ICD Mappings")
    mappings_headers = [
        "Scheme Code",
        "Scheme Name",
        "Age Scope",
        "ICD Code",
        "ICD Display",
        "Bucket Path",
        "Match Type",
        "Manual Override",
        "Source Sheet",
        "Source Row Number",
        "Source Category",
        "Mapping Note",
        "Active",
        "Mapping ID",
        "Node ID",
    ]
    mappings_sheet.append(mappings_headers)
    for mapping in payload["mappings"]:
        mappings_sheet.append(
            [
                payload["scheme"]["scheme_code"],
                payload["scheme"]["scheme_name"],
                mapping["age_scope"],
                mapping["icd_code"],
                mapping["icd_to_display"],
                mapping["node_path_label"],
                mapping["match_type"],
                (
                    "Yes"
                    if mapping["match_type"] == MANUAL_OVERRIDE_MATCH_TYPE
                    else "No"
                ),
                mapping["source_sheet"],
                mapping["source_row_number"],
                mapping["source_category"],
                mapping["mapping_note"],
                "Yes" if mapping["is_active"] else "No",
                mapping["mapping_id"],
                mapping["node_id"],
            ]
        )

    _format_xlsx_sheet(nodes_sheet)
    _format_xlsx_sheet(mappings_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def get_cod_bucket_node_mappings_payload(
    *,
    scheme_code: str,
    node_id,
) -> dict:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    node = db.session.get(MasCodBucketNode, node_id)
    if node is None or node.scheme_id != scheme.scheme_id:
        raise LookupError(f"Unknown COD bucket node: {node_id}")
    if node.node_type != NODE_TYPE_FIELD:
        raise ValueError("ICD mappings can only be loaded for field nodes.")

    icd_master_sq = _icd_master_display_subquery()
    mapping_rows = db.session.execute(
        sa.select(
            MapIcdCodBucket.mapping_id,
            MapIcdCodBucket.icd_code,
            MapIcdCodBucket.node_id,
            MapIcdCodBucket.match_type,
            MapIcdCodBucket.mapping_note,
            MapIcdCodBucket.source_sheet,
            MapIcdCodBucket.source_row_number,
            icd_master_sq.c.icd_to_display,
        )
        .select_from(MapIcdCodBucket)
        .outerjoin(
            icd_master_sq,
            icd_master_sq.c.icd_code == MapIcdCodBucket.icd_code,
        )
        .where(
            MapIcdCodBucket.scheme_id == scheme.scheme_id,
            MapIcdCodBucket.age_scope == node.age_scope,
            MapIcdCodBucket.node_id == node.node_id,
        )
        .order_by(MapIcdCodBucket.icd_code.asc())
    ).mappings().all()

    return {
        "node": {
            "node_id": str(node.node_id),
            "node_type": node.node_type,
            "node_label": node.node_label,
            "path_label": _node_path_label(node),
            "age_scope": node.age_scope,
        },
        "mappings": [
            {
                "mapping_id": str(row["mapping_id"]),
                "icd_code": row["icd_code"],
                "icd_to_display": row["icd_to_display"] or row["icd_code"],
                "node_id": str(row["node_id"]),
                "path_label": _node_path_label(node),
                "match_type": row["match_type"],
                "mapping_note": row["mapping_note"],
                "source_sheet": row["source_sheet"],
                "source_row_number": row["source_row_number"],
            }
            for row in mapping_rows
        ],
    }


def _node_depth_from_target(
    target_node: MasCodBucketNode,
    node: MasCodBucketNode,
    node_map: dict[uuid.UUID, MasCodBucketNode],
) -> int | None:
    depth = 0
    current = node
    while current is not None:
        if current.node_id == target_node.node_id:
            return depth
        current = node_map.get(current.parent_node_id)
        depth += 1
    return None


def _descendant_nodes_for_target(
    target_node: MasCodBucketNode,
    all_nodes: list[MasCodBucketNode],
) -> list[MasCodBucketNode]:
    node_map = {node.node_id: node for node in all_nodes}
    descendants = []
    for node in all_nodes:
        current = node
        while current is not None:
            if current.node_id == target_node.node_id:
                descendants.append(node)
                break
            current = node_map.get(current.parent_node_id)
    return descendants


def _make_unique_child_node_code(
    *,
    scheme_id,
    age_scope: str | None,
    parent_node_id,
    node_type: str,
    base_label: str,
) -> str:
    node_code_base = _slugify(base_label, fallback_prefix=node_type)
    node_code = node_code_base
    suffix = 2
    while db.session.scalar(
        sa.select(MasCodBucketNode.node_id).where(
            MasCodBucketNode.scheme_id == scheme_id,
            MasCodBucketNode.age_scope == age_scope,
            MasCodBucketNode.parent_node_id == parent_node_id,
            MasCodBucketNode.node_type == node_type,
            MasCodBucketNode.node_code == node_code,
        )
    ):
        node_code = f"{node_code_base}_{suffix}"
        suffix += 1
    return node_code


def _ensure_unmapped_replacement_leaf(
    *,
    scheme: MasCodBucketScheme,
    target_node: MasCodBucketNode,
    subtree_nodes: list[MasCodBucketNode],
    all_nodes: list[MasCodBucketNode],
) -> MasCodBucketNode:
    subtree_node_ids = {node.node_id for node in subtree_nodes}
    candidate_fields = [
        node for node in subtree_nodes
        if node.node_type == NODE_TYPE_FIELD
    ]
    if target_node.node_type == NODE_TYPE_FIELD:
        required_types = [NODE_TYPE_FIELD]
    else:
        node_map = {node.node_id: node for node in all_nodes}
        deepest_path_types = [target_node.node_type]
        deepest_depth = -1
        for field_node in candidate_fields:
            lineage = []
            current = field_node
            while current is not None:
                lineage.append(current)
                if current.node_id == target_node.node_id:
                    break
                current = node_map.get(current.parent_node_id)
            if not lineage or lineage[-1].node_id != target_node.node_id:
                continue
            lineage.reverse()
            if len(lineage) > deepest_depth:
                deepest_depth = len(lineage)
                deepest_path_types = [node.node_type for node in lineage]
        required_types = deepest_path_types
        if required_types[-1] != NODE_TYPE_FIELD:
            required_types.append(NODE_TYPE_FIELD)

    parent_node = target_node.parent
    current_parent = parent_node
    current_parent_id = parent_node.node_id if parent_node else None
    replacement_leaf = None
    starting_sort_order = target_node.sort_order

    for depth, node_type in enumerate(required_types):
        existing = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == target_node.age_scope,
                MasCodBucketNode.parent_node_id == current_parent_id,
                MasCodBucketNode.node_type == node_type,
                MasCodBucketNode.node_label == "Unmapped",
                MasCodBucketNode.is_active.is_(True),
                sa.not_(MasCodBucketNode.node_id.in_(subtree_node_ids)),
            )
        )
        if existing is None:
            sort_order = starting_sort_order if depth == 0 else 1
            existing = MasCodBucketNode(
                scheme_id=scheme.scheme_id,
                age_scope=target_node.age_scope,
                node_type=node_type,
                parent=current_parent,
                node_code=_make_unique_child_node_code(
                    scheme_id=scheme.scheme_id,
                    age_scope=target_node.age_scope,
                    parent_node_id=current_parent_id,
                    node_type=node_type,
                    base_label="unmapped",
                ),
                node_label="Unmapped",
                sort_order=sort_order,
                is_active=True,
            )
            db.session.add(existing)
            db.session.flush()
        current_parent = existing
        current_parent_id = existing.node_id
        replacement_leaf = existing

    return replacement_leaf


def delete_cod_bucket_node(
    *,
    scheme_code: str,
    node_id,
    mapping_disposition: str,
) -> dict:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    target_node = db.session.get(MasCodBucketNode, node_id)
    if target_node is None or target_node.scheme_id != scheme.scheme_id:
        raise LookupError(f"Unknown COD bucket node: {node_id}")

    if mapping_disposition not in {
        NODE_DELETE_DISPOSITION_UNMAP,
        NODE_DELETE_DISPOSITION_MOVE_TO_UNMAPPED,
    }:
        raise ValueError("mapping_disposition must be unmap or move_to_unmapped.")

    all_nodes = list(
        db.session.scalars(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == target_node.age_scope,
            )
        )
    )
    subtree_nodes = _descendant_nodes_for_target(target_node, all_nodes)
    subtree_node_ids = {node.node_id for node in subtree_nodes}
    field_node_ids = [
        node.node_id for node in subtree_nodes
        if node.node_type == NODE_TYPE_FIELD
    ]

    mappings = list(
        db.session.scalars(
            sa.select(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.age_scope == target_node.age_scope,
                MapIcdCodBucket.node_id.in_(field_node_ids if field_node_ids else [uuid.uuid4()]),
            )
        )
    ) if field_node_ids else []

    replacement_leaf = None
    if mappings and mapping_disposition == NODE_DELETE_DISPOSITION_MOVE_TO_UNMAPPED:
        replacement_leaf = _ensure_unmapped_replacement_leaf(
            scheme=scheme,
            target_node=target_node,
            subtree_nodes=subtree_nodes,
            all_nodes=all_nodes,
        )
        for mapping in mappings:
            mapping.node_id = replacement_leaf.node_id

    affected_icd_codes = [mapping.icd_code for mapping in mappings]
    deleted_node_path = _node_path_label(target_node)
    db.session.delete(target_node)
    db.session.commit()

    return {
        "deleted_node_id": str(node_id),
        "deleted_path_label": deleted_node_path,
        "deleted_node_count": len(subtree_nodes),
        "affected_icd_count": len(affected_icd_codes),
        "affected_icd_codes": affected_icd_codes,
        "mapping_disposition": mapping_disposition,
        "replacement_leaf_node_id": str(replacement_leaf.node_id) if replacement_leaf else None,
        "replacement_leaf_path_label": _node_path_label(replacement_leaf) if replacement_leaf else None,
    }


def search_cod_bucket_icd_candidates(
    *,
    scheme_code: str,
    age_scope: str | None,
    query: str,
    selected_node_id=None,
    unmapped_only: bool = False,
    limit: int = _ICD_SEARCH_LIMIT,
) -> list[dict]:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    normalized_query = _normalize_search_query(query)
    if len(normalized_query) < 2:
        return []

    reporting_catalog_sq = _reporting_icd_catalog_subquery()
    escaped_query = _escape_like(normalized_query)
    code_prefix = f"{escaped_query}%"
    title_prefix = f"{escaped_query}%"
    title_contains = f"%{escaped_query}%"
    lower_code = sa.func.lower(reporting_catalog_sq.c.code)
    lower_title = sa.func.lower(reporting_catalog_sq.c.title)

    rank_expr = sa.case(
        (lower_code == normalized_query, 0),
        (lower_code.like(code_prefix, escape=_LIKE_ESCAPE), 1),
        (lower_title.like(title_prefix, escape=_LIKE_ESCAPE), 2),
        (lower_title.like(title_contains, escape=_LIKE_ESCAPE), 3),
        else_=4,
    )

    stmt = (
        sa.select(
            reporting_catalog_sq.c.code.label("icd_code"),
            reporting_catalog_sq.c.title.label("icd_to_display"),
            reporting_catalog_sq.c.is_coding_selectable.label("is_coding_selectable"),
            MapIcdCodBucket.mapping_id,
            MapIcdCodBucket.node_id,
        )
        .select_from(reporting_catalog_sq)
        .outerjoin(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.age_scope == age_scope,
                MapIcdCodBucket.icd_code == reporting_catalog_sq.c.code,
            ),
        )
        .where(
            sa.or_(
                lower_code.like(code_prefix, escape=_LIKE_ESCAPE),
                lower_title.like(title_prefix, escape=_LIKE_ESCAPE),
                lower_title.like(title_contains, escape=_LIKE_ESCAPE),
            )
        )
        .order_by(rank_expr, reporting_catalog_sq.c.code.asc())
        .limit(limit)
    )
    if unmapped_only:
        stmt = stmt.where(MapIcdCodBucket.mapping_id.is_(None))

    rows = db.session.execute(stmt).mappings().all()
    mapped_node_ids = {
        row["node_id"]
        for row in rows
        if row["node_id"] is not None
    }
    mapped_nodes = {
        node.node_id: node
        for node in db.session.scalars(
            sa.select(MasCodBucketNode).where(MasCodBucketNode.node_id.in_(mapped_node_ids))
        )
    } if mapped_node_ids else {}

    return [
        {
            "icd_code": row["icd_code"],
            "icd_to_display": row["icd_to_display"] or row["icd_code"],
            "current_node_id": str(row["node_id"]) if row["node_id"] else None,
            "current_path_label": (
                _node_path_label(mapped_nodes[row["node_id"]])
                if row["node_id"] in mapped_nodes
                else None
            ),
            "is_mapped": row["node_id"] is not None,
            "is_selected_leaf": bool(selected_node_id and row["node_id"] == selected_node_id),
            "is_assignable_in_coding": bool(row["is_coding_selectable"]),
            "coding_status_label": (
                None if row["is_coding_selectable"] else "Currently not assignable in coding"
            ),
        }
        for row in rows
    ]


def list_cod_bucket_unmapped_icd_rows(
    *,
    scheme_code: str,
) -> dict:
    scheme = get_cod_bucket_scheme(scheme_code)
    if scheme is None:
        raise LookupError(f"Unknown COD bucket scheme: {scheme_code}")

    mapped_codes_sq = (
        sa.select(MapIcdCodBucket.icd_code.label("icd_code"))
        .where(
            MapIcdCodBucket.scheme_id == scheme.scheme_id,
            MapIcdCodBucket.is_active.is_(True),
        )
        .group_by(MapIcdCodBucket.icd_code)
        .subquery()
    )

    reporting_catalog_sq = _reporting_icd_catalog_subquery()
    rows = db.session.execute(
        sa.select(
            reporting_catalog_sq.c.code,
            reporting_catalog_sq.c.title,
            reporting_catalog_sq.c.semantic_level,
            reporting_catalog_sq.c.is_coding_selectable,
            reporting_catalog_sq.c.chapter_code,
            reporting_catalog_sq.c.chapter_title,
            reporting_catalog_sq.c.three_character_code,
            reporting_catalog_sq.c.three_character_title,
        )
        .select_from(reporting_catalog_sq)
        .outerjoin(mapped_codes_sq, mapped_codes_sq.c.icd_code == reporting_catalog_sq.c.code)
        .where(
            mapped_codes_sq.c.icd_code.is_(None),
        )
        .order_by(
            reporting_catalog_sq.c.chapter_code.asc(),
            reporting_catalog_sq.c.three_character_code.asc(),
            reporting_catalog_sq.c.code.asc(),
        )
    ).mappings().all()
    final_cod_counts: dict[str, int] = {}
    smartva_counts: dict[str, int] = {}
    analytics_views_available = bool(
        db.session.scalar(sa.text(f"SELECT to_regclass('{COD_MV_NAME}')"))
        and db.session.scalar(sa.text(f"SELECT to_regclass('{DEMOGRAPHICS_MV_NAME}')"))
    )
    if analytics_views_available:
        cod = sa.table(
            COD_MV_NAME,
            sa.column("va_sid"),
            sa.column("final_icd"),
        )
        demo = sa.table(
            DEMOGRAPHICS_MV_NAME,
            sa.column("va_sid"),
            sa.column("has_human_final_cod"),
        )
        reporting_alias = sa.orm.aliased(MapIcd10LegacyReportingAlias)
        final_cod_counts = {
            row["final_icd"]: int(row["final_cod_count"])
            for row in db.session.execute(
                sa.select(
                    cod.c.final_icd,
                    sa.func.count().label("final_cod_count"),
                )
                .select_from(cod)
                .join(demo, demo.c.va_sid == cod.c.va_sid)
                .outerjoin(reporting_alias, reporting_alias.legacy_code == cod.c.final_icd)
                .outerjoin(
                    mapped_codes_sq,
                    mapped_codes_sq.c.icd_code == sa.func.coalesce(
                        reporting_alias.reporting_code,
                        cod.c.final_icd,
                    ),
                )
                .where(
                    cod.c.final_icd.is_not(None),
                    demo.c.has_human_final_cod.is_(True),
                    mapped_codes_sq.c.icd_code.is_(None),
                )
                .group_by(cod.c.final_icd)
            ).mappings()
        }
        latest_smartva_sq = (
            sa.select(
                VaSmartvaResults.va_sid.label("va_sid"),
                sa.func.upper(VaSmartvaResults.va_smartva_cause1icd).label("cause1_icd"),
                sa.func.upper(VaSmartvaResults.va_smartva_cause2icd).label("cause2_icd"),
                sa.func.upper(VaSmartvaResults.va_smartva_cause3icd).label("cause3_icd"),
            )
            .where(VaSmartvaResults.va_smartva_status == "active")
            .distinct(VaSmartvaResults.va_sid)
            .order_by(
                VaSmartvaResults.va_sid,
                VaSmartvaResults.va_smartva_updatedat.desc(),
                VaSmartvaResults.va_smartva_id.desc(),
            )
            .subquery()
        )
        smartva_alias1 = sa.orm.aliased(MapIcd10LegacyReportingAlias)
        smartva_alias2 = sa.orm.aliased(MapIcd10LegacyReportingAlias)
        smartva_alias3 = sa.orm.aliased(MapIcd10LegacyReportingAlias)
        smartva_codes_sq = sa.union_all(
            sa.select(
                latest_smartva_sq.c.va_sid,
                latest_smartva_sq.c.cause1_icd.label("smartva_icd"),
                sa.func.coalesce(
                    smartva_alias1.reporting_code,
                    latest_smartva_sq.c.cause1_icd,
                ).label("mapped_icd"),
            )
            .select_from(latest_smartva_sq)
            .outerjoin(
                smartva_alias1,
                smartva_alias1.legacy_code == latest_smartva_sq.c.cause1_icd,
            )
            .where(latest_smartva_sq.c.cause1_icd.is_not(None)),
            sa.select(
                latest_smartva_sq.c.va_sid,
                latest_smartva_sq.c.cause2_icd.label("smartva_icd"),
                sa.func.coalesce(
                    smartva_alias2.reporting_code,
                    latest_smartva_sq.c.cause2_icd,
                ).label("mapped_icd"),
            )
            .select_from(latest_smartva_sq)
            .outerjoin(
                smartva_alias2,
                smartva_alias2.legacy_code == latest_smartva_sq.c.cause2_icd,
            )
            .where(latest_smartva_sq.c.cause2_icd.is_not(None)),
            sa.select(
                latest_smartva_sq.c.va_sid,
                latest_smartva_sq.c.cause3_icd.label("smartva_icd"),
                sa.func.coalesce(
                    smartva_alias3.reporting_code,
                    latest_smartva_sq.c.cause3_icd,
                ).label("mapped_icd"),
            )
            .select_from(latest_smartva_sq)
            .outerjoin(
                smartva_alias3,
                smartva_alias3.legacy_code == latest_smartva_sq.c.cause3_icd,
            )
            .where(latest_smartva_sq.c.cause3_icd.is_not(None)),
        ).subquery()
        smartva_counts = {
            row["smartva_icd"]: int(row["smartva_count"])
            for row in db.session.execute(
                sa.select(
                    smartva_codes_sq.c.smartva_icd,
                    sa.func.count(
                        sa.distinct(
                            sa.tuple_(
                                smartva_codes_sq.c.va_sid,
                                smartva_codes_sq.c.smartva_icd,
                            )
                        )
                    ).label("smartva_count"),
                )
                .select_from(smartva_codes_sq)
                .outerjoin(
                    mapped_codes_sq,
                    mapped_codes_sq.c.icd_code == smartva_codes_sq.c.mapped_icd,
                )
                .where(
                    smartva_codes_sq.c.smartva_icd.is_not(None),
                    mapped_codes_sq.c.icd_code.is_(None),
                )
                .group_by(smartva_codes_sq.c.smartva_icd)
            ).mappings()
        }

    payload_rows = [
        {
            "chapter": (
                f'{row["chapter_code"]} {row["chapter_title"]}'
                if row["chapter_code"] and row["chapter_title"]
                else (row["chapter_code"] or "")
            ),
            "three_character_code": row["three_character_code"] or "",
            "three_character_title": row["three_character_title"] or "",
            "detailed_code": row["code"] if row["semantic_level"] == "detailed_code" else "",
            "detailed_title": row["title"] if row["semantic_level"] == "detailed_code" else "",
            "semantic_level": row["semantic_level"],
            "code": row["code"],
            "title": row["title"],
            "final_cod_count": final_cod_counts.get(row["code"], 0),
            "is_utilized_in_final_cod": final_cod_counts.get(row["code"], 0) > 0,
            "smartva_count": smartva_counts.get(row["code"], 0),
            "is_utilized_in_smartva": smartva_counts.get(row["code"], 0) > 0,
            "is_assignable_in_coding": bool(row["is_coding_selectable"]),
            "coding_status_label": (
                None
                if row["is_coding_selectable"]
                else "Currently not assignable in coding"
            ),
        }
        for row in rows
    ]
    known_codes = {row["code"] for row in payload_rows}
    orphan_codes = sorted(
        set(code for code in final_cod_counts if code not in known_codes)
        | set(code for code in smartva_counts if code not in known_codes)
    )
    for icd_code in orphan_codes:
        final_count = final_cod_counts.get(icd_code, 0)
        smartva_count = smartva_counts.get(icd_code, 0)
        payload_rows.append(
            {
                "chapter": "",
                "three_character_code": "",
                "three_character_title": "",
                "detailed_code": "",
                "detailed_title": "",
                "semantic_level": "unknown",
                "code": icd_code,
                "title": "",
                "final_cod_count": final_count,
                "smartva_count": smartva_count,
                "is_utilized_in_final_cod": final_count > 0,
                "is_utilized_in_smartva": smartva_count > 0,
                "is_assignable_in_coding": False,
                "coding_status_label": "Currently not assignable in coding",
            }
        )

    return {
        "scheme": {
            "scheme_code": scheme.scheme_code,
            "scheme_name": scheme.scheme_name,
        },
        "rows": payload_rows,
        "total_rows": len(payload_rows),
    }


def _age_bound_days_sql(value_column, unit_column):
    return sa.case(
        (
            sa.and_(value_column.is_not(None), unit_column == AGE_UNIT_DAYS),
            sa.cast(value_column, sa.Numeric()),
        ),
        (
            sa.and_(value_column.is_not(None), unit_column == AGE_UNIT_MONTHS),
            sa.cast(value_column, sa.Numeric()) * sa.literal(_DAYS_PER_MONTH),
        ),
        (
            sa.and_(value_column.is_not(None), unit_column == AGE_UNIT_YEARS),
            sa.cast(value_column, sa.Numeric()) * sa.literal(365),
        ),
        else_=None,
    )


def aggregate_coded_submissions_by_bucket(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
    collapse_scope: bool = False,
) -> list[dict]:
    """Return grouped coded-form counts by bucket for the given scheme."""
    if allowed_project_site_pairs is not None and not allowed_project_site_pairs:
        return []
    scheme, base_rows = _cod_bucket_aggregate_base_subquery(
        scheme_code=scheme_code,
        project_id=project_id,
        site_id=site_id,
        form_id=form_id,
        gender=gender,
        submission_date_from=submission_date_from,
        submission_date_to=submission_date_to,
        allowed_project_site_pairs=allowed_project_site_pairs,
    )

    field_node = sa.orm.aliased(MasCodBucketNode)
    parent_node = sa.orm.aliased(MasCodBucketNode)
    grandparent_node = sa.orm.aliased(MasCodBucketNode)

    category_label = sa.case(
        (field_node.parent_node_id.is_(None), sa.null()),
        (parent_node.node_type == NODE_TYPE_CATEGORY, parent_node.node_label),
        else_=grandparent_node.node_label,
    ).label("bucket_category")
    category_sort_order = sa.case(
        (field_node.parent_node_id.is_(None), 0),
        (parent_node.node_type == NODE_TYPE_CATEGORY, parent_node.sort_order),
        else_=grandparent_node.sort_order,
    ).label("bucket_category_sort_order")
    subcategory_label = sa.case(
        (field_node.parent_node_id.is_(None), sa.null()),
        (parent_node.node_type == NODE_TYPE_SUBCATEGORY, parent_node.node_label),
        else_=sa.null(),
    ).label("bucket_subcategory")
    subcategory_sort_order = sa.case(
        (field_node.parent_node_id.is_(None), 0),
        (parent_node.node_type == NODE_TYPE_SUBCATEGORY, parent_node.sort_order),
        else_=0,
    ).label("bucket_subcategory_sort_order")
    gender_normalized = sa.func.lower(sa.func.coalesce(base_rows.c.gender, "unknown"))
    male_count = sa.func.sum(
        sa.case((gender_normalized == "male", 1), else_=0)
    ).label("male_count")
    female_count = sa.func.sum(
        sa.case((gender_normalized == "female", 1), else_=0)
    ).label("female_count")
    unknown_count = sa.func.sum(
        sa.case(
            (
                sa.and_(
                    gender_normalized != "male",
                    gender_normalized != "female",
                ),
                1,
            ),
            else_=0,
        )
    ).label("unknown_count")

    select_columns = [
        base_rows.c.age_scope,
        base_rows.c.age_scope_label,
        base_rows.c.age_scope_sort_order,
        category_label,
        category_sort_order,
        subcategory_label,
        subcategory_sort_order,
        field_node.node_label.label("bucket_field"),
        field_node.sort_order.label("bucket_field_sort_order"),
        male_count,
        female_count,
        unknown_count,
        sa.func.count().label("coded_count"),
    ]
    group_by_columns = [
        base_rows.c.age_scope,
        base_rows.c.age_scope_label,
        base_rows.c.age_scope_sort_order,
        category_label,
        category_sort_order,
        subcategory_label,
        subcategory_sort_order,
        field_node.node_label,
        field_node.sort_order,
    ]
    order_by_columns = [
        base_rows.c.age_scope_sort_order.asc(),
        category_sort_order.asc(),
        subcategory_sort_order.asc(),
        field_node.sort_order.asc(),
    ]

    if not collapse_scope:
        select_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *select_columns,
        ]
        group_by_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *group_by_columns,
        ]
        order_by_columns = [
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            *order_by_columns,
        ]

    query = (
        sa.select(*select_columns)
        .select_from(base_rows)
        .join(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                sa.or_(
                    sa.and_(
                        MapIcdCodBucket.age_scope.is_(None),
                        base_rows.c.age_scope.is_(None),
                    ),
                    MapIcdCodBucket.age_scope == base_rows.c.age_scope,
                ),
            ),
        )
        .join(field_node, field_node.node_id == MapIcdCodBucket.node_id)
        .outerjoin(parent_node, parent_node.node_id == field_node.parent_node_id)
        .outerjoin(grandparent_node, grandparent_node.node_id == parent_node.parent_node_id)
        .group_by(*group_by_columns)
        .order_by(*order_by_columns)
    )
    rows = [dict(row) for row in db.session.execute(query).mappings().all()]

    if not _uses_reporting_age_band_detail_sections(scheme):
        return rows

    reporting_age_scope = _reporting_age_band_sql(
        base_rows.c.age_normalized_days
    ).label("age_scope")
    reporting_age_scope_label = reporting_age_scope.label("age_scope_label")
    reporting_age_scope_sort_order = _detailed_reporting_age_scope_sort_sql(
        base_rows.c.age_normalized_days
    ).label("age_scope_sort_order")

    reporting_select_columns = [
        reporting_age_scope,
        reporting_age_scope_label,
        reporting_age_scope_sort_order,
        category_label,
        category_sort_order,
        subcategory_label,
        subcategory_sort_order,
        field_node.node_label.label("bucket_field"),
        field_node.sort_order.label("bucket_field_sort_order"),
        male_count,
        female_count,
        unknown_count,
        sa.func.count().label("coded_count"),
    ]
    reporting_group_by_columns = [
        reporting_age_scope,
        reporting_age_scope_label,
        reporting_age_scope_sort_order,
        category_label,
        category_sort_order,
        subcategory_label,
        subcategory_sort_order,
        field_node.node_label,
        field_node.sort_order,
    ]
    reporting_order_by_columns = [
        reporting_age_scope_sort_order.asc(),
        category_sort_order.asc(),
        subcategory_sort_order.asc(),
        field_node.sort_order.asc(),
    ]

    if not collapse_scope:
        reporting_select_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *reporting_select_columns,
        ]
        reporting_group_by_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *reporting_group_by_columns,
        ]
        reporting_order_by_columns = [
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            *reporting_order_by_columns,
        ]

    reporting_query = (
        sa.select(*reporting_select_columns)
        .select_from(base_rows)
        .join(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                MapIcdCodBucket.age_scope.is_(None),
                base_rows.c.age_scope.is_(None),
            ),
        )
        .join(field_node, field_node.node_id == MapIcdCodBucket.node_id)
        .outerjoin(parent_node, parent_node.node_id == field_node.parent_node_id)
        .outerjoin(grandparent_node, grandparent_node.node_id == parent_node.parent_node_id)
        .group_by(*reporting_group_by_columns)
        .order_by(*reporting_order_by_columns)
    )
    rows.extend(dict(row) for row in db.session.execute(reporting_query).mappings().all())
    return rows


def summarize_unmatched_coded_submissions_by_bucket(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
    collapse_scope: bool = False,
) -> list[dict]:
    """Return counts for coded submissions that do not match the active scheme."""
    if allowed_project_site_pairs is not None and not allowed_project_site_pairs:
        return []
    scheme, base_rows = _cod_bucket_aggregate_base_subquery(
        scheme_code=scheme_code,
        project_id=project_id,
        site_id=site_id,
        form_id=form_id,
        gender=gender,
        submission_date_from=submission_date_from,
        submission_date_to=submission_date_to,
        allowed_project_site_pairs=allowed_project_site_pairs,
    )

    select_columns = [
        base_rows.c.age_scope,
        base_rows.c.age_scope_label,
        base_rows.c.age_scope_sort_order,
        sa.func.count().label("unmatched_count"),
    ]
    group_by_columns = [base_rows.c.age_scope, base_rows.c.age_scope_label, base_rows.c.age_scope_sort_order]
    order_by_columns = [base_rows.c.age_scope_sort_order.asc()]

    if not collapse_scope:
        select_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *select_columns,
        ]
        group_by_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *group_by_columns,
        ]
        order_by_columns = [
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            *order_by_columns,
        ]

    query = (
        sa.select(*select_columns)
        .select_from(base_rows)
        .outerjoin(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                sa.or_(
                    sa.and_(
                        MapIcdCodBucket.age_scope.is_(None),
                        base_rows.c.age_scope.is_(None),
                    ),
                    MapIcdCodBucket.age_scope == base_rows.c.age_scope,
                ),
            ),
        )
        .where(MapIcdCodBucket.mapping_id.is_(None))
        .group_by(*group_by_columns)
        .order_by(*order_by_columns)
    )
    rows = [dict(row) for row in db.session.execute(query).mappings().all()]

    if not _uses_reporting_age_band_detail_sections(scheme):
        return rows

    reporting_age_scope = _reporting_age_band_sql(
        base_rows.c.age_normalized_days
    ).label("age_scope")
    reporting_age_scope_label = reporting_age_scope.label("age_scope_label")
    reporting_age_scope_sort_order = _detailed_reporting_age_scope_sort_sql(
        base_rows.c.age_normalized_days
    ).label("age_scope_sort_order")

    reporting_select_columns = [
        reporting_age_scope,
        reporting_age_scope_label,
        reporting_age_scope_sort_order,
        sa.func.count().label("unmatched_count"),
    ]
    reporting_group_by_columns = [
        reporting_age_scope,
        reporting_age_scope_label,
        reporting_age_scope_sort_order,
    ]
    reporting_order_by_columns = [reporting_age_scope_sort_order.asc()]

    if not collapse_scope:
        reporting_select_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *reporting_select_columns,
        ]
        reporting_group_by_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *reporting_group_by_columns,
        ]
        reporting_order_by_columns = [
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            *reporting_order_by_columns,
        ]

    reporting_query = (
        sa.select(*reporting_select_columns)
        .select_from(base_rows)
        .outerjoin(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                MapIcdCodBucket.age_scope.is_(None),
                base_rows.c.age_scope.is_(None),
            ),
        )
        .where(MapIcdCodBucket.mapping_id.is_(None))
        .group_by(*reporting_group_by_columns)
        .order_by(*reporting_order_by_columns)
    )
    rows.extend(dict(row) for row in db.session.execute(reporting_query).mappings().all())
    return rows


def list_unmatched_coded_submission_icds_by_bucket(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
    collapse_scope: bool = False,
) -> list[dict]:
    """Return unmatched ICD codes and counts for the active scheme."""
    if allowed_project_site_pairs is not None and not allowed_project_site_pairs:
        return []
    scheme, base_rows = _cod_bucket_aggregate_base_subquery(
        scheme_code=scheme_code,
        project_id=project_id,
        site_id=site_id,
        form_id=form_id,
        gender=gender,
        submission_date_from=submission_date_from,
        submission_date_to=submission_date_to,
        allowed_project_site_pairs=allowed_project_site_pairs,
    )

    select_columns = [
        base_rows.c.age_scope,
        base_rows.c.age_scope_label,
        base_rows.c.age_scope_sort_order,
        base_rows.c.final_icd.label("icd_code"),
        base_rows.c.reporting_icd,
        sa.func.count().label("unmatched_count"),
    ]
    group_by_columns = [
        base_rows.c.age_scope,
        base_rows.c.age_scope_label,
        base_rows.c.age_scope_sort_order,
        base_rows.c.final_icd,
        base_rows.c.reporting_icd,
    ]
    order_by_columns = [
        base_rows.c.age_scope_sort_order.asc(),
        sa.func.count().desc(),
        base_rows.c.final_icd.asc(),
    ]

    if not collapse_scope:
        select_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *select_columns,
        ]
        group_by_columns = [
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            *group_by_columns,
        ]
        order_by_columns = [
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            *order_by_columns,
        ]

    query = (
        sa.select(*select_columns)
        .select_from(base_rows)
        .outerjoin(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                sa.or_(
                    sa.and_(
                        MapIcdCodBucket.age_scope.is_(None),
                        base_rows.c.age_scope.is_(None),
                    ),
                    MapIcdCodBucket.age_scope == base_rows.c.age_scope,
                ),
            ),
        )
        .where(MapIcdCodBucket.mapping_id.is_(None))
        .group_by(*group_by_columns)
        .order_by(*order_by_columns)
    )
    rows = [dict(row) for row in db.session.execute(query).mappings().all()]

    if _uses_reporting_age_band_detail_sections(scheme):
        reporting_age_scope = _reporting_age_band_sql(
            base_rows.c.age_normalized_days
        ).label("age_scope")
        reporting_age_scope_label = reporting_age_scope.label("age_scope_label")
        reporting_age_scope_sort_order = _detailed_reporting_age_scope_sort_sql(
            base_rows.c.age_normalized_days
        ).label("age_scope_sort_order")

        reporting_select_columns = [
            reporting_age_scope,
            reporting_age_scope_label,
            reporting_age_scope_sort_order,
            base_rows.c.final_icd.label("icd_code"),
            base_rows.c.reporting_icd,
            sa.func.count().label("unmatched_count"),
        ]
        reporting_group_by_columns = [
            reporting_age_scope,
            reporting_age_scope_label,
            reporting_age_scope_sort_order,
            base_rows.c.final_icd,
            base_rows.c.reporting_icd,
        ]
        reporting_order_by_columns = [
            reporting_age_scope_sort_order.asc(),
            sa.func.count().desc(),
            base_rows.c.final_icd.asc(),
        ]

        if not collapse_scope:
            reporting_select_columns = [
                base_rows.c.project_id,
                base_rows.c.site_id,
                base_rows.c.form_id,
                *reporting_select_columns,
            ]
            reporting_group_by_columns = [
                base_rows.c.project_id,
                base_rows.c.site_id,
                base_rows.c.form_id,
                *reporting_group_by_columns,
            ]
            reporting_order_by_columns = [
                base_rows.c.project_id.asc(),
                base_rows.c.site_id.asc(),
                base_rows.c.form_id.asc(),
                *reporting_order_by_columns,
            ]

        reporting_query = (
            sa.select(*reporting_select_columns)
            .select_from(base_rows)
            .outerjoin(
                MapIcdCodBucket,
                sa.and_(
                    MapIcdCodBucket.scheme_id == scheme.scheme_id,
                    MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                    MapIcdCodBucket.age_scope.is_(None),
                    base_rows.c.age_scope.is_(None),
                ),
            )
            .where(MapIcdCodBucket.mapping_id.is_(None))
            .group_by(*reporting_group_by_columns)
            .order_by(*reporting_order_by_columns)
        )
        rows.extend(
            dict(row) for row in db.session.execute(reporting_query).mappings().all()
        )

    reporting_icd_codes = [row["reporting_icd"] for row in rows if row["reporting_icd"]]
    master_rows = {
        row.code: row
        for row in db.session.scalars(
            sa.select(MasIcd1020192).where(MasIcd1020192.code.in_(reporting_icd_codes))
        )
    } if reporting_icd_codes else {}

    classified_rows = []
    for row in rows:
        item = dict(row)
        master_row = master_rows.get(item["reporting_icd"])
        is_master_coding_eligible = bool(
            master_row
            and master_row.is_active
            and master_row.is_coding_selectable
            and master_row.semantic_level in {"three_character", "detailed_code"}
        )
        item["category"] = (
            "not_included_in_scheme"
            if is_master_coding_eligible
            else "not_eligible_for_coding"
        )
        item["category_label"] = (
            "ICD codes not included in CoD Categories"
            if is_master_coding_eligible
            else "ICD codes not eligible for coding"
        )
        item["is_master_coding_eligible"] = is_master_coding_eligible
        item.pop("reporting_icd", None)
        classified_rows.append(item)
    return classified_rows


def _matched_cod_bucket_rows_subquery(*, scheme, base_rows):
    field_node = sa.orm.aliased(MasCodBucketNode)
    parent_node = sa.orm.aliased(MasCodBucketNode)
    grandparent_node = sa.orm.aliased(MasCodBucketNode)

    category_label = sa.case(
        (field_node.parent_node_id.is_(None), sa.null()),
        (parent_node.node_type == NODE_TYPE_CATEGORY, parent_node.node_label),
        else_=grandparent_node.node_label,
    ).label("bucket_category")
    subcategory_label = sa.case(
        (field_node.parent_node_id.is_(None), sa.null()),
        (parent_node.node_type == NODE_TYPE_SUBCATEGORY, parent_node.node_label),
        else_=sa.null(),
    ).label("bucket_subcategory")

    query = (
        sa.select(
            base_rows.c.age_scope,
            base_rows.c.age_scope_label,
            base_rows.c.age_scope_sort_order,
            base_rows.c.age_normalized_days,
            base_rows.c.gender,
            base_rows.c.country,
            base_rows.c.submission_year,
            category_label,
            subcategory_label,
            field_node.node_label.label("bucket_field"),
        )
        .select_from(base_rows)
        .join(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                sa.or_(
                    sa.and_(
                        MapIcdCodBucket.age_scope.is_(None),
                        base_rows.c.age_scope.is_(None),
                    ),
                    MapIcdCodBucket.age_scope == base_rows.c.age_scope,
                ),
            ),
        )
        .join(field_node, field_node.node_id == MapIcdCodBucket.node_id)
        .outerjoin(parent_node, parent_node.node_id == field_node.parent_node_id)
        .outerjoin(grandparent_node, grandparent_node.node_id == parent_node.parent_node_id)
    )
    return query.subquery("matched_cod_bucket_rows")


def summarize_cod_bucket_reporting_breakdowns(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
    top_n: int = 10,
) -> dict:
    """Return dashboard summaries for the current COD bucket reporting scope."""
    if allowed_project_site_pairs is not None and not allowed_project_site_pairs:
        return {
            "scheme_used": scheme_code,
            "top_causes": [],
            "top_causes_by_age": {"all": []},
            "first_level_counts": [],
            "first_level_counts_by_age": {"all": []},
            "age_filters": [{"key": "all", "label": "All ages"}],
            "age_sex_distribution": [],
            "gender_distribution": [],
            "heatmap": {
                "dimension": "country",
                "view": "top_causes",
                "views": {
                    "top_causes": {},
                    "first_level_counts": {},
                },
            },
            "treemap": [],
            "matched_total": 0,
        }

    scheme, base_rows = _cod_bucket_aggregate_base_subquery(
        scheme_code=scheme_code,
        project_id=project_id,
        site_id=site_id,
        form_id=form_id,
        gender=gender,
        submission_date_from=submission_date_from,
        submission_date_to=submission_date_to,
        allowed_project_site_pairs=allowed_project_site_pairs,
    )
    matched_rows = _matched_cod_bucket_rows_subquery(scheme=scheme, base_rows=base_rows)

    matched_total = int(
        db.session.scalar(sa.select(sa.func.count()).select_from(matched_rows)) or 0
    )
    if matched_total == 0:
        return {
            "scheme_used": scheme.scheme_name,
            "top_causes": [],
            "top_causes_by_age": {"all": []},
            "first_level_counts": [],
            "first_level_counts_by_age": {"all": []},
            "age_filters": [{"key": "all", "label": "All ages"}],
            "age_sex_distribution": [],
            "gender_distribution": [],
            "heatmap": {
                "dimension": "country",
                "view": "top_causes",
                "views": {
                    "top_causes": {},
                    "first_level_counts": {},
                },
            },
            "treemap": [],
            "matched_total": 0,
        }

    top_cause_rows = db.session.execute(
        sa.select(
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
        )
        .order_by(
            sa.func.count().desc(),
            matched_rows.c.bucket_field.asc(),
            matched_rows.c.bucket_category.asc().nullslast(),
            matched_rows.c.bucket_subcategory.asc().nullslast(),
        )
        .limit(top_n)
    ).mappings().all()

    first_level_label = (
        matched_rows.c.bucket_field
        if scheme.scheme_code == SCHEME_CODE_CMEA10
        else sa.func.coalesce(matched_rows.c.bucket_category, matched_rows.c.bucket_field)
    ).label("first_level_label")
    first_level_rows = db.session.execute(
        sa.select(
            first_level_label,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(first_level_label)
        .order_by(sa.func.count().desc(), first_level_label.asc())
    ).mappings().all()

    if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
        age_band_rows = db.session.execute(
            sa.select(
                matched_rows.c.age_scope.label("age_band_key"),
                matched_rows.c.age_scope_label.label("age_band"),
                matched_rows.c.age_scope_sort_order.label("sort_order"),
                matched_rows.c.gender,
                sa.func.count().label("coded_count"),
            )
            .select_from(matched_rows)
            .group_by(
                matched_rows.c.age_scope,
                matched_rows.c.age_scope_label,
                matched_rows.c.age_scope_sort_order,
                matched_rows.c.gender,
            )
            .order_by(
                matched_rows.c.age_scope_sort_order.asc(),
                matched_rows.c.gender.asc(),
            )
        ).mappings().all()
    else:
        age_band_label = _reporting_age_band_sql(
            matched_rows.c.age_normalized_days
        ).label("age_band")
        age_band_sort_order = _reporting_age_band_sort_sql(
            matched_rows.c.age_normalized_days
        ).label("sort_order")
        age_band_rows = db.session.execute(
            sa.select(
                age_band_label.label("age_band_key"),
                age_band_label,
                age_band_sort_order,
                matched_rows.c.gender,
                sa.func.count().label("coded_count"),
            )
            .select_from(matched_rows)
            .group_by(
                age_band_label,
                age_band_sort_order,
                matched_rows.c.gender,
            )
            .order_by(age_band_sort_order.asc(), matched_rows.c.gender.asc())
        ).mappings().all()

    gender_rows = db.session.execute(
        sa.select(
            matched_rows.c.gender,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(matched_rows.c.gender)
        .order_by(sa.func.count().desc(), matched_rows.c.gender.asc().nullslast())
    ).mappings().all()

    top_causes = []
    for index, row in enumerate(top_cause_rows, start=1):
        category = row["bucket_category"]
        subcategory = row["bucket_subcategory"]
        field = row["bucket_field"]
        label_parts = [part for part in (category, subcategory, field) if part]
        coded_count = int(row["coded_count"] or 0)
        top_causes.append(
            {
                "rank": index,
                "bucket_category": category,
                "bucket_subcategory": subcategory,
                "bucket_field": field,
                "display_label": " / ".join(label_parts) if label_parts else field,
                "coded_count": coded_count,
                "percent": round((coded_count / matched_total) * 100, 1),
            }
        )

    first_level_counts = []
    for row in first_level_rows:
        coded_count = int(row["coded_count"] or 0)
        first_level_counts.append(
            {
                "label": row["first_level_label"],
                "coded_count": coded_count,
                "percent": round((coded_count / matched_total) * 100, 1),
            }
        )

    age_sex_buckets = {}
    for row in age_band_rows:
        key = row["age_band_key"] or row["age_band"]
        if key not in age_sex_buckets:
            age_sex_buckets[key] = {
                "age_band": row["age_band"] or row["age_band_key"] or "Unknown",
                "sort_order": int(row["sort_order"] or 0),
                "male_count": 0,
                "female_count": 0,
                "unknown_count": 0,
            }
        bucket = age_sex_buckets[key]
        coded_count = int(row["coded_count"] or 0)
        gender_key = _normalize_gender_label(row["gender"])
        if gender_key == "Male":
            bucket["male_count"] += coded_count
        elif gender_key == "Female":
            bucket["female_count"] += coded_count
        else:
            bucket["unknown_count"] += coded_count

    age_sex_distribution = []
    for row in sorted(
        age_sex_buckets.values(),
        key=lambda item: (item["sort_order"], item["age_band"]),
    ):
        total_count = row["male_count"] + row["female_count"] + row["unknown_count"]
        age_sex_distribution.append(
            {
                "age_band": row["age_band"],
                "sort_order": row["sort_order"],
                "male_count": row["male_count"],
                "female_count": row["female_count"],
                "unknown_count": row["unknown_count"],
                "total_count": total_count,
                "male_percent": round((row["male_count"] / matched_total) * 100, 1),
                "female_percent": round((row["female_count"] / matched_total) * 100, 1),
                "unknown_percent": round((row["unknown_count"] / matched_total) * 100, 1),
                "total_percent": round((total_count / matched_total) * 100, 1),
            }
        )

    if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
        age_filter_key_expr = sa.func.coalesce(
            matched_rows.c.age_scope,
            sa.literal("__none__"),
        ).label("age_filter_key")
        age_filter_label_expr = sa.func.coalesce(
            matched_rows.c.age_scope_label,
            sa.literal("All Ages"),
        ).label("age_filter_label")
        age_filter_sort_expr = sa.func.coalesce(
            matched_rows.c.age_scope_sort_order,
            sa.literal(999),
        ).label("age_filter_sort_order")
    else:
        age_filter_key_expr = _reporting_age_band_sql(
            matched_rows.c.age_normalized_days
        ).label("age_filter_key")
        age_filter_label_expr = age_filter_key_expr.label("age_filter_label")
        age_filter_sort_expr = _reporting_age_band_sort_sql(
            matched_rows.c.age_normalized_days
        ).label("age_filter_sort_order")

    age_filter_rows = db.session.execute(
        sa.select(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
        )
        .select_from(matched_rows)
        .group_by(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
        )
        .order_by(age_filter_sort_expr.asc(), age_filter_label_expr.asc())
    ).mappings().all()
    age_filters = [{"key": "all", "label": "All ages"}] + [
        {
            "key": row["age_filter_key"],
            "label": row["age_filter_label"],
        }
        for row in age_filter_rows
    ]

    top_cause_by_age_rows = db.session.execute(
        sa.select(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
        )
        .order_by(
            age_filter_sort_expr.asc(),
            sa.func.count().desc(),
            matched_rows.c.bucket_field.asc(),
            matched_rows.c.bucket_category.asc().nullslast(),
            matched_rows.c.bucket_subcategory.asc().nullslast(),
        )
    ).mappings().all()

    first_level_by_age_rows = db.session.execute(
        sa.select(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
            first_level_label,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(
            age_filter_key_expr,
            age_filter_label_expr,
            age_filter_sort_expr,
            first_level_label,
        )
        .order_by(
            age_filter_sort_expr.asc(),
            sa.func.count().desc(),
            first_level_label.asc(),
        )
    ).mappings().all()

    top_causes_by_age = {"all": top_causes}
    age_totals_by_key = {}
    for row in top_cause_by_age_rows:
        age_key = row["age_filter_key"]
        age_totals_by_key[age_key] = age_totals_by_key.get(age_key, 0) + int(
            row["coded_count"] or 0
        )
    grouped_top_by_age = {}
    for row in top_cause_by_age_rows:
        grouped_top_by_age.setdefault(row["age_filter_key"], []).append(row)
    for age_key, rows_for_age in grouped_top_by_age.items():
        total_for_age = age_totals_by_key.get(age_key, 0) or 1
        top_causes_by_age[age_key] = []
        for index, row in enumerate(rows_for_age[:top_n], start=1):
            category = row["bucket_category"]
            subcategory = row["bucket_subcategory"]
            field = row["bucket_field"]
            label_parts = [part for part in (category, subcategory, field) if part]
            coded_count = int(row["coded_count"] or 0)
            top_causes_by_age[age_key].append(
                {
                    "rank": index,
                    "bucket_category": category,
                    "bucket_subcategory": subcategory,
                    "bucket_field": field,
                    "display_label": " / ".join(label_parts) if label_parts else field,
                    "coded_count": coded_count,
                    "percent": round((coded_count / total_for_age) * 100, 1),
                }
            )

    first_level_counts_by_age = {"all": first_level_counts}
    age_first_level_totals = {}
    for row in first_level_by_age_rows:
        age_key = row["age_filter_key"]
        age_first_level_totals[age_key] = age_first_level_totals.get(age_key, 0) + int(
            row["coded_count"] or 0
        )
    grouped_first_level_by_age = {}
    for row in first_level_by_age_rows:
        grouped_first_level_by_age.setdefault(row["age_filter_key"], []).append(row)
    for age_key, rows_for_age in grouped_first_level_by_age.items():
        total_for_age = age_first_level_totals.get(age_key, 0) or 1
        first_level_counts_by_age[age_key] = [
            {
                "label": row["first_level_label"],
                "coded_count": int(row["coded_count"] or 0),
                "percent": round((int(row["coded_count"] or 0) / total_for_age) * 100, 1),
            }
            for row in rows_for_age
        ]

    gender_distribution = []
    for row in gender_rows:
        coded_count = int(row["coded_count"] or 0)
        gender_label = _normalize_gender_label(row["gender"])
        gender_distribution.append(
            {
                "gender": gender_label,
                "coded_count": coded_count,
                "percent": round((coded_count / matched_total) * 100, 1),
            }
        )

    def _serialize_heatmap_dimensions(
        rows,
        *,
        ordered_labels: list[str],
        row_label_key: str,
        display_lookup: dict[str, str],
    ) -> dict[str, list[dict]]:
        heatmap_dimensions = {
            "country": {},
            "year": {},
            "sex": {},
            "age": {},
        }
        for row in rows:
            row_key = row[row_label_key]
            coded_count = int(row["coded_count"] or 0)
            country_label = row["country"] or "Unknown"
            year_label = str(row["submission_year"] or "Unknown")
            sex_label = _normalize_gender_label(row["gender"])
            if scheme.scheme_code == SCHEME_CODE_SRS_INDIA:
                age_label = row["age_scope_label"] or "Unknown"
                age_sort = int(row["age_scope_sort_order"] or 0)
            else:
                age_label = row["reporting_age_band"]
                age_sort = int(row["reporting_age_band_sort_order"] or 0)

            for dimension, label, sort_order in (
                ("country", country_label, None),
                ("year", year_label, None),
                ("sex", sex_label, None),
                ("age", age_label, age_sort),
            ):
                matrix = heatmap_dimensions[dimension]
                if label not in matrix:
                    matrix[label] = {
                        "label": label,
                        "sort_order": sort_order,
                        "values": {},
                    }
                matrix[label]["values"][row_key] = matrix[label]["values"].get(row_key, 0) + coded_count

        serialized = {}
        for dimension, value_map in heatmap_dimensions.items():
            if dimension == "age":
                ordered_values = sorted(
                    value_map.values(),
                    key=lambda item: (
                        item["sort_order"] if item["sort_order"] is not None else 999,
                        item["label"],
                    ),
                )
            else:
                ordered_values = sorted(value_map.values(), key=lambda item: item["label"])
            serialized[dimension] = [
                {
                    "label": item["label"],
                    "values": [
                        {
                            "cause": label,
                            "display_label": display_lookup.get(label, label),
                            "coded_count": item["values"].get(label, 0),
                            "percent": round((item["values"].get(label, 0) / matched_total) * 100, 1),
                        }
                        for label in ordered_labels
                    ],
                }
                for item in ordered_values
            ]
        return serialized

    top_cause_labels = [row["bucket_field"] for row in top_cause_rows]
    top_cause_heatmap_base = db.session.execute(
        sa.select(
            matched_rows.c.bucket_field,
            matched_rows.c.country,
            matched_rows.c.submission_year,
            matched_rows.c.gender,
            matched_rows.c.age_scope_label,
            matched_rows.c.age_scope_sort_order,
            _reporting_age_band_sql(matched_rows.c.age_normalized_days).label("reporting_age_band"),
            _reporting_age_band_sort_sql(matched_rows.c.age_normalized_days).label("reporting_age_band_sort_order"),
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .where(matched_rows.c.bucket_field.in_(top_cause_labels))
        .group_by(
            matched_rows.c.bucket_field,
            matched_rows.c.country,
            matched_rows.c.submission_year,
            matched_rows.c.gender,
            matched_rows.c.age_scope_label,
            matched_rows.c.age_scope_sort_order,
            _reporting_age_band_sql(matched_rows.c.age_normalized_days),
            _reporting_age_band_sort_sql(matched_rows.c.age_normalized_days),
        )
    ).mappings().all() if top_cause_labels else []

    top_cause_lookup = {row["bucket_field"]: row["display_label"] for row in top_causes}
    serialized_top_cause_heatmap = _serialize_heatmap_dimensions(
        top_cause_heatmap_base,
        ordered_labels=top_cause_labels,
        row_label_key="bucket_field",
        display_lookup=top_cause_lookup,
    )

    first_level_labels = [row["label"] for row in first_level_counts]
    first_level_heatmap_base = db.session.execute(
        sa.select(
            first_level_label.label("first_level_label"),
            matched_rows.c.country,
            matched_rows.c.submission_year,
            matched_rows.c.gender,
            matched_rows.c.age_scope_label,
            matched_rows.c.age_scope_sort_order,
            _reporting_age_band_sql(matched_rows.c.age_normalized_days).label("reporting_age_band"),
            _reporting_age_band_sort_sql(matched_rows.c.age_normalized_days).label("reporting_age_band_sort_order"),
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .where(first_level_label.in_(first_level_labels))
        .group_by(
            first_level_label,
            matched_rows.c.country,
            matched_rows.c.submission_year,
            matched_rows.c.gender,
            matched_rows.c.age_scope_label,
            matched_rows.c.age_scope_sort_order,
            _reporting_age_band_sql(matched_rows.c.age_normalized_days),
            _reporting_age_band_sort_sql(matched_rows.c.age_normalized_days),
        )
    ).mappings().all() if first_level_labels else []
    first_level_lookup = {row["label"]: row["label"] for row in first_level_counts}
    serialized_first_level_heatmap = _serialize_heatmap_dimensions(
        first_level_heatmap_base,
        ordered_labels=first_level_labels,
        row_label_key="first_level_label",
        display_lookup=first_level_lookup,
    )

    treemap_rows = db.session.execute(
        sa.select(
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
            sa.func.count().label("coded_count"),
        )
        .select_from(matched_rows)
        .group_by(
            matched_rows.c.bucket_category,
            matched_rows.c.bucket_subcategory,
            matched_rows.c.bucket_field,
        )
        .order_by(sa.func.count().desc(), matched_rows.c.bucket_field.asc())
    ).mappings().all()
    treemap = [
        {
            "group": row["bucket_field"] if scheme.scheme_code == SCHEME_CODE_CMEA10 else (row["bucket_category"] or row["bucket_field"]),
            "subcategory": row["bucket_subcategory"],
            "label": row["bucket_field"],
            "coded_count": int(row["coded_count"] or 0),
            "percent": round((int(row["coded_count"] or 0) / matched_total) * 100, 1),
        }
        for row in treemap_rows
    ]

    return {
        "scheme_used": scheme.scheme_name,
        "top_causes": top_causes,
        "top_causes_by_age": top_causes_by_age,
        "first_level_counts": first_level_counts,
        "first_level_counts_by_age": first_level_counts_by_age,
        "age_filters": age_filters,
        "age_sex_distribution": age_sex_distribution,
        "gender_distribution": gender_distribution,
        "heatmap": {
            "dimension": "country",
            "view": "top_causes",
            "views": {
                "top_causes": serialized_top_cause_heatmap,
                "first_level_counts": serialized_first_level_heatmap,
            },
        },
        "treemap": treemap,
        "matched_total": matched_total,
    }


def export_cod_bucket_reporting_csv(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
) -> str:
    """Return a CSV export for the current COD bucket reporting scope."""
    if allowed_project_site_pairs is not None and not allowed_project_site_pairs:
        return ""

    scheme, base_rows = _cod_bucket_aggregate_base_subquery(
        scheme_code=scheme_code,
        project_id=project_id,
        site_id=site_id,
        form_id=form_id,
        gender=gender,
        submission_date_from=submission_date_from,
        submission_date_to=submission_date_to,
        allowed_project_site_pairs=allowed_project_site_pairs,
    )

    query = (
        sa.select(
            base_rows.c.va_sid,
            base_rows.c.project_id,
            base_rows.c.site_id,
            base_rows.c.form_id,
            base_rows.c.submission_year,
            base_rows.c.final_icd,
            base_rows.c.reporting_icd,
            base_rows.c.gender,
            base_rows.c.age_scope_label,
            base_rows.c.age_normalized_days,
            base_rows.c.age_normalized_years,
            MapIcdCodBucket.node_id,
        )
        .select_from(base_rows)
        .outerjoin(
            MapIcdCodBucket,
            sa.and_(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == base_rows.c.reporting_icd,
                sa.or_(
                    sa.and_(
                        MapIcdCodBucket.age_scope.is_(None),
                        base_rows.c.age_scope.is_(None),
                    ),
                    MapIcdCodBucket.age_scope == base_rows.c.age_scope,
                ),
            ),
        )
        .order_by(
            base_rows.c.project_id.asc(),
            base_rows.c.site_id.asc(),
            base_rows.c.form_id.asc(),
            base_rows.c.submission_year.asc(),
            base_rows.c.va_sid.asc(),
        )
    )
    rows = [dict(row) for row in db.session.execute(query).mappings().all()]

    node_rows = db.session.execute(
        sa.select(
            MasCodBucketNode.node_id,
            MasCodBucketNode.parent_node_id,
            MasCodBucketNode.node_label,
        ).where(
            MasCodBucketNode.scheme_id == scheme.scheme_id,
            MasCodBucketNode.is_active.is_(True),
        )
    ).mappings().all()
    nodes_by_id = {row["node_id"]: dict(row) for row in node_rows}
    path_cache: dict[uuid.UUID, list[str]] = {}

    def _node_path(node_id):
        if node_id is None:
            return []
        if node_id in path_cache:
            return path_cache[node_id]
        labels = []
        current_id = node_id
        while current_id is not None:
            node = nodes_by_id.get(current_id)
            if node is None:
                break
            labels.append(node["node_label"])
            current_id = node["parent_node_id"]
        path = list(reversed(labels))
        path_cache[node_id] = path
        return path

    max_levels = max((len(_node_path(row["node_id"])) for row in rows), default=0)
    headers = [
        "SID",
        "Age",
        "Age group",
        "Sex",
        "Final authoritative COD",
        *[f"Scheme level {index}" for index in range(1, max_levels + 1)],
        "Project",
        "Site",
        "Form code",
        "Year",
    ]

    handle = StringIO()
    writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        path = _node_path(row["node_id"])
        age_group = (
            row["age_scope_label"]
            if scheme.scheme_code == SCHEME_CODE_SRS_INDIA
            else _reporting_age_band_label_for_days(row["age_normalized_days"])
        )
        export_row = {
            "SID": row["va_sid"],
            "Age": row["age_normalized_years"],
            "Age group": age_group,
            "Sex": _normalize_gender_label(row["gender"]),
            "Final authoritative COD": row["final_icd"],
            "Project": row["project_id"],
            "Site": row["site_id"],
            "Form code": row["form_id"],
            "Year": int(row["submission_year"]) if row["submission_year"] is not None else "",
        }
        for index in range(max_levels):
            export_row[f"Scheme level {index + 1}"] = path[index] if index < len(path) else ""
        writer.writerow(export_row)
    return handle.getvalue()


def _cod_bucket_aggregate_base_subquery(
    *,
    scheme_code: str,
    project_id: str | None = None,
    site_id: str | None = None,
    form_id: str | None = None,
    gender: str | None = None,
    submission_date_from=None,
    submission_date_to=None,
    allowed_project_site_pairs: set[tuple[str, str]] | None = None,
):
    """Return the filtered coded submission set joined to scheme age bands."""
    scheme = db.session.scalar(
        sa.select(MasCodBucketScheme).where(
            MasCodBucketScheme.scheme_code == scheme_code,
            MasCodBucketScheme.is_active.is_(True),
        )
    )
    if scheme is None:
        raise LookupError(f"Unknown active COD bucket scheme: {scheme_code}")

    core = sa.table(
        CORE_MV_NAME,
        sa.column("va_sid"),
        sa.column("project_id"),
        sa.column("site_id"),
        sa.column("submission_date"),
    )
    demo = sa.table(
        DEMOGRAPHICS_MV_NAME,
        sa.column("va_sid"),
        sa.column("analytics_age_band"),
        sa.column("analytics_age_normalized_days"),
        sa.column("has_human_final_cod"),
    )
    cod = sa.table(
        COD_MV_NAME,
        sa.column("va_sid"),
        sa.column("final_icd"),
    )
    reporting_alias = sa.orm.aliased(MapIcd10LegacyReportingAlias)

    submission_age_days = sa.cast(demo.c.analytics_age_normalized_days, sa.Numeric())
    age_band = sa.orm.aliased(MasCodBucketSchemeAgeBand)
    gender_filter = _normalize_gender_filter(gender)

    conditions = [
        cod.c.final_icd.is_not(None),
        demo.c.has_human_final_cod.is_(True),
        _gender_filter_clause(VaSubmissions.va_deceased_gender, gender_filter),
    ]
    if project_id:
        conditions.append(core.c.project_id == project_id)
    if site_id:
        conditions.append(core.c.site_id == site_id)
    if form_id:
        conditions.append(VaSubmissions.va_form_id == form_id)
    if submission_date_from:
        conditions.append(core.c.submission_date >= submission_date_from)
    if submission_date_to:
        conditions.append(core.c.submission_date <= submission_date_to)
    if allowed_project_site_pairs is not None:
        conditions.append(
            sa.tuple_(core.c.project_id, core.c.site_id).in_(list(allowed_project_site_pairs))
        )

    query = (
        sa.select(
            core.c.va_sid.label("va_sid"),
            core.c.project_id.label("project_id"),
            core.c.site_id.label("site_id"),
            VaSubmissions.va_form_id.label("form_id"),
            age_band.age_scope.label("age_scope"),
            age_band.age_label.label("age_scope_label"),
            age_band.sort_order.label("age_scope_sort_order"),
            submission_age_days.label("age_normalized_days"),
            VaSubmissions.va_deceased_age_normalized_years.label("age_normalized_years"),
            VaSubmissions.va_deceased_gender.label("gender"),
            VaForms.form_smartvacountry.label("country"),
            sa.extract("year", core.c.submission_date).label("submission_year"),
            cod.c.final_icd.label("final_icd"),
            _reporting_icd_sql(
                cod.c.final_icd,
                reporting_alias.reporting_code,
            ).label("reporting_icd"),
        )
        .select_from(core)
        .join(demo, demo.c.va_sid == core.c.va_sid)
        .join(cod, cod.c.va_sid == core.c.va_sid)
        .join(VaSubmissions, VaSubmissions.va_sid == core.c.va_sid)
        .join(VaForms, VaForms.form_id == VaSubmissions.va_form_id)
        .outerjoin(reporting_alias, reporting_alias.legacy_code == cod.c.final_icd)
        .join(
            age_band,
            sa.and_(
                age_band.scheme_id == scheme.scheme_id,
                age_band.is_active.is_(True),
                submission_age_days >= _age_bound_days_sql(
                    age_band.min_age_value,
                    age_band.min_age_unit,
                ),
                submission_age_days < _age_bound_days_sql(
                    age_band.max_age_value,
                    age_band.max_age_unit,
                ),
            ),
        )
        .where(sa.and_(*conditions))
    )
    return scheme, query.subquery("cod_bucket_base")
