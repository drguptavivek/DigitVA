"""rebuild icd10 and cod bucket master data

Revision ID: d6e7f8a9b0c1
Revises: c1d2e3f4a5b6
Create Date: 2026-04-27T22:00:00.000000
"""

from __future__ import annotations

import csv
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path

from alembic import op
from openpyxl import load_workbook
import sqlalchemy as sa


revision = "d6e7f8a9b0c1"
down_revision = "c1d2e3f4a5b6"
branch_labels = None
depends_on = None


ICD_CSV_PATH = Path(
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "icd10-2019-base-2026-04-27/icd10_2019_hierarchy.csv"
)
WHO_POLICY_PATH = Path(
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "who-2022-va-icd-cod-2026-04-27/who_2022_icd10_2019_2_policy_reviewed.json"
)
SRS_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "srs-india-cod-2026-04-27/icd-10-CODES_SRS_India.xlsx"
)
CMEA10_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "cmea10-cod-2026-04-27/icd-10-CODES_CMEA10_mapped.xlsx"
)
WHO_COD_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/migration-artifacts/"
    "who-2022-va-icd-cod-2026-04-27/WHO_2022_VA_Bucket_Mapping_document_derived.xlsx"
)

SOURCE_VERSION = "ICD-10-2019"
CHUNK_SIZE = 1000

SCHEME_CODE_SRS_INDIA = "SRS_INDIA"
SCHEME_CODE_CMEA10 = "CMEA10"
SCHEME_CODE_WHO_2022_VA = "WHO_2022_VA"

AGE_SCOPE_ADULT_OVER5Y = "adult_over5y"
AGE_SCOPE_CHILD_1_59M = "child_1_59m"
AGE_SCOPE_NEONATE = "neonate"

AGE_UNIT_DAYS = "days"
AGE_UNIT_MONTHS = "months"
AGE_UNIT_YEARS = "years"

NODE_TYPE_CATEGORY = "category"
NODE_TYPE_SUBCATEGORY = "subcategory"
NODE_TYPE_FIELD = "field"

SRS_SCOPE_CONFIG = (
    (
        AGE_SCOPE_ADULT_OVER5Y,
        "SRS_India_over5y_main_group",
        "SRS_India_over5y_sub_group",
        "SRS_India_over5y_disease",
        "SRS_India_over5y_match_type",
        "SRS_India_over5y_note",
    ),
    (
        AGE_SCOPE_NEONATE,
        "SRS_India_neonate_main_group",
        "SRS_India_neonate_sub_group",
        "SRS_India_neonate_disease",
        "SRS_India_neonate_match_type",
        "SRS_India_neonate_note",
    ),
    (
        AGE_SCOPE_CHILD_1_59M,
        "SRS_India_1_59mth_main_group",
        "SRS_India_1_59mth_sub_group",
        "SRS_India_1_59mth_disease",
        "SRS_India_1_59mth_type",
        "SRS_India_1_59mth_note",
    ),
)


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _artifact_path(relative_path: Path) -> Path:
    path = _repo_root() / relative_path
    if not path.exists():
        raise ValueError(f"Migration artifact not found: {path}")
    return path


def _optional_text(value):
    cleaned = str(value or "").strip()
    return cleaned or None


def _normalize_label(value) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value).strip())
    return normalized or None


def _normalize_icd_code(value) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip().upper()
    return normalized or None


def _slugify(value: str, *, fallback_prefix: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or fallback_prefix


def _parse_bool(value):
    normalized = (value or "").strip().lower()
    if not normalized:
        return None
    if normalized == "true":
        return True
    if normalized == "false":
        return False
    raise ValueError(f"Invalid boolean value in ICD seed CSV: {value!r}")


def _table(name: str, *columns) -> sa.Table:
    return sa.table(name, *columns)


ICD_TABLE = _table(
    "mas_icd10_2019_2",
    sa.column("code", sa.String(length=16)),
    sa.column("title", sa.Text()),
    sa.column("node_type", sa.String(length=16)),
    sa.column("semantic_level", sa.String(length=32)),
    sa.column("sort_order", sa.Integer()),
    sa.column("parent_code", sa.String(length=16)),
    sa.column("chapter_code", sa.String(length=16)),
    sa.column("chapter_title", sa.Text()),
    sa.column("block_code", sa.String(length=16)),
    sa.column("block_title", sa.Text()),
    sa.column("three_character_code", sa.String(length=16)),
    sa.column("three_character_title", sa.Text()),
    sa.column("has_children", sa.Boolean()),
    sa.column("is_leaf", sa.Boolean()),
    sa.column("is_three_character_code", sa.Boolean()),
    sa.column("is_detailed_code", sa.Boolean()),
    sa.column("is_coding_selectable", sa.Boolean()),
    sa.column("sex_selectable", sa.String(length=16)),
    sa.column("age_group_selectable", sa.String(length=32)),
    sa.column("policy_status", sa.String(length=32)),
    sa.column("restriction_note", sa.Text()),
    sa.column("source_version", sa.String(length=32)),
    sa.column("source_path", sa.String(length=512)),
    sa.column("is_active", sa.Boolean()),
    sa.column("created_at", sa.DateTime(timezone=True)),
    sa.column("updated_at", sa.DateTime(timezone=True)),
)

SCHEME_TABLE = _table(
    "mas_cod_bucket_schemes",
    sa.column("scheme_id", sa.Uuid(as_uuid=True)),
    sa.column("scheme_code", sa.String(length=32)),
    sa.column("scheme_name", sa.String(length=128)),
    sa.column("scheme_description", sa.Text()),
    sa.column("source_path", sa.String(length=512)),
    sa.column("mapping_version", sa.Integer()),
    sa.column("is_active", sa.Boolean()),
    sa.column("created_at", sa.DateTime(timezone=True)),
    sa.column("updated_at", sa.DateTime(timezone=True)),
)

AGE_BAND_TABLE = _table(
    "mas_cod_bucket_scheme_age_bands",
    sa.column("age_band_id", sa.Uuid(as_uuid=True)),
    sa.column("scheme_id", sa.Uuid(as_uuid=True)),
    sa.column("age_scope", sa.String(length=32)),
    sa.column("age_label", sa.String(length=128)),
    sa.column("min_age_value", sa.Integer()),
    sa.column("min_age_unit", sa.String(length=8)),
    sa.column("max_age_value", sa.Integer()),
    sa.column("max_age_unit", sa.String(length=8)),
    sa.column("level_count", sa.Integer()),
    sa.column("sort_order", sa.Integer()),
    sa.column("is_active", sa.Boolean()),
    sa.column("created_at", sa.DateTime(timezone=True)),
    sa.column("updated_at", sa.DateTime(timezone=True)),
)

NODE_TABLE = _table(
    "mas_cod_bucket_nodes",
    sa.column("node_id", sa.Uuid(as_uuid=True)),
    sa.column("scheme_id", sa.Uuid(as_uuid=True)),
    sa.column("age_scope", sa.String(length=32)),
    sa.column("node_type", sa.String(length=16)),
    sa.column("parent_node_id", sa.Uuid(as_uuid=True)),
    sa.column("node_code", sa.String(length=128)),
    sa.column("node_label", sa.String(length=256)),
    sa.column("sort_order", sa.Integer()),
    sa.column("is_active", sa.Boolean()),
    sa.column("created_at", sa.DateTime(timezone=True)),
    sa.column("updated_at", sa.DateTime(timezone=True)),
)

MAPPING_TABLE = _table(
    "map_icd_cod_buckets",
    sa.column("mapping_id", sa.Uuid(as_uuid=True)),
    sa.column("scheme_id", sa.Uuid(as_uuid=True)),
    sa.column("age_scope", sa.String(length=32)),
    sa.column("icd_code", sa.String(length=16)),
    sa.column("node_id", sa.Uuid(as_uuid=True)),
    sa.column("source_sheet", sa.String(length=128)),
    sa.column("source_row_number", sa.Integer()),
    sa.column("source_category", sa.String(length=256)),
    sa.column("match_type", sa.String(length=32)),
    sa.column("mapping_note", sa.Text()),
    sa.column("is_active", sa.Boolean()),
    sa.column("created_at", sa.DateTime(timezone=True)),
    sa.column("updated_at", sa.DateTime(timezone=True)),
)


def _load_sheet_rows(workbook_path: Path, sheet_name: str):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(header).strip() if header is not None else "" for header in next(rows)]
        for row_number, row in enumerate(rows, start=2):
            yield row_number, dict(zip(headers, row, strict=False))
    finally:
        workbook.close()


def _insert_chunks(bind, table, rows: list[dict], *, chunk_size: int = CHUNK_SIZE) -> None:
    for start in range(0, len(rows), chunk_size):
        bind.execute(table.insert(), rows[start : start + chunk_size])


def _rebuild_icd_base(bind, now: datetime) -> None:
    csv_path = _artifact_path(ICD_CSV_PATH)
    bind.execute(sa.text("TRUNCATE TABLE mas_icd10_2019_2"))

    seed_rows = []
    with csv_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for sort_order, row in enumerate(reader, start=1):
            code = (row.get("code") or "").strip()
            if not code:
                raise ValueError("ICD seed CSV contains a row with empty code")
            seed_rows.append(
                {
                    "code": code,
                    "title": (row.get("title") or "").strip(),
                    "node_type": (row.get("node_type") or "").strip(),
                    "semantic_level": (row.get("semantic_level") or "").strip(),
                    "sort_order": sort_order,
                    "parent_code": _optional_text(row.get("parent_code")),
                    "chapter_code": _optional_text(row.get("chapter_code")),
                    "chapter_title": _optional_text(row.get("chapter_title")),
                    "block_code": _optional_text(row.get("block_code")),
                    "block_title": _optional_text(row.get("block_title")),
                    "three_character_code": _optional_text(row.get("three_character_code")),
                    "three_character_title": _optional_text(row.get("three_character_title")),
                    "has_children": bool(_parse_bool(row.get("has_children"))),
                    "is_leaf": bool(_parse_bool(row.get("is_leaf"))),
                    "is_three_character_code": bool(_parse_bool(row.get("is_three_character_code"))),
                    "is_detailed_code": bool(_parse_bool(row.get("is_detailed_code"))),
                    "is_coding_selectable": None,
                    "sex_selectable": None,
                    "age_group_selectable": None,
                    "policy_status": "unreviewed",
                    "restriction_note": None,
                    "source_version": SOURCE_VERSION,
                    "source_path": str(ICD_CSV_PATH),
                    "is_active": True,
                    "created_at": now,
                    "updated_at": now,
                }
            )
    _insert_chunks(bind, ICD_TABLE, seed_rows)


def _apply_reviewed_who_policy(bind) -> None:
    policy_path = _artifact_path(WHO_POLICY_PATH)
    payload = json.loads(policy_path.read_text(encoding="utf-8"))
    items = payload.get("items") or []

    bind.execute(
        sa.text(
            """
            UPDATE mas_icd10_2019_2
            SET is_coding_selectable = FALSE,
                sex_selectable = NULL,
                age_group_selectable = NULL,
                policy_status = 'reviewed',
                restriction_note = NULL
            WHERE is_active IS TRUE
              AND semantic_level IN ('three_character', 'detailed_code')
            """
        )
    )

    update_stmt = sa.text(
        """
        UPDATE mas_icd10_2019_2
        SET is_coding_selectable = :is_coding_selectable,
            sex_selectable = :sex_selectable,
            age_group_selectable = :age_group_selectable,
            policy_status = 'reviewed',
            restriction_note = :restriction_note
        WHERE code = :code
          AND is_active IS TRUE
          AND semantic_level IN ('three_character', 'detailed_code')
        """
    )
    bind.execute(
        update_stmt,
        [
            {
                "code": item["code"],
                "is_coding_selectable": bool(item.get("is_coding_selectable")),
                "sex_selectable": item.get("sex_selectable"),
                "age_group_selectable": item.get("age_group_selectable"),
                "restriction_note": item.get("restriction_note"),
            }
            for item in items
        ],
    )


def _builtin_age_band_metadata(scheme_code: str, age_scope: str | None) -> dict:
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_ADULT_OVER5Y:
        return {
            "age_label": "Adult / Over 5 Years",
            "min_age_value": 5,
            "min_age_unit": AGE_UNIT_YEARS,
            "max_age_value": 120,
            "max_age_unit": AGE_UNIT_YEARS,
            "level_count": 3,
            "sort_order": 1,
        }
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_CHILD_1_59M:
        return {
            "age_label": "Child / 1-59 Months",
            "min_age_value": 1,
            "min_age_unit": AGE_UNIT_MONTHS,
            "max_age_value": 60,
            "max_age_unit": AGE_UNIT_MONTHS,
            "level_count": 3,
            "sort_order": 2,
        }
    if scheme_code == SCHEME_CODE_SRS_INDIA and age_scope == AGE_SCOPE_NEONATE:
        return {
            "age_label": "Neonate",
            "min_age_value": 0,
            "min_age_unit": AGE_UNIT_DAYS,
            "max_age_value": 29,
            "max_age_unit": AGE_UNIT_DAYS,
            "level_count": 3,
            "sort_order": 3,
        }
    if scheme_code == SCHEME_CODE_CMEA10 and age_scope is None:
        return {
            "age_label": "All Ages",
            "min_age_value": 0,
            "min_age_unit": AGE_UNIT_DAYS,
            "max_age_value": 120,
            "max_age_unit": AGE_UNIT_YEARS,
            "level_count": 1,
            "sort_order": 1,
        }
    if scheme_code == SCHEME_CODE_WHO_2022_VA and age_scope is None:
        return {
            "age_label": "All Ages",
            "min_age_value": 0,
            "min_age_unit": AGE_UNIT_DAYS,
            "max_age_value": 120,
            "max_age_unit": AGE_UNIT_YEARS,
            "level_count": 2,
            "sort_order": 1,
        }
    raise ValueError(f"Unsupported built-in scheme/age scope: {scheme_code} / {age_scope}")


def _get_or_create_scheme(
    bind,
    *,
    scheme_code: str,
    scheme_name: str,
    scheme_description: str,
    source_path: Path,
    now: datetime,
) -> uuid.UUID:
    existing = bind.execute(
        sa.select(SCHEME_TABLE.c.scheme_id).where(SCHEME_TABLE.c.scheme_code == scheme_code)
    ).scalar_one_or_none()
    if existing is not None:
        bind.execute(
            SCHEME_TABLE.update()
            .where(SCHEME_TABLE.c.scheme_id == existing)
            .values(
                scheme_name=scheme_name,
                scheme_description=scheme_description,
                source_path=str(source_path),
                mapping_version=1,
                is_active=True,
                updated_at=now,
            )
        )
        return existing

    scheme_id = uuid.uuid4()
    bind.execute(
        SCHEME_TABLE.insert().values(
            scheme_id=scheme_id,
            scheme_code=scheme_code,
            scheme_name=scheme_name,
            scheme_description=scheme_description,
            source_path=str(source_path),
            mapping_version=1,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
    )
    return scheme_id


def _replace_scheme_contents(bind, scheme_id: uuid.UUID) -> None:
    bind.execute(MAPPING_TABLE.delete().where(MAPPING_TABLE.c.scheme_id == scheme_id))
    bind.execute(NODE_TABLE.delete().where(NODE_TABLE.c.scheme_id == scheme_id))
    bind.execute(AGE_BAND_TABLE.delete().where(AGE_BAND_TABLE.c.scheme_id == scheme_id))


def _create_age_band(bind, *, scheme_id: uuid.UUID, age_scope: str | None, meta: dict, now: datetime) -> uuid.UUID:
    age_band_id = uuid.uuid4()
    bind.execute(
        AGE_BAND_TABLE.insert().values(
            age_band_id=age_band_id,
            scheme_id=scheme_id,
            age_scope=age_scope,
            age_label=meta["age_label"],
            min_age_value=meta["min_age_value"],
            min_age_unit=meta["min_age_unit"],
            max_age_value=meta["max_age_value"],
            max_age_unit=meta["max_age_unit"],
            level_count=meta["level_count"],
            sort_order=meta["sort_order"],
            is_active=True,
            created_at=now,
            updated_at=now,
        )
    )
    return age_band_id


def _create_node(
    bind,
    *,
    scheme_id: uuid.UUID,
    age_scope: str | None,
    node_type: str,
    node_label: str,
    sort_order: int,
    now: datetime,
    parent_node_id: uuid.UUID | None = None,
    node_code_suffix: str | None = None,
) -> uuid.UUID:
    node_id = uuid.uuid4()
    bind.execute(
        NODE_TABLE.insert().values(
            node_id=node_id,
            scheme_id=scheme_id,
            age_scope=age_scope,
            node_type=node_type,
            parent_node_id=parent_node_id,
            node_code=_slugify(node_code_suffix or node_label, fallback_prefix=node_type),
            node_label=node_label,
            sort_order=sort_order,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
    )
    return node_id


def _insert_mapping(
    bind,
    *,
    scheme_id: uuid.UUID,
    age_scope: str | None,
    icd_code: str,
    node_id: uuid.UUID,
    source_sheet: str,
    source_row_number: int,
    source_category: str | None,
    match_type: str | None,
    mapping_note: str | None,
    now: datetime,
) -> None:
    bind.execute(
        MAPPING_TABLE.insert().values(
            mapping_id=uuid.uuid4(),
            scheme_id=scheme_id,
            age_scope=age_scope,
            icd_code=icd_code,
            node_id=node_id,
            source_sheet=source_sheet,
            source_row_number=source_row_number,
            source_category=source_category,
            match_type=match_type,
            mapping_note=mapping_note,
            is_active=True,
            created_at=now,
            updated_at=now,
        )
    )


def _load_srs_rows_by_scope(workbook_path: Path) -> dict[str, list[dict]]:
    rows_by_scope = {
        AGE_SCOPE_ADULT_OVER5Y: [],
        AGE_SCOPE_CHILD_1_59M: [],
        AGE_SCOPE_NEONATE: [],
    }
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD_Mapped"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        if not icd_code:
            continue
        for age_scope, main_key, sub_key, disease_key, match_key, note_key in SRS_SCOPE_CONFIG:
            disease = _normalize_label(payload.get(disease_key))
            if not disease:
                continue
            rows_by_scope[age_scope].append(
                {
                    "row_number": row_number,
                    "icd_code": icd_code,
                    "source_category": _normalize_label(payload.get("category")),
                    "main_group": _normalize_label(payload.get(main_key)),
                    "sub_group": _normalize_label(payload.get(sub_key)),
                    "disease": disease,
                    "match_type": _normalize_label(payload.get(match_key)),
                    "mapping_note": _normalize_label(payload.get(note_key)),
                }
            )
    return rows_by_scope


def _populate_srs_scope(bind, *, scheme_id: uuid.UUID, age_scope: str, rows: list[dict], now: datetime) -> None:
    category_nodes: dict[str, uuid.UUID] = {}
    subcategory_nodes: dict[tuple[str, str], uuid.UUID] = {}
    field_nodes: dict[tuple[str, str | None, str], uuid.UUID] = {}
    category_order = 0
    subcategory_order = 0
    field_order = 0

    for row in rows:
        category_label = row["main_group"] or "Unspecified"
        subcategory_label = row["sub_group"]
        field_label = row["disease"]

        if category_label not in category_nodes:
            category_order += 1
            category_nodes[category_label] = _create_node(
                bind,
                scheme_id=scheme_id,
                age_scope=age_scope,
                node_type=NODE_TYPE_CATEGORY,
                node_label=category_label,
                sort_order=category_order,
                now=now,
            )
        parent_node_id = category_nodes[category_label]

        if subcategory_label:
            subcategory_key = (category_label, subcategory_label)
            if subcategory_key not in subcategory_nodes:
                subcategory_order += 1
                subcategory_nodes[subcategory_key] = _create_node(
                    bind,
                    scheme_id=scheme_id,
                    age_scope=age_scope,
                    node_type=NODE_TYPE_SUBCATEGORY,
                    node_label=subcategory_label,
                    sort_order=subcategory_order,
                    parent_node_id=category_nodes[category_label],
                    now=now,
                )
            parent_node_id = subcategory_nodes[subcategory_key]

        field_key = (category_label, subcategory_label, field_label)
        if field_key not in field_nodes:
            field_order += 1
            field_nodes[field_key] = _create_node(
                bind,
                scheme_id=scheme_id,
                age_scope=age_scope,
                node_type=NODE_TYPE_FIELD,
                node_label=field_label,
                sort_order=field_order,
                parent_node_id=parent_node_id,
                now=now,
            )

        _insert_mapping(
            bind,
            scheme_id=scheme_id,
            age_scope=age_scope,
            icd_code=row["icd_code"],
            node_id=field_nodes[field_key],
            source_sheet="ICD_Mapped",
            source_row_number=row["row_number"],
            source_category=row["source_category"],
            match_type=row["match_type"],
            mapping_note=row["mapping_note"],
            now=now,
        )


def _rebuild_srs_india(bind, now: datetime) -> None:
    source_path = SRS_WORKBOOK_PATH
    workbook_path = _artifact_path(source_path)
    scheme_id = _get_or_create_scheme(
        bind,
        scheme_code=SCHEME_CODE_SRS_INDIA,
        scheme_name="SRS India",
        scheme_description="Age-scoped SRS India cause-of-death reporting hierarchy imported from migration artifacts.",
        source_path=source_path,
        now=now,
    )
    _replace_scheme_contents(bind, scheme_id)
    rows_by_scope = _load_srs_rows_by_scope(workbook_path)
    for age_scope in (AGE_SCOPE_ADULT_OVER5Y, AGE_SCOPE_CHILD_1_59M, AGE_SCOPE_NEONATE):
        _create_age_band(
            bind,
            scheme_id=scheme_id,
            age_scope=age_scope,
            meta=_builtin_age_band_metadata(SCHEME_CODE_SRS_INDIA, age_scope),
            now=now,
        )
        _populate_srs_scope(
            bind,
            scheme_id=scheme_id,
            age_scope=age_scope,
            rows=rows_by_scope[age_scope],
            now=now,
        )


def _rebuild_cmea10(bind, now: datetime) -> None:
    source_path = CMEA10_WORKBOOK_PATH
    workbook_path = _artifact_path(source_path)
    scheme_id = _get_or_create_scheme(
        bind,
        scheme_code=SCHEME_CODE_CMEA10,
        scheme_name="CMEA10",
        scheme_description="Flat ICD-10 to CMEA10 cause bucket mapping imported from migration artifacts.",
        source_path=source_path,
        now=now,
    )
    _replace_scheme_contents(bind, scheme_id)
    _create_age_band(
        bind,
        scheme_id=scheme_id,
        age_scope=None,
        meta=_builtin_age_band_metadata(SCHEME_CODE_CMEA10, None),
        now=now,
    )
    field_nodes: dict[str, uuid.UUID] = {}
    field_order = 0
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD10_CMEA10"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        field_label = _normalize_label(payload.get("CMEA10"))
        if not icd_code or not field_label:
            continue
        if field_label not in field_nodes:
            field_order += 1
            field_nodes[field_label] = _create_node(
                bind,
                scheme_id=scheme_id,
                age_scope=None,
                node_type=NODE_TYPE_FIELD,
                node_label=field_label,
                sort_order=field_order,
                now=now,
            )
        _insert_mapping(
            bind,
            scheme_id=scheme_id,
            age_scope=None,
            icd_code=icd_code,
            node_id=field_nodes[field_label],
            source_sheet="ICD10_CMEA10",
            source_row_number=row_number,
            source_category=_normalize_label(payload.get("category")),
            match_type=None,
            mapping_note=None,
            now=now,
        )


def _rebuild_who_2022_va(bind, now: datetime) -> None:
    source_path = WHO_COD_WORKBOOK_PATH
    workbook_path = _artifact_path(source_path)
    scheme_id = _get_or_create_scheme(
        bind,
        scheme_code=SCHEME_CODE_WHO_2022_VA,
        scheme_name="WHO 2022 VA",
        scheme_description="WHO 2022 verbal autopsy COD bucket mapping imported from migration artifacts.",
        source_path=source_path,
        now=now,
    )
    _replace_scheme_contents(bind, scheme_id)
    _create_age_band(
        bind,
        scheme_id=scheme_id,
        age_scope=None,
        meta=_builtin_age_band_metadata(SCHEME_CODE_WHO_2022_VA, None),
        now=now,
    )

    category_nodes: dict[str, uuid.UUID] = {}
    field_nodes: dict[tuple[str, str], uuid.UUID] = {}
    category_order = 0
    field_order = 0
    for row_number, payload in _load_sheet_rows(workbook_path, "ICD_Mapped"):
        icd_code = _normalize_icd_code(payload.get("icd_code"))
        section_label = _normalize_label(payload.get("WHO_2022_VA_section"))
        va_code = _normalize_label(payload.get("WHO_2022_VA_code"))
        va_title = _normalize_label(payload.get("WHO_2022_VA_cause"))
        if not icd_code or not section_label or not va_code or not va_title:
            continue
        if section_label not in category_nodes:
            category_order += 1
            category_nodes[section_label] = _create_node(
                bind,
                scheme_id=scheme_id,
                age_scope=None,
                node_type=NODE_TYPE_CATEGORY,
                node_label=section_label,
                sort_order=category_order,
                now=now,
            )
        field_key = (section_label, va_code)
        if field_key not in field_nodes:
            field_order += 1
            field_nodes[field_key] = _create_node(
                bind,
                scheme_id=scheme_id,
                age_scope=None,
                node_type=NODE_TYPE_FIELD,
                node_label=va_title,
                sort_order=field_order,
                parent_node_id=category_nodes[section_label],
                node_code_suffix=va_code,
                now=now,
            )
        _insert_mapping(
            bind,
            scheme_id=scheme_id,
            age_scope=None,
            icd_code=icd_code,
            node_id=field_nodes[field_key],
            source_sheet="ICD_Mapped",
            source_row_number=row_number,
            source_category=_normalize_label(payload.get("category")),
            match_type=_normalize_label(payload.get("WHO_2022_VA_match_type")),
            mapping_note=_normalize_label(payload.get("WHO_2022_VA_note")),
            now=now,
        )


def upgrade():
    bind = op.get_bind()
    now = datetime.now(timezone.utc)
    _rebuild_icd_base(bind, now)
    _apply_reviewed_who_policy(bind)
    _rebuild_srs_india(bind, now)
    _rebuild_cmea10(bind, now)
    _rebuild_who_2022_va(bind, now)


def downgrade():
    # This migration rebuilds master-data policy state from reviewed artifacts.
    # Downgrading would require restoring a site-specific previous policy
    # snapshot, so no destructive reverse operation is attempted here.
    pass
