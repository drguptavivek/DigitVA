from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.generate_who_2022_icd_policy import _code_sort_key

DEFAULT_POLICY_PATH = Path(
    "docs/icd-causegrp-mappings/generated/who_2022_icd10_2019_2_policy.json"
)
DEFAULT_DECISION_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/"
    "CMEA10_Blank_WHO_2022_Assignable_Audit- decision.xlsx"
)


def _decision_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Audit"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        for row in rows:
            yield dict(zip(headers, row, strict=False))
    finally:
        workbook.close()


def disabled_codes_from_decisions(workbook_path: Path) -> set[str]:
    disabled_codes: set[str] = set()
    for row in _decision_rows(workbook_path):
        code = str(row.get("icd_code") or "").strip().upper()
        final_decision = str(row.get("final_decision") or "").strip().lower()
        if code and final_decision == "disable who coding":
            disabled_codes.add(code)
    return disabled_codes


def apply_decisions(
    *,
    policy_path: Path = DEFAULT_POLICY_PATH,
    decision_workbook_path: Path = DEFAULT_DECISION_WORKBOOK_PATH,
) -> dict[str, int]:
    payload = json.loads(policy_path.read_text(encoding="utf-8"))
    disabled_codes = disabled_codes_from_decisions(decision_workbook_path)
    retained_items = [
        item for item in payload.get("items", []) if str(item.get("code")) not in disabled_codes
    ]
    removed_count = len(payload.get("items", [])) - len(retained_items)
    retained_items.sort(key=lambda item: _code_sort_key(str(item["code"])))
    payload["items"] = retained_items
    payload["row_count"] = len(retained_items)
    payload["assignability_decision_source"] = str(decision_workbook_path)
    payload["disabled_by_assignability_review_count"] = removed_count
    policy_path.write_text(
        json.dumps(payload, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    return {
        "disabled_decision_codes": len(disabled_codes),
        "removed_policy_items": removed_count,
        "row_count": len(retained_items),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Apply reviewed WHO 2022 ICD assignability decisions to generated policy JSON."
    )
    parser.add_argument("--policy", type=Path, default=DEFAULT_POLICY_PATH)
    parser.add_argument("--decisions", type=Path, default=DEFAULT_DECISION_WORKBOOK_PATH)
    args = parser.parse_args()
    result = apply_decisions(
        policy_path=args.policy,
        decision_workbook_path=args.decisions,
    )
    print(
        f"Applied {result['disabled_decision_codes']} disable decisions; "
        f"removed {result['removed_policy_items']} policy rows; "
        f"{result['row_count']} rows remain."
    )


if __name__ == "__main__":
    main()
