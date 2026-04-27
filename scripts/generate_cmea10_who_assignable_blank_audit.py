from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.generate_who_2022_icd_policy import _code_sort_key

DEFAULT_CMEA10_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/icd-10-CODES_CMEA10_mapped.xlsx"
)
DEFAULT_WHO_POLICY_PATH = Path(
    "docs/icd-causegrp-mappings/generated/who_2022_icd10_2019_2_policy.json"
)
DEFAULT_WHO_BUCKET_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/WHO_2022_VA_Bucket_Mapping.xlsx"
)
DEFAULT_OUTPUT_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/CMEA10_Blank_WHO_2022_Assignable_Audit.xlsx"
)

LIKELY_NON_ASSIGNABLE_PREFIXES = {"H"}
LIKELY_NON_ASSIGNABLE_CODES = {
    "K00",
    "K01",
    "K02",
    "K03",
    "K04",
    "K05",
    "K06",
    "K07",
    "K08",
    "K09",
    "K10",
    "K11",
    "K12",
    "K13",
    "K14",
    "K64",
    "L63",
    "L65",
    "L67",
    "L71",
}
LIKELY_CMEA10_OMISSION_CODES = {
    "A40",
    "A41",
    "A82",
    "A97",
    "C16",
    "C22",
    "I30",
    "I33",
    "I40",
    "K65",
    "K67",
    "K81",
    "L00",
    "L02",
    "L03",
    "N17",
    "N18",
    "N19",
    "O99",
}


def _load_who_assignable(policy_path: Path) -> dict[str, dict[str, object]]:
    payload = json.loads(policy_path.read_text(encoding="utf-8"))
    return {
        str(item["code"]): item
        for item in payload.get("items", [])
        if item.get("is_coding_selectable")
    }


def _load_who_bucket_rows(workbook_path: Path) -> dict[str, dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["ICD_Mapped"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        result = {}
        for row in rows:
            payload = dict(zip(headers, row, strict=False))
            code = str(payload.get("icd_code") or "").strip()
            if code:
                result[code] = payload
        return result
    finally:
        workbook.close()


def _load_cmea10_blank_rows(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["ICD10_CMEA10"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        result = []
        for row_number, row in enumerate(rows, start=2):
            payload = dict(zip(headers, row, strict=False))
            code = str(payload.get("icd_code") or "").strip().upper()
            bucket = str(payload.get("CMEA10") or "").strip()
            if code and not bucket:
                payload["source_row_number"] = row_number
                payload["icd_code"] = code
                result.append(payload)
        return sorted(result, key=lambda row: _code_sort_key(str(row["icd_code"])))
    finally:
        workbook.close()


def _proposed_action(code: str) -> str:
    if code[0] in LIKELY_NON_ASSIGNABLE_PREFIXES or code in LIKELY_NON_ASSIGNABLE_CODES:
        return "Disable WHO coding"
    if code in LIKELY_CMEA10_OMISSION_CODES:
        return "Add CMEA10 bucket"
    return "Clinical review"


def generate_audit_workbook(
    *,
    cmea10_workbook_path: Path = DEFAULT_CMEA10_WORKBOOK_PATH,
    who_policy_path: Path = DEFAULT_WHO_POLICY_PATH,
    who_bucket_workbook_path: Path = DEFAULT_WHO_BUCKET_WORKBOOK_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> dict[str, int]:
    who_assignable = _load_who_assignable(who_policy_path)
    who_bucket_rows = _load_who_bucket_rows(who_bucket_workbook_path)
    cmea_blank_rows = _load_cmea10_blank_rows(cmea10_workbook_path)

    audit_rows = [
        row
        for row in cmea_blank_rows
        if str(row["icd_code"]) in who_assignable
    ]

    workbook = Workbook()
    audit = workbook.active
    audit.title = "Audit"
    audit.append(
        [
            "icd_code",
            "icd_to_display",
            "category",
            "WHO_2022_VA_section",
            "WHO_2022_VA_code",
            "WHO_2022_VA_cause",
            "semantic_level",
            "sex_selectable",
            "age_group_selectable",
            "source_row_number",
            "proposed_action",
            "final_decision",
            "reviewer_notes",
        ]
    )
    action_counts: dict[str, int] = {}
    for row in audit_rows:
        code = str(row["icd_code"])
        who_item = who_assignable[code]
        bucket = who_bucket_rows.get(code, {})
        action = _proposed_action(code)
        action_counts[action] = action_counts.get(action, 0) + 1
        audit.append(
            [
                code,
                row.get("icd_to_display") or f"{code}-{who_item.get('title') or ''}",
                row.get("category") or "",
                bucket.get("WHO_2022_VA_section") or "",
                bucket.get("WHO_2022_VA_code") or "",
                bucket.get("WHO_2022_VA_cause") or "",
                who_item.get("semantic_level") or "",
                who_item.get("sex_selectable") or "",
                who_item.get("age_group_selectable") or "",
                row.get("source_row_number") or "",
                action,
                "",
                "",
            ]
        )

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["CMEA10 blank rows", len(cmea_blank_rows)])
    summary.append(["CMEA10 blank rows still WHO 2022 assignable", len(audit_rows)])
    for action, count in sorted(action_counts.items()):
        summary.append([action, count])

    notes = workbook.create_sheet("Notes")
    notes.append(["topic", "note"])
    notes.append(
        [
            "Purpose",
            "Audit CMEA10 source rows with blank CMEA10 buckets that are still assignable under the WHO 2022 ICD policy.",
        ]
    )
    notes.append(
        [
            "Review workflow",
            "Use proposed_action as a starting heuristic only; final_decision and reviewer_notes are for clinical/data review.",
        ]
    )
    notes.append(
        [
            "No data changes",
            "This workbook is an audit artifact only. It does not change WHO ICD allowability or CMEA10 mappings.",
        ]
    )

    _format_workbook(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "cmea_blank_rows": len(cmea_blank_rows),
        "audit_rows": len(audit_rows),
        **action_counts,
    }


def _format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_len + 2, 10),
                55,
            )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate audit workbook for CMEA10 blank rows still assignable under WHO 2022."
    )
    parser.add_argument("--cmea10-workbook", type=Path, default=DEFAULT_CMEA10_WORKBOOK_PATH)
    parser.add_argument("--who-policy", type=Path, default=DEFAULT_WHO_POLICY_PATH)
    parser.add_argument("--who-bucket-workbook", type=Path, default=DEFAULT_WHO_BUCKET_WORKBOOK_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    summary = generate_audit_workbook(
        cmea10_workbook_path=args.cmea10_workbook,
        who_policy_path=args.who_policy,
        who_bucket_workbook_path=args.who_bucket_workbook,
        output_path=args.output,
    )
    print(
        f"Wrote {summary['audit_rows']} audit rows to {args.output} "
        f"from {summary['cmea_blank_rows']} CMEA10 blank rows."
    )


if __name__ == "__main__":
    main()
