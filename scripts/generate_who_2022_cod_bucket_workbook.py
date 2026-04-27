from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.generate_who_2022_icd_policy import (
    DEFAULT_ICD_CSV_PATH,
    DEFAULT_WORKBOOK_PATH,
    _code_sort_key,
    _expand_expressions,
    _extract_icd_expressions,
    _iter_sheet_records,
    _load_icd_rows,
)

DEFAULT_POLICY_PATH = Path(
    "docs/icd-causegrp-mappings/generated/who_2022_icd10_2019_2_policy.json"
)
DEFAULT_OUTPUT_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/WHO_2022_VA_Bucket_Mapping.xlsx"
)


def _load_policy_items(policy_path: Path) -> list[dict[str, object]]:
    payload = json.loads(policy_path.read_text(encoding="utf-8"))
    return sorted(payload["items"], key=lambda item: _code_sort_key(str(item["code"])))


def _load_road_traffic_codes(
    *,
    workbook_path: Path,
    icd_rows: list[dict[str, str]],
) -> set[str]:
    expressions: list[str] = []
    for record in _iter_sheet_records(workbook_path, "RoadTraffic_Footnote"):
        for value in record.values():
            expressions.extend(_extract_icd_expressions(value))
    return _expand_expressions(expressions, icd_rows, include_descendants=True)


def _transport_codes(icd_rows: list[dict[str, str]]) -> set[str]:
    return _expand_expressions(["V01-V99", "Y85"], icd_rows, include_descendants=True)


def _crosswalk_rows(
    *,
    workbook_path: Path,
    icd_rows: list[dict[str, str]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    road_codes = _load_road_traffic_codes(workbook_path=workbook_path, icd_rows=icd_rows)
    transport_non_road_codes = _transport_codes(icd_rows) - road_codes

    for source_row_number, record in enumerate(
        _iter_sheet_records(workbook_path, "VA_2022_Crosswalk"),
        start=2,
    ):
        va_code = str(record.get("va_code") or "").strip()
        if not va_code:
            continue

        raw_expression = str(record.get("icd10_codes_raw") or "")
        if va_code == "VAs-12.01":
            expanded_codes = road_codes
            match_type = "road_traffic_footnote"
        elif va_code == "VAs-12.02":
            expanded_codes = transport_non_road_codes
            match_type = "transport_non_road"
        else:
            expressions = _extract_icd_expressions(raw_expression)
            expanded_codes = _expand_expressions(
                expressions,
                icd_rows,
                include_descendants=True,
            )
            match_type = "range" if any("-" in expression for expression in expressions) else "exact"

        rows.append(
            {
                "source_row_number": source_row_number,
                "section": str(record.get("section") or "").strip(),
                "va_code": va_code,
                "va_title": str(record.get("va_title") or "").strip(),
                "icd10_codes_raw": raw_expression,
                "notes": str(record.get("notes") or "").strip(),
                "match_type": match_type,
                "expanded_codes": expanded_codes,
            }
        )
    return rows


def _is_x10_x19(code: str) -> bool:
    return code.startswith("X") and code.split(".", 1)[0] in {
        f"X{number:02d}" for number in range(10, 20)
    }


def _match_priority(row: dict[str, object], *, code: str) -> tuple[int, int]:
    match_type = str(row.get("match_type") or "")
    va_code = str(row.get("va_code") or "")
    if code == "P95" and va_code == "VAs-11.02":
        return -1, 0
    if _is_x10_x19(code) and va_code == "VAs-12.99":
        return -1, 0
    is_residual = va_code.endswith(".99") or va_code == "VAs-98" or va_code == "VAs-99"
    if match_type in {"exact", "road_traffic_footnote", "transport_non_road"}:
        match_rank = 0
    else:
        match_rank = 1
    residual_rank = 1 if is_residual else 0
    return match_rank, residual_rank


def generate_workbook(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    icd_csv_path: Path = DEFAULT_ICD_CSV_PATH,
    policy_path: Path = DEFAULT_POLICY_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> dict[str, int]:
    policy_items = _load_policy_items(policy_path)
    policy_by_code = {str(item["code"]): item for item in policy_items}
    icd_rows = [
        row for row in _load_icd_rows(icd_csv_path) if row["code"] in policy_by_code
    ]
    crosswalk_rows = _crosswalk_rows(workbook_path=workbook_path, icd_rows=icd_rows)

    matches_by_code: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in crosswalk_rows:
        for code in row["expanded_codes"]:
            if code in policy_by_code:
                matches_by_code[code].append(row)

    workbook = Workbook()
    mapped_sheet = workbook.active
    mapped_sheet.title = "ICD_Mapped"
    mapped_headers = [
        "disease_id",
        "icd_code",
        "icd_to_display",
        "category",
        "WHO_2022_VA_section",
        "WHO_2022_VA_code",
        "WHO_2022_VA_cause",
        "WHO_2022_VA_match_type",
        "WHO_2022_VA_note",
        "semantic_level",
        "sex_selectable",
        "age_group_selectable",
        "chapter_code",
        "chapter_title",
        "block_code",
        "block_title",
        "three_character_code",
        "three_character_title",
        "ambiguous_match_count",
        "ambiguous_matches",
        "source_row_number",
        "source_icd_expression",
    ]
    mapped_sheet.append(mapped_headers)

    ambiguous_rows: list[list[object]] = []
    unmatched_rows: list[list[object]] = []
    for disease_id, item in enumerate(policy_items, start=1):
        code = str(item["code"])
        matches = sorted(
            matches_by_code.get(code, []),
            key=lambda row: (*_match_priority(row, code=code), int(row["source_row_number"])),
        )
        primary = matches[0] if matches else {}
        ambiguous = matches[1:]
        ambiguous_match_text = "; ".join(
            f'{row["va_code"]} {row["va_title"]}' for row in ambiguous
        )

        mapped_sheet.append(
            [
                disease_id,
                code,
                f'{code}-{item.get("title") or ""}',
                item.get("block_title") or item.get("chapter_title") or "",
                primary.get("section", ""),
                primary.get("va_code", ""),
                primary.get("va_title", ""),
                primary.get("match_type", ""),
                primary.get("notes", ""),
                item.get("semantic_level", ""),
                item.get("sex_selectable", ""),
                item.get("age_group_selectable", ""),
                item.get("chapter_code", ""),
                item.get("chapter_title", ""),
                item.get("block_code", ""),
                item.get("block_title", ""),
                item.get("three_character_code", ""),
                item.get("three_character_title", ""),
                len(matches),
                ambiguous_match_text,
                primary.get("source_row_number", ""),
                primary.get("icd10_codes_raw", ""),
            ]
        )
        if not matches:
            unmatched_rows.append([code, item.get("title") or "", item.get("semantic_level") or ""])
        if ambiguous:
            primary_va_code = primary.get("va_code")
            for row in matches:
                ambiguous_rows.append(
                    [
                        code,
                        item.get("title") or "",
                        row["section"],
                        row["va_code"],
                        row["va_title"],
                        row["match_type"],
                        row["icd10_codes_raw"],
                        row["notes"],
                        "Primary" if row["va_code"] == primary_va_code else "Alternative",
                    ]
                )

    ambiguous_sheet = workbook.create_sheet("Ambiguous_Matches")
    ambiguous_sheet.append(
        [
            "icd_code",
            "icd_title",
            "section",
            "va_code",
            "va_title",
            "match_type",
            "source_icd_expression",
            "notes",
            "Final Decision",
        ]
    )
    for row in ambiguous_rows:
        ambiguous_sheet.append(row)

    unmatched_sheet = workbook.create_sheet("Unmatched_Valid_Codes")
    unmatched_sheet.append(["icd_code", "icd_title", "semantic_level"])
    for row in unmatched_rows:
        unmatched_sheet.append(row)

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["WHO 2022 valid ICD rows", len(policy_items)])
    summary_sheet.append(["Mapped ICD rows", len(policy_items) - len(unmatched_rows)])
    summary_sheet.append(["Unmatched valid ICD rows", len(unmatched_rows)])
    summary_sheet.append(["ICD rows with ambiguous bucket matches", len({row[0] for row in ambiguous_rows})])
    summary_sheet.append(["Three-character rows", sum(1 for item in policy_items if item.get("semantic_level") == "three_character")])
    summary_sheet.append(["Dotted detailed rows", sum(1 for item in policy_items if "." in str(item.get("code")))])

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["topic", "note"])
    notes_sheet.append(
        [
            "Generation",
            "Generated from WHO_2022_VA_Crosswalk.xlsx, icd10_2019_hierarchy.csv, and who_2022_icd10_2019_2_policy.json.",
        ]
    )
    notes_sheet.append(
        [
            "Universe",
            "Rows are all active/selectable WHO 2022 ICD codes from the generated policy JSON, including 3-character and dotted detailed codes.",
        ]
    )
    notes_sheet.append(
        [
            "Ambiguity",
            "When a code matches multiple WHO VA buckets, ICD_Mapped keeps the first crosswalk match as primary and lists the rest in ambiguous_matches.",
        ]
    )
    notes_sheet.append(
        [
            "Road traffic",
            "VAs-12.01 is expanded from the RoadTraffic_Footnote sheet; VAs-12.02 is transport codes V01-V99/Y85 not in that road-traffic set.",
        ]
    )

    _format_workbook(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "valid_rows": len(policy_items),
        "mapped_rows": len(policy_items) - len(unmatched_rows),
        "unmatched_rows": len(unmatched_rows),
        "ambiguous_codes": len({row[0] for row in ambiguous_rows}),
    }


def _format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_len + 2, 10), 55)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate expanded WHO 2022 VA ICD-to-bucket mapping workbook."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--icd-csv", type=Path, default=DEFAULT_ICD_CSV_PATH)
    parser.add_argument("--policy", type=Path, default=DEFAULT_POLICY_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    summary = generate_workbook(
        workbook_path=args.workbook,
        icd_csv_path=args.icd_csv,
        policy_path=args.policy,
        output_path=args.output,
    )
    print(
        f"Wrote {summary['valid_rows']} WHO 2022 ICD rows to {args.output} "
        f"({summary['mapped_rows']} mapped, {summary['unmatched_rows']} unmatched, "
        f"{summary['ambiguous_codes']} ambiguous)."
    )


if __name__ == "__main__":
    main()
