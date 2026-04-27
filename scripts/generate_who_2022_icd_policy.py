from __future__ import annotations

import argparse
import csv
import json
import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_WORKBOOK_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/WHO_2022_VA_Crosswalk.xlsx"
)
DEFAULT_ICD_CSV_PATH = Path("docs/icd-causegrp-mappings/generated/icd10_2019_hierarchy.csv")
DEFAULT_OUTPUT_PATH = Path(
    "docs/icd-causegrp-mappings/generated/who_2022_icd10_2019_2_policy.json"
)
SOURCE_VERSION = "WHO_2022_ICD10_2019_2"
EDITABLE_LEVELS = {"three_character", "detailed_code"}
ICD_EXPR_RE = re.compile(r"\b[A-Z]\d{2}(?:\.\d+)?(?:\s*-\s*[A-Z]?\d{2}(?:\.\d+)?)?\b")


@dataclass(frozen=True)
class IcdCode:
    letter: str
    number: int
    decimal: tuple[int, ...]


def _parse_code(code: str) -> IcdCode:
    normalized = code.strip().upper()
    match = re.fullmatch(r"([A-Z])(\d{2})(?:\.(\d+))?", normalized)
    if not match:
        raise ValueError(f"Invalid ICD code expression: {code!r}")
    decimal = tuple(int(char) for char in (match.group(3) or ""))
    return IcdCode(match.group(1), int(match.group(2)), decimal)


def _code_sort_key(code: str) -> tuple[str, int, tuple[int, ...]]:
    parsed = _parse_code(code)
    return parsed.letter, parsed.number, parsed.decimal


def _is_active(value: str | None) -> bool:
    return (value or "true").strip().lower() not in {"false", "0", "no"}


def _load_icd_rows(icd_csv_path: Path) -> list[dict[str, str]]:
    with icd_csv_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = [
            row
            for row in reader
            if row.get("code")
            and row.get("semantic_level") in EDITABLE_LEVELS
            and _is_active(row.get("is_active"))
        ]
    return sorted(rows, key=lambda row: (int(row.get("sort_order") or 0), row["code"]))


def _iter_sheet_records(workbook_path: Path, sheet_name: str) -> Iterable[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        for values in rows:
            yield dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


def _extract_icd_expressions(text: object) -> list[str]:
    if text is None:
        return []
    return [match.group(0).replace(" ", "").upper() for match in ICD_EXPR_RE.finditer(str(text))]


def _range_bounds(expression: str) -> tuple[str, str]:
    if "-" not in expression:
        return expression, expression
    start, end = expression.split("-", 1)
    if re.match(r"^\d", end):
        end = f"{start[0]}{end}"
    return start, end


def _row_matches_expression(
    row_code: str,
    expression: str,
    *,
    include_descendants: bool = False,
) -> bool:
    start, end = _range_bounds(expression)
    row = _parse_code(row_code)
    start_code = _parse_code(start)
    end_code = _parse_code(end)
    has_range = start != end

    if not has_range:
        if row_code == start:
            return True
        return include_descendants and not start_code.decimal and row_code.startswith(f"{start}.")

    if row.letter < start_code.letter or row.letter > end_code.letter:
        return False
    if row.letter == start_code.letter and row.number < start_code.number:
        return False
    if row.letter == end_code.letter and row.number > end_code.number:
        return False
    if row.letter == start_code.letter and row.number == start_code.number:
        if start_code.decimal and (not row.decimal or row.decimal < start_code.decimal):
            return False
    if row.letter == end_code.letter and row.number == end_code.number:
        if end_code.decimal and (not row.decimal or row.decimal > end_code.decimal):
            return False
    if include_descendants:
        return True
    if not row.decimal:
        if row.letter == start_code.letter and row.number == start_code.number:
            return not start_code.decimal
        if row.letter == end_code.letter and row.number == end_code.number:
            return not end_code.decimal
        return True
    if start_code.decimal and row.letter == start_code.letter and row.number == start_code.number:
        return True
    if end_code.decimal and row.letter == end_code.letter and row.number == end_code.number:
        return True
    return False


def _expand_expressions(
    expressions: Iterable[str],
    icd_rows: list[dict[str, str]],
    *,
    include_descendants: bool = False,
) -> set[str]:
    expanded: set[str] = set()
    for expression in expressions:
        matched_codes: set[str] = set()
        for row in icd_rows:
            if _row_matches_expression(
                row["code"].upper(),
                expression,
                include_descendants=include_descendants,
            ):
                matched_codes.add(row["code"])
        expanded.update(matched_codes)
    return expanded


def _apply_rule(
    policy: dict[str, dict[str, object]],
    codes: set[str],
    *,
    sex: str,
    age_group: str,
) -> None:
    for code in codes:
        if code not in policy:
            continue
        policy[code]["sex_selectable"] = sex
        policy[code]["age_group_selectable"] = age_group


def _item_for_row(row: dict[str, str]) -> dict[str, object]:
    return {
        "code": row["code"],
        "title": row.get("title") or "",
        "semantic_level": row.get("semantic_level"),
        "chapter_code": row.get("chapter_code") or None,
        "chapter_title": row.get("chapter_title") or None,
        "block_code": row.get("block_code") or None,
        "block_title": row.get("block_title") or None,
        "three_character_code": row.get("three_character_code") or None,
        "three_character_title": row.get("three_character_title") or None,
        "is_coding_selectable": True,
        "sex_selectable": "both",
        "age_group_selectable": "all",
    }


def generate_policy(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    icd_csv_path: Path = DEFAULT_ICD_CSV_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> dict[str, object]:
    icd_rows = _load_icd_rows(icd_csv_path)
    row_by_code = {row["code"]: row for row in icd_rows}

    crosswalk_expressions: list[str] = []
    for record in _iter_sheet_records(workbook_path, "VA_2022_Crosswalk"):
        crosswalk_expressions.extend(_extract_icd_expressions(record.get("icd10_codes_raw")))

    allowed_codes = _expand_expressions(crosswalk_expressions, icd_rows)
    policy = {
        code: _item_for_row(row_by_code[code])
        for code in allowed_codes
        if code in row_by_code
    }

    road_expressions: list[str] = []
    for record in _iter_sheet_records(workbook_path, "RoadTraffic_Footnote"):
        for value in record.values():
            road_expressions.extend(_extract_icd_expressions(value))
    road_codes = _expand_expressions(road_expressions, icd_rows)
    for code in road_codes:
        if code in row_by_code:
            policy[code] = _item_for_row(row_by_code[code])

    neonate_expressions = [
        "P05",
        "P07",
        "P20-P22",
        "P23-P24",
        "P36",
        "A33",
        "Q00-Q99",
        "P00-P04",
        "P08-P15",
        "P25-P35",
        "P37-P94",
        "P96",
        "P95",
    ]
    _apply_rule(
        policy,
        _expand_expressions(neonate_expressions, icd_rows),
        sex="both",
        age_group="neonate",
    )
    _apply_rule(
        policy,
        _expand_expressions(["R95"], icd_rows),
        sex="both",
        age_group="infant",
    )
    _apply_rule(
        policy,
        {code for code in policy if code.startswith("O")},
        sex="female",
        age_group="adult",
    )
    _apply_rule(
        policy,
        _expand_expressions(["C51-C58"], icd_rows),
        sex="female",
        age_group="all",
    )
    _apply_rule(
        policy,
        _expand_expressions(["C60-C63"], icd_rows),
        sex="male",
        age_group="all",
    )
    _apply_rule(
        policy,
        _expand_expressions(["C50"], icd_rows),
        sex="both",
        age_group="all",
    )
    _apply_rule(
        policy,
        _expand_expressions(["D25-D28"], icd_rows),
        sex="female",
        age_group="all",
    )
    _apply_rule(
        policy,
        _expand_expressions(["D29"], icd_rows),
        sex="male",
        age_group="all",
    )

    never_codes = _expand_expressions(["S00-T99"], icd_rows, include_descendants=True)
    for code in never_codes:
        policy.pop(code, None)

    items = sorted(policy.values(), key=lambda item: _code_sort_key(str(item["code"])))
    payload = {
        "source_version": SOURCE_VERSION,
        "row_count": len(items),
        "items": items,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate WHO 2022 ICD10-2019-2 policy JSON.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--icd-csv", type=Path, default=DEFAULT_ICD_CSV_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    payload = generate_policy(
        workbook_path=args.workbook,
        icd_csv_path=args.icd_csv,
        output_path=args.output,
    )
    print(f"Wrote {payload['row_count']} policy rows to {args.output}")


if __name__ == "__main__":
    main()
