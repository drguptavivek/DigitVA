from __future__ import annotations

import argparse
from collections.abc import Iterable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.generate_who_2022_icd_policy import (
    DEFAULT_ICD_CSV_PATH,
    DEFAULT_WORKBOOK_PATH,
    _code_sort_key,
    _expand_expressions,
    _extract_icd_expressions,
    _iter_sheet_records,
    _load_icd_rows,
)

DEFAULT_OUTPUT_PATH = Path(
    "docs/icd-causegrp-mappings/ICD-to-VA-Buckets/WHO_2022_VA_RTA_NonRTA_Review.xlsx"
)


def _load_road_traffic_expressions(workbook_path: Path) -> list[str]:
    expressions: list[str] = []
    for record in _iter_sheet_records(workbook_path, "RoadTraffic_Footnote"):
        for value in record.values():
            expressions.extend(_extract_icd_expressions(value))
    return expressions


def _transport_expressions() -> list[str]:
    return ["V01-V99", "Y85"]


def _is_non_road_transport_code(code: str) -> bool:
    base_code = code.split(".", 1)[0]
    return base_code in {f"V{number:02d}" for number in range(90, 100)} or code == "Y85.9"


def _is_rail_non_road_event(*, code: str, title: str) -> bool:
    title_lower = title.lower()
    if code.startswith("V81.0"):
        return True
    if code.startswith("V82.0"):
        return True
    if code.startswith("V81.") or code.startswith("V82."):
        return "traffic accident" not in title_lower
    return False


def _proposed_va_code(*, code: str, title: str, in_footnote: bool) -> str:
    if in_footnote and not _is_non_road_transport_code(code) and not _is_rail_non_road_event(
        code=code,
        title=title,
    ):
        return "VAs-12.01"
    return "VAs-12.02"


def _title_review_flag(*, code: str, title: str, proposed_code: str) -> str:
    title_lower = title.lower()
    if "nontraffic" in title_lower and proposed_code == "VAs-12.01":
        return "Review: title says nontraffic but proposed bucket is RTA."
    if _is_non_road_transport_code(code) and proposed_code == "VAs-12.01":
        return "Review: water/air/space/other transport should not be RTA."
    if (
        proposed_code == "VAs-12.01"
        and "traffic" not in title_lower
        and code != "Y85.0"
    ):
        return "Review: listed in footnote but title does not explicitly say traffic."
    if proposed_code == "VAs-12.02" and "traffic" in title_lower and "nontraffic" not in title_lower:
        return "Review: title says traffic but proposed bucket is non-RTA."
    return ""


def _transport_rows(icd_csv_path: Path) -> list[dict[str, str]]:
    rows = _load_icd_rows(icd_csv_path)
    transport_codes = _expand_expressions(
        _transport_expressions(),
        rows,
        include_descendants=True,
    )
    return sorted(
        [row for row in rows if row["code"] in transport_codes],
        key=lambda row: _code_sort_key(row["code"]),
    )


def _append_review_rows(
    *,
    sheet,
    rows: Iterable[dict[str, str]],
    road_codes: set[str],
) -> None:
    for row in rows:
        code = row["code"]
        title = row.get("title") or ""
        in_footnote = code in road_codes
        proposed_code = _proposed_va_code(
            code=code,
            title=title,
            in_footnote=in_footnote,
        )
        proposed_cause = (
            "Road traffic accident"
            if proposed_code == "VAs-12.01"
            else "Other transport accident"
        )
        review_flag = _title_review_flag(
            code=code,
            title=row.get("title") or "",
            proposed_code=proposed_code,
        )
        sheet.append(
            [
                code,
                title,
                row.get("semantic_level") or "",
                row.get("chapter_code") or "",
                row.get("block_code") or "",
                row.get("three_character_code") or "",
                "Yes" if in_footnote else "No",
                proposed_code,
                proposed_cause,
                review_flag,
                proposed_code,
                "",
            ]
        )


def _format_workbook(workbook: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_len + 2, 12),
                60,
            )


def generate_review_workbook(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    icd_csv_path: Path = DEFAULT_ICD_CSV_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
) -> dict[str, int]:
    icd_rows = _transport_rows(icd_csv_path)
    road_expressions = _load_road_traffic_expressions(workbook_path)
    road_codes = _expand_expressions(
        road_expressions,
        icd_rows,
        include_descendants=True,
    )

    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "RTA_NonRTA_Review"
    headers = [
        "icd_code",
        "icd_title",
        "semantic_level",
        "chapter_code",
        "block_code",
        "three_character_code",
        "in_who_footnote_list",
        "proposed_va_code",
        "proposed_va_cause",
        "review_flag",
        "final_decision",
        "reviewer_note",
    ]
    review_sheet.append(headers)
    _append_review_rows(sheet=review_sheet, rows=icd_rows, road_codes=road_codes)

    footnote_sheet = workbook.create_sheet("Footnote_Expressions")
    footnote_sheet.append(["source", "icd_expression"])
    for expression in road_expressions:
        footnote_sheet.append(["RoadTraffic_Footnote", expression])

    summary_sheet = workbook.create_sheet("Summary")
    proposed_rta_count = sum(
        1
        for row in icd_rows
        if _proposed_va_code(
            code=row["code"],
            title=row.get("title") or "",
            in_footnote=row["code"] in road_codes,
        )
        == "VAs-12.01"
    )
    summary_sheet.append(["metric", "value"])
    summary_sheet.append(["transport_codes_reviewed", len(icd_rows)])
    summary_sheet.append(["codes_in_who_footnote_list", len(road_codes)])
    summary_sheet.append(["proposed_rta_vas_12_01", proposed_rta_count])
    summary_sheet.append(["proposed_non_rta_vas_12_02", len(icd_rows) - proposed_rta_count])
    summary_sheet.append(
        [
            "review_rule",
            "Codes in RoadTraffic_Footnote are proposed as VAs-12.01 except V90-V99, Y85.9, and rail/streetcar events that do not explicitly say traffic accident; these are proposed as VAs-12.02. Review flags call out titles that do not explicitly say traffic.",
        ]
    )

    _format_workbook(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "transport_codes_reviewed": len(icd_rows),
        "codes_in_who_footnote_list": len(road_codes),
        "proposed_rta": proposed_rta_count,
        "proposed_non_rta": len(icd_rows) - proposed_rta_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate WHO 2022 VA VAs-12.01/VAs-12.02 RTA vs non-RTA review workbook."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--icd-csv", type=Path, default=DEFAULT_ICD_CSV_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    args = parser.parse_args()
    summary = generate_review_workbook(
        workbook_path=args.workbook,
        icd_csv_path=args.icd_csv,
        output_path=args.output,
    )
    print(
        f"Wrote {summary['transport_codes_reviewed']} transport rows to {args.output} "
        f"({summary['proposed_rta']} proposed RTA, "
        f"{summary['proposed_non_rta']} proposed non-RTA)."
    )


if __name__ == "__main__":
    main()
