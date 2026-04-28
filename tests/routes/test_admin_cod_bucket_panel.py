from datetime import datetime, timezone
from io import BytesIO
import uuid

import sqlalchemy as sa
from openpyxl import load_workbook

from app import db
from app.models import (
    MapIcdCodBucket,
    MasCodBucketNode,
    MasCodBucketScheme,
    MasCodBucketSchemeAgeBand,
    MasIcd1020192,
    # Deprecated as of 2026-04-20: retained here only for legacy fixture coverage.
    VaIcdCodes,
)
from tests.base import BaseTestCase


class AdminCodBucketPanelTests(BaseTestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        scheme = MasCodBucketScheme(
            scheme_code=f"TEST_ADMIN_{uuid.uuid4().hex[:8].upper()}",
            scheme_name="Admin COD Scheme",
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        cls.scheme_code = scheme.scheme_code
        cls.scheme_id = scheme.scheme_id
        db.session.add(
            MasCodBucketSchemeAgeBand(
                scheme_id=scheme.scheme_id,
                age_scope="adult_over5y",
                age_label="Adult / Over 5 Years",
                min_age_value=5,
                min_age_unit="years",
                max_age_value=120,
                max_age_unit="years",
                level_count=3,
                sort_order=1,
                is_active=True,
            )
        )
        db.session.flush()
        cls.adult_age_band_id = db.session.scalar(
            sa.select(MasCodBucketSchemeAgeBand.age_band_id).where(
                MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id,
                MasCodBucketSchemeAgeBand.age_scope == "adult_over5y",
            )
        )

        category = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope="adult_over5y",
            node_type="category",
            node_code="injuries",
            node_label="Injuries",
            sort_order=1,
            is_active=True,
        )
        subcategory = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope="adult_over5y",
            node_type="subcategory",
            parent=category,
            node_code="road_injuries",
            node_label="Road Injuries",
            sort_order=1,
            is_active=True,
        )
        field_a = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope="adult_over5y",
            node_type="field",
            parent=subcategory,
            node_code="pedestrian",
            node_label="Pedestrian Road Injury",
            sort_order=1,
            is_active=True,
        )
        field_b = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope="adult_over5y",
            node_type="field",
            parent=subcategory,
            node_code="vehicle",
            node_label="Vehicle Occupant Injury",
            sort_order=2,
            is_active=True,
        )
        db.session.add_all([category, subcategory, field_a, field_b])
        db.session.flush()
        cls.field_a_id = field_a.node_id
        cls.field_b_id = field_b.node_id

        for code, title, is_selectable in (
            ("A00", "Cholera", True),
            ("V01", "Pedestrian injured in collision with pedal cycle", True),
            ("V02", "Pedestrian injured in collision with two- or three-wheeled motor vehicle", True),
            ("V03", "Pedestrian injured in collision with car, pick-up truck or van", True),
            ("Z91", "Personal history of risk-factors, not elsewhere classified", False),
        ):
            row = db.session.get(MasIcd1020192, code)
            if row is None:
                row = MasIcd1020192(
                    code=code,
                    title=title,
                    node_type="category",
                    semantic_level="three_character",
                    sort_order=1,
                    parent_code=None,
                    chapter_code="XXI",
                    chapter_title="Factors influencing health status and contact with health services",
                    block_code=f"{code}-{code}",
                    block_title=title,
                    three_character_code=code,
                    three_character_title=title,
                    has_children=False,
                    is_leaf=True,
                    is_three_character_code=True,
                    is_detailed_code=False,
                    is_coding_selectable=is_selectable,
                    sex_selectable="both" if is_selectable else None,
                    age_group_selectable="all" if is_selectable else None,
                    policy_status="allowed" if is_selectable else "unreviewed",
                    source_version="2019-test",
                    source_path="tests",
                    is_active=True,
                )
                db.session.add(row)
            else:
                row.title = title
                row.semantic_level = "three_character"
                row.is_three_character_code = True
                row.is_detailed_code = False
                row.is_coding_selectable = is_selectable
                row.sex_selectable = "both" if is_selectable else None
                row.age_group_selectable = "all" if is_selectable else None
                row.is_active = True

        db.session.add_all(
            [
                VaIcdCodes(
                    disease_id=910001,
                    icd_code="V01",
                    icd_to_display="V01-Admin test pedestrian injury",
                    category="test",
                ),
                VaIcdCodes(
                    disease_id=910002,
                    icd_code="V02",
                    icd_to_display="V02-Admin test cyclist injury",
                    category="test",
                ),
                VaIcdCodes(
                    disease_id=910003,
                    icd_code="V03",
                    icd_to_display="V03-Admin test pedestrian collision",
                    category="test",
                ),
                VaIcdCodes(
                    disease_id=910004,
                    icd_code="ZZ91",
                    icd_to_display="ZZ91-Admin unique unmapped syndrome",
                    category="test",
                ),
            ]
        )

        mapping = MapIcdCodBucket(
            scheme_id=scheme.scheme_id,
            age_scope="adult_over5y",
            icd_code="V01",
            node_id=field_a.node_id,
            is_active=True,
        )
        db.session.add(mapping)
        db.session.commit()
        cls.mapping_id = mapping.mapping_id
        cls.category_id = category.node_id

    def test_cod_bucket_panel_renders_for_admin(self):
        self._login(self.base_admin_id)
        response = self.client.get("/admin/panels/cod-buckets")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"COD Bucket Mapping", response.data)
        self.assertIn(b"New Scheme", response.data)

    def test_cod_bucket_scheme_detail_returns_editor_payload(self):
        self._login(self.base_admin_id)
        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}?age_scope=adult_over5y"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["scheme"]["scheme_code"], self.scheme_code)
        self.assertEqual(payload["selected_age_scope"], "adult_over5y")
        self.assertGreaterEqual(len(payload["nodes"]), 4)
        self.assertTrue(any(node["node_id"] == str(self.category_id) for node in payload["nodes"]))
        self.assertTrue(any(node["node_id"] == str(self.field_a_id) for node in payload["nodes"]))
        self.assertNotIn("mappings", payload)
        self.assertNotIn("field_options", payload)
        self.assertEqual(payload["selected_age_band"]["level_count"], 3)

    def test_cod_bucket_scheme_export_returns_json_attachment(self):
        self._login(self.base_admin_id)
        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/export"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/json")
        self.assertIn(
            f'attachment; filename="cod_bucket_scheme_{self.scheme_code.lower()}.json"',
            response.headers.get("Content-Disposition", ""),
        )

        payload = response.get_json()
        self.assertEqual(payload["scheme"]["scheme_code"], self.scheme_code)
        self.assertEqual(payload["scheme"]["scheme_name"], "Admin COD Scheme")
        self.assertEqual(len(payload["age_bands"]), 1)
        self.assertGreaterEqual(len(payload["nodes"]), 4)
        self.assertTrue(any(item["icd_code"] == "V01" for item in payload["mappings"]))
        v01_mapping = next(item for item in payload["mappings"] if item["icd_code"] == "V01")
        self.assertIn("Road Injuries", v01_mapping["node_path_label"])

    def test_cod_bucket_scheme_export_returns_xlsx_with_bucket_and_manual_override_status(self):
        self._login(self.base_admin_id)
        mapping = db.session.get(MapIcdCodBucket, self.mapping_id)
        mapping.source_sheet = "admin_cod_bucket_editor"
        mapping.match_type = "manual_override"
        mapping.mapping_note = "Manual override to default COD bucket scheme mapping."
        db.session.commit()

        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/export.xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            f'attachment; filename="cod_bucket_scheme_{self.scheme_code.lower()}.xlsx"',
            response.headers.get("Content-Disposition", ""),
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
        self.assertIn("Bucket Nodes", workbook.sheetnames)
        self.assertIn("ICD Mappings", workbook.sheetnames)

        mapping_sheet = workbook["ICD Mappings"]
        mapping_headers = [
            cell.value for cell in next(mapping_sheet.iter_rows(min_row=1, max_row=1))
        ]
        mapping_rows = [
            dict(zip(mapping_headers, row))
            for row in mapping_sheet.iter_rows(min_row=2, values_only=True)
        ]
        v01 = next(row for row in mapping_rows if row["ICD Code"] == "V01")
        self.assertEqual(v01["Manual Override"], "Yes")
        self.assertEqual(v01["Match Type"], "manual_override")
        self.assertIn("Road Injuries", v01["Bucket Path"])

        node_sheet = workbook["Bucket Nodes"]
        node_headers = [cell.value for cell in next(node_sheet.iter_rows(min_row=1, max_row=1))]
        node_rows = [
            dict(zip(node_headers, row))
            for row in node_sheet.iter_rows(min_row=2, values_only=True)
        ]
        self.assertTrue(
            any(row["Node Label"] == "Pedestrian Road Injury" for row in node_rows)
        )

    def test_cod_bucket_scheme_icd_export_returns_xlsx_with_policy_and_override_status(self):
        self._login(self.base_admin_id)
        mapping = db.session.get(MapIcdCodBucket, self.mapping_id)
        mapping.source_sheet = "admin_cod_bucket_editor"
        mapping.match_type = "manual_override"
        mapping.mapping_note = "Manual override to default COD bucket scheme mapping."
        db.session.commit()

        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/icd-export.xlsx"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            f'attachment; filename="cod_bucket_icd10_{self.scheme_code.lower()}.xlsx"',
            response.headers.get("Content-Disposition", ""),
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
        sheet = workbook["ICD10 Codes"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            headers,
            [
                "ICD Code",
                "ICD Title",
                "Semantic Level",
                "Chapter Code",
                "Chapter Title",
                "Block Code",
                "Block Title",
                "Three Character Code",
                "Coding Allowed",
                "Age Selectable",
                "Sex Selectable",
                "Policy Status",
                "Mapped To Scheme",
                "Scheme Age Scope",
                "Scheme Age Band",
                "COD Bucket Path",
                "Match Type",
                "Manual Override",
                "Source Sheet",
                "Source Row Number",
                "Mapping Note",
            ],
        )
        rows = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        v01 = next(row for row in rows if row["ICD Code"] == "V01")
        self.assertEqual(v01["Coding Allowed"], "Yes")
        self.assertEqual(v01["Age Selectable"], "all")
        self.assertEqual(v01["Sex Selectable"], "both")
        self.assertEqual(v01["Mapped To Scheme"], "Yes")
        self.assertEqual(v01["Manual Override"], "Yes")
        self.assertIn("Road Injuries", v01["COD Bucket Path"])
        z91 = next(row for row in rows if row["ICD Code"] == "Z91")
        self.assertEqual(z91["Coding Allowed"], "No")
        self.assertEqual(z91["Mapped To Scheme"], "No")
        self.assertEqual(z91["Manual Override"], "No")

    def test_cod_bucket_scheme_create_returns_feedback_but_still_persists(self):
        self._login(self.base_admin_id)
        response = self.client.post(
            "/admin/api/cod-bucket-schemes",
            json={
                "scheme_name": "Custom Feedback Scheme",
                "scheme_code": f"CUSTOM_{uuid.uuid4().hex[:6].upper()}",
                "age_bands": [
                    {
                        "age_label": "Neonate",
                        "min_age_value": 0,
                        "min_age_unit": "days",
                        "max_age_value": 28,
                        "max_age_unit": "days",
                        "level_count": 3,
                    },
                    {
                        "age_label": "Adult",
                        "min_age_value": 5,
                        "min_age_unit": "years",
                        "max_age_value": 120,
                        "max_age_unit": "years",
                        "level_count": 2,
                    },
                ],
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 201)
        payload = response.get_json()
        self.assertTrue(payload["warnings"])
        self.assertEqual(payload["scheme"]["scheme_name"], "Custom Feedback Scheme")
        self.assertEqual(len(payload["scheme"]["age_bands"]), 2)

    def test_cod_bucket_scheme_update_can_edit_name_and_add_age_band(self):
        self._login(self.base_admin_id)
        response = self.client.patch(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}",
            json={
                "scheme_name": "Admin COD Scheme Updated",
                "age_bands": [
                    {
                        "age_band_id": str(self.adult_age_band_id),
                        "age_label": "Adult / 5+ Years",
                        "min_age_value": 5,
                        "min_age_unit": "years",
                        "max_age_value": 120,
                        "max_age_unit": "years",
                        "level_count": 3,
                    },
                    {
                        "age_label": "Child / 1-59 Months",
                        "min_age_value": 1,
                        "min_age_unit": "months",
                        "max_age_value": 60,
                        "max_age_unit": "months",
                        "level_count": 2,
                    },
                ],
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["scheme"]["scheme_name"], "Admin COD Scheme Updated")
        self.assertEqual(len(payload["scheme"]["age_bands"]), 2)
        self.assertTrue(
            any(item["label"] == "Child / 1-59 Months" for item in payload["scheme"]["age_bands"])
        )

    def test_cod_bucket_scheme_update_blocks_removing_non_empty_age_band(self):
        self._login(self.base_admin_id)
        response = self.client.patch(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}",
            json={
                "scheme_name": "Admin COD Scheme",
                "age_bands": [
                    {
                        "age_label": "Child / 1-59 Months",
                        "min_age_value": 1,
                        "min_age_unit": "months",
                        "max_age_value": 60,
                        "max_age_unit": "months",
                        "level_count": 2,
                    }
                ],
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertIn("Cannot remove age band", payload["error"])

    def test_cod_bucket_scheme_reset_default_rejects_non_source_backed_scheme(self):
        self._login(self.base_admin_id)
        response = self.client.post(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/reset-default",
            json={"age_scope": "adult_over5y"},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertIn("cannot be reset from source", payload["error"].lower())

    def test_cod_bucket_node_mappings_returns_selected_leaf_only(self):
        self._login(self.base_admin_id)
        isolated_mapping = MapIcdCodBucket(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            icd_code="A00",
            node_id=self.field_a_id,
            is_active=True,
        )
        db.session.add(isolated_mapping)
        db.session.commit()

        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/nodes/{self.field_a_id}/mappings"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["node"]["node_id"], str(self.field_a_id))
        self.assertIn("A00", [mapping["icd_code"] for mapping in payload["mappings"]])
        self.assertTrue(any(
            mapping["icd_code"] == "A00"
            and mapping["icd_to_display"] == "A00-Cholera"
            for mapping in payload["mappings"]
        ))

    def test_cod_bucket_icd_search_includes_current_path_context(self):
        self._login(self.base_admin_id)
        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/icd-search"
            f"?age_scope=adult_over5y&q=V01&selected_node_id={self.field_a_id}"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        v01 = next(item for item in payload["results"] if item["icd_code"] == "V01")
        self.assertTrue(v01["is_mapped"])
        self.assertTrue(v01["is_selected_leaf"])
        self.assertIn("Pedestrian Road Injury", v01["current_path_label"])

    def test_cod_bucket_icd_search_can_filter_to_unmapped(self):
        self._login(self.base_admin_id)
        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/icd-search"
            "?age_scope=adult_over5y&q=V0&unmapped_only=1"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(any(item["icd_code"] == "V02" for item in payload["results"]))
        self.assertFalse(any(item["icd_code"] == "V01" for item in payload["results"]))

    def test_cod_bucket_icd_search_includes_non_assignable_master_codes_with_marker(self):
        self._login(self.base_admin_id)
        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/icd-search"
            "?age_scope=adult_over5y&q=Z91"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        z91 = next(item for item in payload["results"] if item["icd_code"] == "Z91")
        self.assertFalse(z91["is_assignable_in_coding"])
        self.assertEqual(z91["coding_status_label"], "Currently not assignable in coding")

    def test_cod_bucket_unmapped_icd_grid_payload_lists_active_scheme_unmapped_codes(self):
        self._login(self.base_admin_id)
        db.session.merge(
            MasIcd1020192(
                code="XY01",
                title="Admin unmapped scheme test code",
                node_type="category",
                semantic_level="detailed_code",
                sort_order=99,
                parent_code="XY0",
                chapter_code="X",
                chapter_title="Test chapter",
                block_code="XY0-XY9",
                block_title="Test block",
                three_character_code="XY0",
                three_character_title="Test three character",
                has_children=False,
                is_leaf=True,
                is_three_character_code=False,
                is_detailed_code=True,
                is_coding_selectable=True,
                sex_selectable="both",
                age_group_selectable="all",
                policy_status="allowed",
                source_version="2019-test",
                source_path="tests",
                is_active=True,
            )
        )
        db.session.commit()

        response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/unmapped-icd"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["scheme"]["scheme_code"], self.scheme_code)
        self.assertTrue(any(item["code"] == "XY01" for item in payload["rows"]))
        self.assertFalse(any(item["code"] == "V01" for item in payload["rows"]))
        z91 = next(item for item in payload["rows"] if item["code"] == "Z91")
        self.assertFalse(z91["is_assignable_in_coding"])
        self.assertEqual(z91["coding_status_label"], "Currently not assignable in coding")

    def test_cod_bucket_node_patch_updates_label_and_sort_order(self):
        self._login(self.base_admin_id)
        response = self.client.patch(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/nodes/{self.category_id}",
            json={"node_label": "External Injuries", "sort_order": 5},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        node = db.session.get(MasCodBucketNode, self.category_id)
        self.assertEqual(node.node_label, "External Injuries")
        self.assertEqual(node.sort_order, 5)

    def test_cod_bucket_node_delete_leaf_can_unmap_associated_icds(self):
        self._login(self.base_admin_id)
        field = MasCodBucketNode(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            node_type="field",
            parent_node_id=db.session.get(MasCodBucketNode, self.field_b_id).parent_node_id,
            node_code=f"delete_leaf_{uuid.uuid4().hex[:8]}",
            node_label="Delete Leaf Field",
            sort_order=99,
            is_active=True,
        )
        db.session.add(field)
        db.session.flush()
        db.session.add(
            VaIcdCodes(
                disease_id=910006,
                icd_code="ZZ93",
                icd_to_display="ZZ93-Admin delete leaf syndrome",
                category="test",
            )
        )
        mapping = MapIcdCodBucket(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            icd_code="ZZ93",
            node_id=field.node_id,
            is_active=True,
        )
        db.session.add(mapping)
        db.session.commit()

        response = self.client.delete(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/nodes/{field.node_id}",
            json={"mapping_disposition": "unmap"},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(db.session.get(MasCodBucketNode, field.node_id))
        remaining = db.session.scalar(
            sa.select(sa.func.count())
            .select_from(MapIcdCodBucket)
            .where(
                MapIcdCodBucket.scheme_id == self.scheme_id,
                MapIcdCodBucket.age_scope == "adult_over5y",
                MapIcdCodBucket.icd_code == "ZZ93",
            )
        )
        self.assertEqual(remaining, 0)

    def test_cod_bucket_node_delete_higher_level_can_move_icds_to_unmapped(self):
        self._login(self.base_admin_id)
        category = MasCodBucketNode(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            node_type="category",
            node_code=f"delete_category_{uuid.uuid4().hex[:8]}",
            node_label="Delete Category",
            sort_order=120,
            is_active=True,
        )
        db.session.add(category)
        db.session.flush()
        subcategory = MasCodBucketNode(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            node_type="subcategory",
            parent=category,
            node_code=f"delete_subcategory_{uuid.uuid4().hex[:8]}",
            node_label="Delete Subcategory",
            sort_order=1,
            is_active=True,
        )
        db.session.add(subcategory)
        db.session.flush()
        field = MasCodBucketNode(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            node_type="field",
            parent=subcategory,
            node_code=f"delete_field_{uuid.uuid4().hex[:8]}",
            node_label="Delete Field",
            sort_order=1,
            is_active=True,
        )
        db.session.add(field)
        db.session.add_all(
            [
                VaIcdCodes(
                    disease_id=910007,
                    icd_code="ZZ94",
                    icd_to_display="ZZ94-Admin cascade delete syndrome",
                    category="test",
                ),
                VaIcdCodes(
                    disease_id=910008,
                    icd_code="ZZ95",
                    icd_to_display="ZZ95-Admin cascade delete syndrome two",
                    category="test",
                ),
            ]
        )
        db.session.flush()
        db.session.add_all(
            [
                MapIcdCodBucket(
                    scheme_id=self.scheme_id,
                    age_scope="adult_over5y",
                    icd_code="ZZ94",
                    node_id=field.node_id,
                    is_active=True,
                ),
                MapIcdCodBucket(
                    scheme_id=self.scheme_id,
                    age_scope="adult_over5y",
                    icd_code="ZZ95",
                    node_id=field.node_id,
                    is_active=True,
                ),
            ]
        )
        db.session.commit()

        response = self.client.delete(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/nodes/{category.node_id}",
            json={"mapping_disposition": "move_to_unmapped"},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertIn("Unmapped", payload["replacement_leaf_path_label"])
        self.assertIsNone(db.session.get(MasCodBucketNode, category.node_id))
        reassigned = db.session.scalars(
            sa.select(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == self.scheme_id,
                MapIcdCodBucket.age_scope == "adult_over5y",
                MapIcdCodBucket.icd_code.in_(["ZZ94", "ZZ95"]),
            )
        ).all()
        self.assertEqual(len(reassigned), 2)
        self.assertEqual({str(row.node_id) for row in reassigned}, {payload["replacement_leaf_node_id"]})

    def test_cod_bucket_mapping_patch_repoints_single_leaf_target(self):
        self._login(self.base_admin_id)
        response = self.client.patch(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/mappings/{self.mapping_id}",
            json={"node_id": str(self.field_b_id)},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        mapping = db.session.get(MapIcdCodBucket, self.mapping_id)
        self.assertEqual(mapping.node_id, self.field_b_id)
        self.assertEqual(mapping.source_sheet, "admin_cod_bucket_editor")
        self.assertEqual(mapping.match_type, "manual_override")
        self.assertEqual(
            mapping.mapping_note,
            "Manual override to default COD bucket scheme mapping.",
        )
        count = db.session.scalar(
            sa.select(sa.func.count())
            .select_from(MapIcdCodBucket)
            .where(
                MapIcdCodBucket.scheme_id == self.scheme_id,
                MapIcdCodBucket.age_scope == "adult_over5y",
                MapIcdCodBucket.icd_code == "V01",
            )
        )
        self.assertEqual(count, 1)

    def test_cod_bucket_mapping_delete_unmaps_icd_code(self):
        self._login(self.base_admin_id)
        mapping = MapIcdCodBucket(
            scheme_id=self.scheme_id,
            age_scope="adult_over5y",
            icd_code="V03",
            node_id=self.field_a_id,
            is_active=True,
        )
        db.session.add(mapping)
        db.session.commit()

        response = self.client.delete(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/mappings/{mapping.mapping_id}",
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(db.session.get(MapIcdCodBucket, mapping.mapping_id))

        search_response = self.client.get(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/icd-search"
            "?age_scope=adult_over5y&q=V03&unmapped_only=1"
        )
        self.assertEqual(search_response.status_code, 200)
        payload = search_response.get_json()
        self.assertTrue(any(item["icd_code"] == "V03" for item in payload["results"]))

    def test_cod_bucket_mapping_post_adds_codes_to_selected_leaf(self):
        self._login(self.base_admin_id)
        response = self.client.post(
            f"/admin/api/cod-bucket-schemes/{self.scheme_code}/mappings",
            json={"node_id": str(self.field_a_id), "icd_codes": "V02, V03"},
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 201)
        rows = db.session.scalars(
            sa.select(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == self.scheme_id,
                MapIcdCodBucket.age_scope == "adult_over5y",
                MapIcdCodBucket.node_id == self.field_a_id,
                MapIcdCodBucket.icd_code.in_(["V02", "V03"]),
            )
        ).all()
        self.assertEqual(sorted(row.icd_code for row in rows), ["V02", "V03"])
        for row in rows:
            self.assertEqual(row.source_sheet, "admin_cod_bucket_editor")
            self.assertEqual(row.match_type, "manual_override")
            self.assertEqual(
                row.mapping_note,
                "Manual override to default COD bucket scheme mapping.",
            )
