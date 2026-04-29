from datetime import datetime, timezone
from io import BytesIO
import io
import json
import uuid

import sqlalchemy as sa
from openpyxl import load_workbook

from app import db
from app.models import (
    MapIcdCodBucket,
    MapIcd10LegacyReportingAlias,
    MasCodBucketNode,
    MasCodBucketScheme,
    MasCodBucketSchemeAgeBand,
    MasIcd1020192,
)
from tests.base import BaseTestCase


class TestAdminIcd10Browser(BaseTestCase):
    def setUp(self):
        super().setUp()
        db.session.execute(sa.delete(MapIcd10LegacyReportingAlias))
        db.session.execute(sa.delete(MasIcd1020192))
        db.session.flush()

        now = datetime.now(timezone.utc)
        db.session.add_all(
            [
                MasIcd1020192(
                    code="I",
                    title="Certain infectious and parasitic diseases",
                    node_type="chapter",
                    semantic_level="chapter",
                    sort_order=1,
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    has_children=True,
                    is_leaf=False,
                    is_three_character_code=False,
                    is_detailed_code=False,
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MasIcd1020192(
                    code="A00-A09",
                    title="Intestinal infectious diseases",
                    node_type="block",
                    semantic_level="block",
                    sort_order=2,
                    parent_code="I",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="A00-A09",
                    block_title="Intestinal infectious diseases",
                    has_children=True,
                    is_leaf=False,
                    is_three_character_code=False,
                    is_detailed_code=False,
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MasIcd1020192(
                    code="A00",
                    title="Cholera",
                    node_type="category",
                    semantic_level="three_character",
                    sort_order=3,
                    parent_code="A00-A09",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="A00-A09",
                    block_title="Intestinal infectious diseases",
                    three_character_code="A00",
                    three_character_title="Cholera",
                    has_children=True,
                    is_leaf=False,
                    is_three_character_code=True,
                    is_detailed_code=False,
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MasIcd1020192(
                    code="A00.0",
                    title="Cholera due to Vibrio cholerae 01, biovar cholerae",
                    node_type="category",
                    semantic_level="detailed_code",
                    sort_order=4,
                    parent_code="A00",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="A00-A09",
                    block_title="Intestinal infectious diseases",
                    three_character_code="A00",
                    three_character_title="Cholera",
                    has_children=False,
                    is_leaf=True,
                    is_three_character_code=False,
                    is_detailed_code=True,
                    is_coding_selectable=True,
                    sex_selectable="female",
                    age_group_selectable="adult",
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MapIcd10LegacyReportingAlias(
                    legacy_code="A90",
                    reporting_code="A00",
                    note="Legacy dengue example for browser rendering.",
                ),
                MapIcd10LegacyReportingAlias(
                    legacy_code="I84",
                    reporting_code="A00.0",
                    note="Legacy haemorrhoids example for browser rendering.",
                ),
            ]
        )
        db.session.commit()

    def test_admin_panel_renders_for_admin(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/panels/icd10-browser")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertIn("ICD-10 2019 Browser", body)
        self.assertIn("Legacy ICD Reporting Aliases", body)
        self.assertIn("A90", body)
        self.assertIn("A00", body)
        self.assertIn("I84", body)
        self.assertIn("A00.0", body)

    def test_admin_reporting_alias_create_rejects_current_master_code_as_legacy(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.post(
            "/admin/api/icd10/2019-2/reporting-aliases",
            json={
                "legacy_code": "A00",
                "reporting_code": "A00.0",
                "note": "should fail",
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("already exists in the ICD-10 2019 master", response.get_json()["error"])

    def test_admin_reporting_alias_create_and_delete(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.post(
            "/admin/api/icd10/2019-2/reporting-aliases",
            json={
                "legacy_code": "A91",
                "reporting_code": "A00",
                "note": "Legacy dengue alias",
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 201)
        payload = response.get_json()
        self.assertEqual(payload["legacy_code"], "A91")
        self.assertEqual(payload["reporting_code"], "A00")
        self.assertEqual(payload["reporting_title"], "Cholera")

        saved = db.session.get(MapIcd10LegacyReportingAlias, "A91")
        self.assertIsNotNone(saved)
        self.assertEqual(saved.reporting_code, "A00")

        delete_response = self.client.delete(
            "/admin/api/icd10/2019-2/reporting-aliases/A91",
            headers=self._csrf_headers(),
        )

        self.assertEqual(delete_response.status_code, 200)
        self.assertIsNone(db.session.get(MapIcd10LegacyReportingAlias, "A91"))

    def test_admin_panel_denied_for_project_pi(self):
        self._login(str(self.base_project_pi_user.user_id))

        response = self.client.get("/admin/panels/icd10-browser")

        self.assertEqual(response.status_code, 403)

    def test_admin_children_api_returns_root_nodes(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/api/icd10/2019-2/children")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual([row["code"] for row in payload["children"]], ["I"])

    def test_admin_children_api_returns_block_status_indicator(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/api/icd10/2019-2/children?parent_code=I")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual([row["code"] for row in payload["children"]], ["A00-A09"])
        self.assertEqual(payload["children"][0]["status_indicator"], "yellow")

    def test_admin_children_api_filter_updates_block_counts(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get(
            "/admin/api/icd10/2019-2/children?parent_code=I&coding_filter=active"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual([row["code"] for row in payload["children"]], ["A00-A09"])
        self.assertEqual(payload["children"][0]["child_count"], 1)

        response = self.client.get(
            "/admin/api/icd10/2019-2/children?parent_code=A00-A09&coding_filter=active"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual([row["code"] for row in payload["children"]], ["A00"])
        self.assertEqual(payload["children"][0]["status_indicator"], "yellow")

    def test_admin_children_api_counts_nested_block_descendants(self):
        now = datetime.now(timezone.utc)
        db.session.add_all(
            [
                MasIcd1020192(
                    code="B00-B99",
                    title="Nested root block",
                    node_type="block",
                    semantic_level="block",
                    sort_order=5,
                    parent_code="I",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="B00-B99",
                    block_title="Nested root block",
                    has_children=True,
                    is_leaf=False,
                    is_three_character_code=False,
                    is_detailed_code=False,
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MasIcd1020192(
                    code="B00-B49",
                    title="Nested sub-block",
                    node_type="block",
                    semantic_level="block",
                    sort_order=6,
                    parent_code="B00-B99",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="B00-B99",
                    block_title="Nested root block",
                    has_children=True,
                    is_leaf=False,
                    is_three_character_code=False,
                    is_detailed_code=False,
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
                MasIcd1020192(
                    code="B00",
                    title="Herpesviral infections",
                    node_type="category",
                    semantic_level="three_character",
                    sort_order=7,
                    parent_code="B00-B49",
                    chapter_code="I",
                    chapter_title="Certain infectious and parasitic diseases",
                    block_code="B00-B99",
                    block_title="Nested root block",
                    three_character_code="B00",
                    three_character_title="Herpesviral infections",
                    has_children=False,
                    is_leaf=True,
                    is_three_character_code=True,
                    is_detailed_code=False,
                    is_coding_selectable=True,
                    sex_selectable="both",
                    age_group_selectable="all",
                    policy_status="unreviewed",
                    source_version="ICD-10-2019",
                    source_path="test",
                    is_active=True,
                    created_at=now,
                    updated_at=now,
                ),
            ]
        )
        db.session.commit()
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get(
            "/admin/api/icd10/2019-2/children?parent_code=I&coding_filter=active"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        by_code = {row["code"]: row for row in payload["children"]}
        self.assertEqual(by_code["B00-B99"]["child_count"], 1)
        self.assertEqual(by_code["B00-B99"]["status_indicator"], "green")

    def test_admin_node_api_returns_details(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/api/icd10/2019-2/node/A00")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["code"], "A00")
        self.assertEqual(payload["status_indicator"], "yellow")
        self.assertEqual([row["code"] for row in payload["ancestors"]], ["I", "A00-A09"])

    def test_admin_policy_patch_updates_node(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.patch(
            "/admin/api/icd10/2019-2/node/A00/policy",
            json={
                "is_coding_selectable": True,
                "sex_selectable": "both",
                "age_group_selectable": "adult",
                "restriction_note": "preferred default code",
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["sex_selectable"], "both")

        refreshed = db.session.get(MasIcd1020192, "A00")
        self.assertTrue(refreshed.is_coding_selectable)
        self.assertEqual(refreshed.restriction_note, "preferred default code")

    def test_admin_policy_export_returns_only_curated_code_rows(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/api/icd10/2019-2/policy-export")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/json")
        self.assertIn("attachment;", response.headers["Content-Disposition"])
        payload = response.get_json()
        self.assertEqual(payload["row_count"], 1)
        self.assertEqual(payload["items"][0]["code"], "A00.0")
        self.assertEqual(payload["items"][0]["sex_selectable"], "female")

    def test_admin_policy_xlsx_export_returns_policy_and_cod_override_status(self):
        scheme = MasCodBucketScheme(
            scheme_code=f"TEST_ICD_XLSX_{uuid.uuid4().hex[:8].upper()}",
            scheme_name="ICD XLSX Test Scheme",
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        db.session.add(
            MasCodBucketSchemeAgeBand(
                scheme_id=scheme.scheme_id,
                age_scope="all_ages",
                age_label="All Ages",
                min_age_value=0,
                min_age_unit="days",
                max_age_value=120,
                max_age_unit="years",
                level_count=1,
                sort_order=1,
                is_active=True,
            )
        )
        node = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope="all_ages",
            node_type="field",
            node_code="cholera",
            node_label="Cholera bucket",
            sort_order=1,
            is_active=True,
        )
        db.session.add(node)
        db.session.flush()
        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope="all_ages",
                icd_code="A00.0",
                node_id=node.node_id,
                match_type="manual_override",
                source_sheet="admin_cod_bucket_editor",
                mapping_note="Manual override to default COD bucket scheme mapping.",
                is_active=True,
            )
        )
        db.session.commit()
        self._login(str(self.base_admin_user.user_id))

        response = self.client.get("/admin/api/icd10/2019-2/policy-export.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            'attachment; filename="icd10_2019_2_policy_export.xlsx"',
            response.headers.get("Content-Disposition", ""),
        )
        workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=True)
        sheet = workbook["ICD10 Policy"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        a000 = next(row for row in rows if row["ICD Code"] == "A00.0")
        self.assertEqual(a000["Coding Allowed"], "Yes")
        self.assertEqual(a000["Age Selectable"], "adult")
        self.assertEqual(a000["Sex Selectable"], "female")
        self.assertEqual(a000["COD Manual Override"], "Yes")
        self.assertIn("ICD XLSX Test Scheme", a000["Manual Override Schemes"])

    def test_admin_policy_import_updates_code_rows(self):
        self._login(str(self.base_admin_user.user_id))

        payload = {
            "items": [
                {
                    "code": "A00",
                    "is_coding_selectable": True,
                    "sex_selectable": "both",
                    "age_group_selectable": "adult",
                }
            ]
        }

        response = self.client.post(
            "/admin/api/icd10/2019-2/policy-import",
            data={"file": (io.BytesIO(json.dumps(payload).encode("utf-8")), "policy.json")},
            headers=self._csrf_headers(),
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["total_items"], 1)
        self.assertEqual(body["updated_items"], 1)
        self.assertEqual(body["reset_items"], 1)
        self.assertEqual(body["failed_codes"], [])

        refreshed = db.session.get(MasIcd1020192, "A00")
        self.assertTrue(refreshed.is_coding_selectable)
        self.assertEqual(refreshed.sex_selectable, "both")
        self.assertEqual(refreshed.age_group_selectable, "adult")

        reset_row = db.session.get(MasIcd1020192, "A00.0")
        self.assertFalse(reset_row.is_coding_selectable)
        self.assertIsNone(reset_row.sex_selectable)
        self.assertIsNone(reset_row.age_group_selectable)

    def test_admin_policy_import_reports_unknown_codes(self):
        self._login(str(self.base_admin_user.user_id))

        payload = {
            "items": [
                {
                    "code": "ZZZ",
                    "is_coding_selectable": True,
                    "sex_selectable": "both",
                    "age_group_selectable": "adult",
                }
            ]
        }

        response = self.client.post(
            "/admin/api/icd10/2019-2/policy-import",
            data={"file": (io.BytesIO(json.dumps(payload).encode("utf-8")), "policy.json")},
            headers=self._csrf_headers(),
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertEqual(body["updated_items"], 0)
        self.assertEqual(body["reset_items"], 2)
        self.assertEqual(body["failed_codes"], [{"code": "ZZZ", "reason": "unknown_or_non_editable_code"}])

    def test_admin_policy_patch_denied_for_project_pi(self):
        self._login(str(self.base_project_pi_user.user_id))

        response = self.client.patch(
            "/admin/api/icd10/2019-2/node/A00/policy",
            json={
                "is_coding_selectable": True,
                "sex_selectable": "both",
                "age_group_selectable": "adult",
                "restriction_note": "",
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 403)

    def test_admin_policy_patch_rejects_structural_node(self):
        self._login(str(self.base_admin_user.user_id))

        response = self.client.patch(
            "/admin/api/icd10/2019-2/node/I/policy",
            json={
                "is_coding_selectable": True,
                "sex_selectable": "both",
                "age_group_selectable": "all",
                "restriction_note": "",
            },
            headers=self._csrf_headers(),
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("three-character and detailed ICD codes", response.get_json()["error"])
