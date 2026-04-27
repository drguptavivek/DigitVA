from datetime import datetime, timezone
from decimal import Decimal
from tempfile import NamedTemporaryFile

import sqlalchemy as sa
from openpyxl import Workbook

from app import db
from app.models import (
    MapIcdCodBucket,
    MasCodBucketNode,
    MasCodBucketScheme,
    MasCodBucketSchemeAgeBand,
    MasIcd1020192,
    VaFinalAssessments,
    VaForms,
    VaResearchProjects,
    VaSites,
    VaStatuses,
    VaSubmissionWorkflow,
    VaSubmissions,
)
from app.services.cod_bucket_mapping_service import (
    AGE_SCOPE_ADULT_OVER5Y,
    AGE_SCOPE_CHILD_1_59M,
    DEFAULT_MAX_AGE_UNIT,
    DEFAULT_MAX_AGE_VALUE,
    DEFAULT_MIN_AGE_UNIT,
    DEFAULT_MIN_AGE_VALUE,
    NODE_TYPE_CATEGORY,
    NODE_TYPE_FIELD,
    NODE_TYPE_SUBCATEGORY,
    SCHEME_CODE_CMEA10,
    SCHEME_CODE_SRS_INDIA,
    SCHEME_CODE_WHO_2022_VA,
    aggregate_coded_submissions_by_bucket,
    create_cod_bucket_scheme,
    import_cmea10_scheme,
    import_srs_india_scheme,
    import_who_2022_va_scheme,
    list_unmatched_coded_submission_icds_by_bucket,
    reset_cod_bucket_scheme_age_band_to_source,
    summarize_unmatched_coded_submissions_by_bucket,
)
from app.services.submission_analytics_mv import refresh_submission_analytics_mv
from app.services.submission_analytics_mv import (
    CORE_MV_NAME,
    COD_MV_NAME,
    DEMOGRAPHICS_MV_NAME,
    build_submission_analytics_core_mv_sql,
    build_submission_analytics_demographics_mv_sql,
    build_submission_cod_detail_mv_sql,
)
from tests.base import BaseTestCase


class CodBucketMappingServiceTests(BaseTestCase):
    PROJECT_ID = "CBMAP1"
    SITE_ID = "CBM1"
    FORM_ID = "CBMAP1CBM101"

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        now = datetime.now(timezone.utc)
        db.session.add(
            VaResearchProjects(
                project_id=cls.PROJECT_ID,
                project_code=cls.PROJECT_ID,
                project_name="COD Bucket Mapping Project",
                project_nickname="CodBucket",
                project_status=VaStatuses.active,
                project_registered_at=now,
                project_updated_at=now,
            )
        )
        db.session.flush()
        db.session.add(
            VaSites(
                site_id=cls.SITE_ID,
                project_id=cls.PROJECT_ID,
                site_name="COD Bucket Mapping Site",
                site_abbr=cls.SITE_ID,
                site_status=VaStatuses.active,
                site_registered_at=now,
                site_updated_at=now,
            )
        )
        db.session.flush()
        db.session.add(
            VaForms(
                form_id=cls.FORM_ID,
                project_id=cls.PROJECT_ID,
                site_id=cls.SITE_ID,
                odk_form_id="COD_BUCKET_FORM",
                odk_project_id="77",
                form_type="WHO VA 2022",
                form_status=VaStatuses.active,
                form_registered_at=now,
                form_updated_at=now,
            )
        )
        db.session.commit()

        for mv in (
            COD_MV_NAME,
            DEMOGRAPHICS_MV_NAME,
            CORE_MV_NAME,
            "va_submission_analytics_mv",
        ):
            db.session.execute(sa.text(f"DROP MATERIALIZED VIEW IF EXISTS {mv} CASCADE"))

        db.session.execute(sa.text(build_submission_analytics_core_mv_sql()))
        db.session.execute(
            sa.text(f"CREATE UNIQUE INDEX ix_test_cod_bucket_core_va_sid ON {CORE_MV_NAME} (va_sid)")
        )

        db.session.execute(sa.text(build_submission_analytics_demographics_mv_sql()))
        db.session.execute(
            sa.text(f"CREATE UNIQUE INDEX ix_test_cod_bucket_demo_va_sid ON {DEMOGRAPHICS_MV_NAME} (va_sid)")
        )

        db.session.execute(sa.text(build_submission_cod_detail_mv_sql()))
        db.session.execute(
            sa.text(f"CREATE UNIQUE INDEX ix_test_cod_bucket_detail_va_sid ON {COD_MV_NAME} (va_sid)")
        )
        db.session.commit()

    @classmethod
    def tearDownClass(cls):
        try:
            for mv in (
                COD_MV_NAME,
                DEMOGRAPHICS_MV_NAME,
                CORE_MV_NAME,
                "va_submission_analytics_mv",
            ):
                db.session.execute(sa.text(f"DROP MATERIALIZED VIEW IF EXISTS {mv} CASCADE"))
            db.session.commit()
        finally:
            super().tearDownClass()

    def _make_srs_workbook(self) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ICD_Mapped"
        sheet.append(
            [
                "disease_id",
                "icd_code",
                "icd_to_display",
                "category",
                "SRS_India_over5y_main_group",
                "SRS_India_over5y_sub_group",
                "SRS_India_over5y_disease",
                "SRS_India_over5y_match_type",
                "SRS_India_over5y_note",
                "SRS_India_neonate_main_group",
                "SRS_India_neonate_sub_group",
                "SRS_India_neonate_disease",
                "SRS_India_neonate_match_type",
                "SRS_India_neonate_note",
                "SRS_India_1_59mth_main_group",
                "SRS_India_1_59mth_sub_group",
                "SRS_India_1_59mth_disease",
                "SRS_India_1_59mth_type",
                "SRS_India_1_59mth_note",
            ]
        )
        sheet.append(
            [
                "1",
                "A00",
                "A00-Cholera",
                "Intestinal infectious diseases",
                "Adult Main",
                "Adult Sub",
                "Adult Disease",
                "range",
                "adult note",
                "Neonate Main",
                "Neonate Sub",
                "Neonate Disease",
                "range",
                "neonate note",
                "Child Main",
                "Child Sub",
                "Child Disease",
                "range",
                "child note",
            ]
        )
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            return tmp.name

    def _make_cmea_workbook(self) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ICD10_CMEA10"
        sheet.append(
            [
                "disease_id",
                "icd_code",
                "icd_to_display",
                "category",
                "CMEA10",
            ]
        )
        sheet.append(["1", "I21", "I21-Acute myocardial infarction", "Ischaemic", "Ischemic heart diseases"])
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            return tmp.name

    def _make_who_2022_va_workbook(self) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ICD_Mapped"
        sheet.append(
            [
                "disease_id",
                "icd_code",
                "icd_to_display",
                "category",
                "WHO_2022_VA_section",
                "WHO_2022_VA_code",
                "WHO_2022_VA_cause",
                "WHO_2022_VA_match_type",
                "WHO_2022_VA_note",
            ]
        )
        sheet.append(
            [
                1,
                "A33",
                "A33-Tetanus neonatorum",
                "Tetanus",
                "Neonatal causes of death",
                "VAs-10.05",
                "Neonatal tetanus",
                "exact",
                "Primary override",
            ]
        )
        sheet.append(
            [
                2,
                "O99.0",
                "O99.0-Anaemia complicating pregnancy, childbirth and the puerperium",
                "Maternal disorders",
                "Pregnancy-, childbirth and puerperium-related disorders",
                "VAs-09.07",
                "Anaemia of pregnancy",
                "exact",
                None,
            ]
        )
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            return tmp.name

    def _add_coded_submission(
        self,
        *,
        sid: str,
        icd: str,
        submitted_at: datetime,
        normalized_years: Decimal,
    ) -> None:
        db.session.add(
            VaSubmissions(
                va_sid=sid,
                va_form_id=self.FORM_ID,
                va_submission_date=submitted_at,
                va_odk_updatedat=submitted_at,
                va_data_collector="collector",
                va_odk_reviewstate="approved",
                va_instance_name=sid,
                va_uniqueid_real=sid,
                va_uniqueid_masked=sid,
                va_consent="yes",
                va_narration_language="English",
                va_deceased_age=0,
                va_deceased_age_normalized_days=normalized_years * Decimal("365.25"),
                va_deceased_age_normalized_years=normalized_years,
                va_deceased_age_source="test",
                va_deceased_gender="male",
                va_summary=[],
                va_catcount={},
                va_category_list=[],
            )
        )
        db.session.flush()
        db.session.add(
            VaSubmissionWorkflow(
                va_sid=sid,
                workflow_state="coder_finalized",
                workflow_reason="test",
                workflow_updated_by_role="vasystem",
            )
        )
        db.session.add(
            VaFinalAssessments(
                va_sid=sid,
                va_finassess_by=self.base_coder_user.user_id,
                va_conclusive_cod=icd,
                va_finassess_status=VaStatuses.active,
                va_finassess_createdat=submitted_at,
                va_finassess_updatedat=submitted_at,
            )
        )

    def test_import_srs_india_scheme_creates_age_scoped_hierarchy(self):
        workbook_path = self._make_srs_workbook()

        scheme = import_srs_india_scheme(workbook_path)

        self.assertEqual(scheme.scheme_code, SCHEME_CODE_SRS_INDIA)
        mapping_count = db.session.scalar(
            sa.select(sa.func.count()).select_from(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == scheme.scheme_id
            )
        )
        self.assertEqual(mapping_count, 3)
        adult_field = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_ADULT_OVER5Y,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
                MasCodBucketNode.node_label == "Adult Disease",
            )
        )
        self.assertIsNotNone(adult_field)
        age_bands = db.session.scalars(
            sa.select(MasCodBucketSchemeAgeBand).where(
                MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id
            )
        ).all()
        self.assertEqual(len(age_bands), 3)

    def test_import_cmea10_scheme_creates_flat_field_mapping(self):
        workbook_path = self._make_cmea_workbook()

        scheme = import_cmea10_scheme(workbook_path)

        self.assertEqual(scheme.scheme_code, SCHEME_CODE_CMEA10)
        field_nodes = db.session.scalars(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        ).all()
        self.assertEqual([node.node_label for node in field_nodes], ["Ischemic heart diseases"])
        age_band = db.session.scalar(
            sa.select(MasCodBucketSchemeAgeBand).where(
                MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id
            )
        )
        self.assertEqual(age_band.age_label, "All Ages")
        self.assertEqual(age_band.min_age_value, DEFAULT_MIN_AGE_VALUE)
        self.assertEqual(age_band.min_age_unit, DEFAULT_MIN_AGE_UNIT)
        self.assertEqual(age_band.max_age_value, DEFAULT_MAX_AGE_VALUE)
        self.assertEqual(age_band.max_age_unit, DEFAULT_MAX_AGE_UNIT)

    def test_import_who_2022_va_scheme_creates_section_and_va_leaf_mappings(self):
        workbook_path = self._make_who_2022_va_workbook()

        scheme = import_who_2022_va_scheme(workbook_path)

        self.assertEqual(scheme.scheme_code, SCHEME_CODE_WHO_2022_VA)
        self.assertEqual(scheme.source_path, workbook_path)
        age_band = db.session.scalar(
            sa.select(MasCodBucketSchemeAgeBand).where(
                MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id
            )
        )
        self.assertEqual(age_band.age_label, "All Ages")
        self.assertEqual(age_band.level_count, 2)
        section = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.node_type == NODE_TYPE_CATEGORY,
                MasCodBucketNode.node_label == "Neonatal causes of death",
            )
        )
        self.assertIsNotNone(section)
        leaf = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
                MasCodBucketNode.node_label == "Neonatal tetanus",
            )
        )
        self.assertIsNotNone(leaf)
        self.assertEqual(leaf.parent_node_id, section.node_id)
        mapping_count = db.session.scalar(
            sa.select(sa.func.count()).select_from(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == scheme.scheme_id
            )
        )
        self.assertEqual(mapping_count, 2)
        mapping = db.session.scalar(
            sa.select(MapIcdCodBucket).where(
                MapIcdCodBucket.scheme_id == scheme.scheme_id,
                MapIcdCodBucket.icd_code == "A33",
            )
        )
        self.assertEqual(mapping.node_id, leaf.node_id)
        self.assertEqual(mapping.match_type, "exact")
        self.assertEqual(mapping.mapping_note, "Primary override")

    def test_reset_srs_age_band_to_source_restores_selected_scope_only(self):
        workbook_path = self._make_srs_workbook()
        scheme = import_srs_india_scheme(workbook_path)

        adult_field = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_ADULT_OVER5Y,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        child_field = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_CHILD_1_59M,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        adult_field.node_label = "Changed Adult Disease"
        child_field.node_label = "Changed Child Disease"
        db.session.commit()

        reset_cod_bucket_scheme_age_band_to_source(
            scheme_code=scheme.scheme_code,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
        )

        restored_adult = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_ADULT_OVER5Y,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        untouched_child = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_CHILD_1_59M,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        self.assertEqual(restored_adult.node_label, "Adult Disease")
        self.assertEqual(untouched_child.node_label, "Changed Child Disease")

    def test_reset_srs_scheme_to_source_restores_all_scopes(self):
        workbook_path = self._make_srs_workbook()
        scheme = import_srs_india_scheme(workbook_path)

        adult_field = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_ADULT_OVER5Y,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        child_field = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_CHILD_1_59M,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        adult_field.node_label = "Changed Adult Disease"
        child_field.node_label = "Changed Child Disease"
        db.session.commit()

        reset_cod_bucket_scheme_age_band_to_source(
            scheme_code=scheme.scheme_code,
            age_scope=None,
            reset_entire_scheme=True,
        )

        restored_adult = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_ADULT_OVER5Y,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        restored_child = db.session.scalar(
            sa.select(MasCodBucketNode).where(
                MasCodBucketNode.scheme_id == scheme.scheme_id,
                MasCodBucketNode.age_scope == AGE_SCOPE_CHILD_1_59M,
                MasCodBucketNode.node_type == NODE_TYPE_FIELD,
            )
        )
        self.assertEqual(restored_adult.node_label, "Adult Disease")
        self.assertEqual(restored_child.node_label, "Child Disease")

    def test_create_cod_bucket_scheme_returns_non_blocking_gap_feedback(self):
        scheme, warnings = create_cod_bucket_scheme(
            scheme_name="Created Scheme",
            scheme_code="CREATED_SCHEME",
            age_bands=[
                {
                    "age_label": "Neonate",
                    "min_age_value": 0,
                    "min_age_unit": "days",
                    "max_age_value": 28,
                    "max_age_unit": "days",
                    "level_count": 3,
                },
                {
                    "age_label": "Adult",
                    "min_age_value": 5,
                    "min_age_unit": "years",
                    "max_age_value": 120,
                    "max_age_unit": "years",
                    "level_count": 2,
                },
            ],
        )

        self.assertEqual(scheme.scheme_code, "CREATED_SCHEME")
        self.assertTrue(warnings)
        self.assertEqual(
            db.session.scalar(
                sa.select(sa.func.count())
                .select_from(MasCodBucketSchemeAgeBand)
                .where(MasCodBucketSchemeAgeBand.scheme_id == scheme.scheme_id)
            ),
            2,
        )

    def test_create_cod_bucket_scheme_treats_upper_bound_as_exclusive(self):
        scheme, warnings = create_cod_bucket_scheme(
            scheme_name="Adjacent Scheme",
            scheme_code="ADJACENT_SCHEME",
            age_bands=[
                {
                    "age_label": "Band A",
                    "min_age_value": 0,
                    "min_age_unit": "days",
                    "max_age_value": 29,
                    "max_age_unit": "days",
                    "level_count": 1,
                },
                {
                    "age_label": "Band B",
                    "min_age_value": 29,
                    "min_age_unit": "days",
                    "max_age_value": 60,
                    "max_age_unit": "days",
                    "level_count": 1,
                },
            ],
        )

        self.assertEqual(scheme.scheme_code, "ADJACENT_SCHEME")
        self.assertEqual(warnings, [])

    def test_aggregate_coded_submissions_by_bucket_uses_age_scope(self):
        db.session.execute(
            sa.delete(VaFinalAssessments).where(
                VaFinalAssessments.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(
            sa.delete(VaSubmissionWorkflow).where(
                VaSubmissionWorkflow.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(sa.delete(VaSubmissions).where(VaSubmissions.va_form_id == self.FORM_ID))
        db.session.commit()

        now = datetime.now(timezone.utc)
        scheme = MasCodBucketScheme(
            scheme_code="TEST_SRS",
            scheme_name="Test SRS",
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        db.session.add_all(
            [
                MasCodBucketSchemeAgeBand(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_ADULT_OVER5Y,
                    age_label="Adult / Over 5 Years",
                    min_age_value=5,
                    min_age_unit="years",
                    max_age_value=120,
                    max_age_unit="years",
                    level_count=3,
                    sort_order=1,
                    is_active=True,
                ),
                MasCodBucketSchemeAgeBand(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_CHILD_1_59M,
                    age_label="Child / 1-59 Months",
                    min_age_value=1,
                    min_age_unit="months",
                    max_age_value=59,
                    max_age_unit="months",
                    level_count=2,
                    sort_order=2,
                    is_active=True,
                ),
            ]
        )

        adult_category = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_CATEGORY,
            node_code="adult_main",
            node_label="Adult Main",
            sort_order=1,
        )
        adult_subcategory = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_SUBCATEGORY,
            parent=adult_category,
            node_code="adult_sub",
            node_label="Adult Sub",
            sort_order=1,
        )
        adult_field = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_FIELD,
            parent=adult_subcategory,
            node_code="adult_disease",
            node_label="Adult Disease",
            sort_order=1,
        )
        child_category = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_CHILD_1_59M,
            node_type=NODE_TYPE_CATEGORY,
            node_code="child_main",
            node_label="Child Main",
            sort_order=1,
        )
        child_field = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_CHILD_1_59M,
            node_type=NODE_TYPE_FIELD,
            parent=child_category,
            node_code="child_disease",
            node_label="Child Disease",
            sort_order=1,
        )
        db.session.add_all(
            [
                adult_category,
                adult_subcategory,
                adult_field,
                child_category,
                child_field,
            ]
        )
        db.session.flush()

        db.session.add_all(
            [
                MapIcdCodBucket(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_ADULT_OVER5Y,
                    icd_code="I21",
                    node_id=adult_field.node_id,
                    is_active=True,
                ),
                MapIcdCodBucket(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_CHILD_1_59M,
                    icd_code="A00",
                    node_id=child_field.node_id,
                    is_active=True,
                ),
            ]
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-adult",
            icd="I21",
            submitted_at=now,
            normalized_years=Decimal("52"),
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-child",
            icd="A00",
            submitted_at=now,
            normalized_years=Decimal("4"),
        )
        db.session.commit()
        refresh_submission_analytics_mv(concurrently=False)

        rows = aggregate_coded_submissions_by_bucket(
            scheme_code="TEST_SRS",
            form_id=self.FORM_ID,
        )

        self.assertEqual(len(rows), 2)
        adult_row = next(row for row in rows if row["age_scope"] == AGE_SCOPE_ADULT_OVER5Y)
        self.assertEqual(adult_row["bucket_category"], "Adult Main")
        self.assertEqual(adult_row["bucket_subcategory"], "Adult Sub")
        self.assertEqual(adult_row["bucket_field"], "Adult Disease")
        self.assertEqual(adult_row["coded_count"], 1)

        child_row = next(row for row in rows if row["age_scope"] == AGE_SCOPE_CHILD_1_59M)
        self.assertEqual(child_row["bucket_category"], "Child Main")
        self.assertIsNone(child_row["bucket_subcategory"])
        self.assertEqual(child_row["bucket_field"], "Child Disease")
        self.assertEqual(child_row["coded_count"], 1)

    def test_aggregate_coded_submissions_by_bucket_preserves_display_order(self):
        db.session.execute(
            sa.delete(VaFinalAssessments).where(
                VaFinalAssessments.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(
            sa.delete(VaSubmissionWorkflow).where(
                VaSubmissionWorkflow.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(sa.delete(VaSubmissions).where(VaSubmissions.va_form_id == self.FORM_ID))
        db.session.commit()

        now = datetime.now(timezone.utc)
        scheme = MasCodBucketScheme(
            scheme_code="TEST_ORDER",
            scheme_name="Test Display Order",
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        db.session.add(
            MasCodBucketSchemeAgeBand(
                scheme_id=scheme.scheme_id,
                age_scope=AGE_SCOPE_ADULT_OVER5Y,
                age_label="Adult / Over 5 Years",
                min_age_value=5,
                min_age_unit="years",
                max_age_value=120,
                max_age_unit="years",
                level_count=2,
                sort_order=1,
                is_active=True,
            )
        )

        category_b = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_CATEGORY,
            node_code="category_b",
            node_label="Category B",
            sort_order=2,
        )
        field_b = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_FIELD,
            parent=category_b,
            node_code="field_b",
            node_label="Disease B",
            sort_order=2,
        )
        category_a = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_CATEGORY,
            node_code="category_a",
            node_label="Category A",
            sort_order=1,
        )
        field_a = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_FIELD,
            parent=category_a,
            node_code="field_a",
            node_label="Disease A",
            sort_order=1,
        )
        db.session.add_all([category_b, field_b, category_a, field_a])
        db.session.flush()

        db.session.add_all(
            [
                MapIcdCodBucket(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_ADULT_OVER5Y,
                    icd_code="I21",
                    node_id=field_b.node_id,
                    is_active=True,
                ),
                MapIcdCodBucket(
                    scheme_id=scheme.scheme_id,
                    age_scope=AGE_SCOPE_ADULT_OVER5Y,
                    icd_code="I22",
                    node_id=field_a.node_id,
                    is_active=True,
                ),
            ]
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-order-1",
            icd="I21",
            submitted_at=now,
            normalized_years=Decimal("50"),
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-order-2",
            icd="I21",
            submitted_at=now,
            normalized_years=Decimal("51"),
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-order-3",
            icd="I22",
            submitted_at=now,
            normalized_years=Decimal("52"),
        )
        db.session.commit()
        refresh_submission_analytics_mv(concurrently=False)

        rows = aggregate_coded_submissions_by_bucket(
            scheme_code="TEST_ORDER",
            form_id=self.FORM_ID,
            collapse_scope=True,
        )

        self.assertEqual([row["bucket_category"] for row in rows], ["Category A", "Category B"])
        self.assertEqual(rows[0]["bucket_category_sort_order"], 1)
        self.assertEqual(rows[0]["bucket_field_sort_order"], 1)
        self.assertEqual(rows[0]["coded_count"], 1)
        self.assertEqual(rows[1]["bucket_category_sort_order"], 2)
        self.assertEqual(rows[1]["bucket_field_sort_order"], 2)
        self.assertEqual(rows[1]["coded_count"], 2)

    def test_summarize_unmatched_coded_submissions_by_bucket_counts_dropped_icds(self):
        db.session.execute(
            sa.delete(VaFinalAssessments).where(
                VaFinalAssessments.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(
            sa.delete(VaSubmissionWorkflow).where(
                VaSubmissionWorkflow.va_sid.in_(
                    sa.select(VaSubmissions.va_sid).where(VaSubmissions.va_form_id == self.FORM_ID)
                )
            )
        )
        db.session.execute(sa.delete(VaSubmissions).where(VaSubmissions.va_form_id == self.FORM_ID))
        db.session.commit()

        now = datetime.now(timezone.utc)
        scheme = MasCodBucketScheme(
            scheme_code="TEST_UNMATCHED",
            scheme_name="Test Unmatched Summary",
            mapping_version=1,
            is_active=True,
        )
        db.session.add(scheme)
        db.session.flush()
        db.session.add(
            MasCodBucketSchemeAgeBand(
                scheme_id=scheme.scheme_id,
                age_scope=AGE_SCOPE_ADULT_OVER5Y,
                age_label="Adult / Over 5 Years",
                min_age_value=5,
                min_age_unit="years",
                max_age_value=120,
                max_age_unit="years",
                level_count=1,
                sort_order=1,
                is_active=True,
            )
        )
        field = MasCodBucketNode(
            scheme_id=scheme.scheme_id,
            age_scope=AGE_SCOPE_ADULT_OVER5Y,
            node_type=NODE_TYPE_FIELD,
            node_code="matched_field",
            node_label="Matched Disease",
            sort_order=1,
        )
        db.session.add(field)
        db.session.flush()
        db.session.add(
            MapIcdCodBucket(
                scheme_id=scheme.scheme_id,
                age_scope=AGE_SCOPE_ADULT_OVER5Y,
                icd_code="I21",
                node_id=field.node_id,
                is_active=True,
            )
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-matched",
            icd="I21",
            submitted_at=now,
            normalized_years=Decimal("52"),
        )
        self._add_coded_submission(
            sid="uuid:cod-bucket-unmatched",
            icd="I22",
            submitted_at=now,
            normalized_years=Decimal("53"),
        )
        db.session.merge(
            MasIcd1020192(
                code="I22",
                title="Subsequent myocardial infarction",
                node_type="category",
                semantic_level="three_character",
                sort_order=1,
                parent_code=None,
                chapter_code="IX",
                chapter_title="Diseases of the circulatory system",
                block_code="I20-I25",
                block_title="Ischaemic heart diseases",
                three_character_code="I22",
                three_character_title="Subsequent myocardial infarction",
                has_children=False,
                is_leaf=True,
                is_three_character_code=True,
                is_detailed_code=False,
                is_coding_selectable=True,
                sex_selectable="both",
                age_group_selectable="all",
                policy_status="allowed",
                source_version="2019-test",
                source_path="tests",
                is_active=True,
            )
        )
        db.session.commit()
        refresh_submission_analytics_mv(concurrently=False)

        rows = aggregate_coded_submissions_by_bucket(
            scheme_code="TEST_UNMATCHED",
            form_id=self.FORM_ID,
            collapse_scope=True,
        )
        unmatched_rows = summarize_unmatched_coded_submissions_by_bucket(
            scheme_code="TEST_UNMATCHED",
            form_id=self.FORM_ID,
            collapse_scope=True,
        )
        unmatched_icd_rows = list_unmatched_coded_submission_icds_by_bucket(
            scheme_code="TEST_UNMATCHED",
            form_id=self.FORM_ID,
            collapse_scope=True,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["bucket_field"], "Matched Disease")
        self.assertEqual(rows[0]["coded_count"], 1)
        self.assertEqual(
            unmatched_rows,
            [
                {
                    "age_scope": AGE_SCOPE_ADULT_OVER5Y,
                    "unmatched_count": 1,
                }
            ],
        )
        self.assertEqual(
            unmatched_icd_rows,
            [
                {
                    "age_scope": AGE_SCOPE_ADULT_OVER5Y,
                    "icd_code": "I22",
                    "unmatched_count": 1,
                    "category": "not_included_in_scheme",
                    "category_label": "ICD codes not included in CoD Categories",
                    "is_master_coding_eligible": True,
                }
            ],
        )
