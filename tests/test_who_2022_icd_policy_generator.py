import csv
import json
from pathlib import Path

from openpyxl import Workbook

from scripts.generate_who_2022_icd_policy import generate_policy


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    crosswalk = workbook.active
    crosswalk.title = "VA_2022_Crosswalk"
    crosswalk.append(["section", "va_code", "va_title", "icd10_codes_raw", "notes"])
    crosswalk.append(["Infectious", "VAs-01.04", "Diarrheal diseases", "A00", None])
    crosswalk.append(["External", "VAs-12.99", "Injuries", "S00-T99", None])
    crosswalk.append(["Perinatal", "VAs-10.99", "Perinatal", "P95; R95", None])
    crosswalk.append(["Maternal", "VAs-09.01", "Maternal", "O00-O99", None])
    crosswalk.append(["Neoplasm", "VAs-02.01", "Neoplasm", "C50-C63", None])

    road = workbook.create_sheet("RoadTraffic_Footnote")
    road.append(["va_code", "detail"])
    road.append(["VAs-12.01 / VAs-12.02", "Road traffic codes: V01.1; V10.4-V10.9"])

    notes = workbook.create_sheet("Notes")
    notes.append(["topic", "note"])
    workbook.save(path)


def _write_icd_csv(path: Path) -> None:
    fields = [
        "code",
        "title",
        "semantic_level",
        "sort_order",
        "chapter_code",
        "chapter_title",
        "block_code",
        "block_title",
        "three_character_code",
        "three_character_title",
        "is_active",
    ]
    rows = [
        ("A00", "Cholera", "three_character", 1),
        ("A00.0", "Cholera due to Vibrio cholerae", "detailed_code", 2),
        ("S00", "Superficial injury of head", "three_character", 3),
        ("S00.0", "Superficial injury of scalp", "detailed_code", 4),
        ("T99", "Other effects of external causes", "three_character", 5),
        ("V01", "Pedestrian injured in collision with pedal cycle", "three_character", 6),
        ("V01.1", "Pedestrian injured in collision with pedal cycle", "detailed_code", 7),
        ("V10", "Pedal cyclist injured in collision with pedestrian", "three_character", 8),
        ("V10.4", "Pedal cyclist injured in transport accident", "detailed_code", 9),
        ("P95", "Fetal death of unspecified cause", "three_character", 10),
        ("R95", "Sudden infant death syndrome", "three_character", 11),
        ("O00", "Ectopic pregnancy", "three_character", 12),
        ("O00.1", "Tubal pregnancy", "detailed_code", 13),
        ("C50", "Malignant neoplasm of breast", "three_character", 14),
        ("C50.0", "Malignant neoplasm of nipple and areola", "detailed_code", 15),
        ("C51", "Malignant neoplasm of vulva", "three_character", 16),
        ("C51.0", "Malignant neoplasm of labium majus", "detailed_code", 17),
        ("C60", "Malignant neoplasm of penis", "three_character", 18),
        ("C60.0", "Malignant neoplasm of prepuce", "detailed_code", 19),
    ]

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for code, title, semantic_level, sort_order in rows:
            writer.writerow(
                {
                    "code": code,
                    "title": title,
                    "semantic_level": semantic_level,
                    "sort_order": str(sort_order),
                    "chapter_code": code[0],
                    "chapter_title": "Chapter",
                    "block_code": "",
                    "block_title": "",
                    "three_character_code": code.split(".")[0],
                    "three_character_title": title,
                    "is_active": "true",
                }
            )


def test_who_2022_policy_generator_applies_allowability_rules(tmp_path):
    workbook_path = tmp_path / "WHO_2022_VA_Crosswalk.xlsx"
    icd_csv_path = tmp_path / "icd10.csv"
    output_path = tmp_path / "who_2022_icd10_2019_2_policy.json"
    _write_workbook(workbook_path)
    _write_icd_csv(icd_csv_path)

    payload = generate_policy(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        output_path=output_path,
    )

    assert output_path.exists()
    assert json.loads(output_path.read_text(encoding="utf-8")) == payload
    assert payload["source_version"] == "WHO_2022_ICD10_2019_2"
    assert payload["row_count"] == len(payload["items"])

    by_code = {item["code"]: item for item in payload["items"]}
    assert by_code["A00"]["sex_selectable"] == "both"
    assert by_code["A00"]["age_group_selectable"] == "all"
    assert "A00.0" not in by_code
    assert "S00" not in by_code
    assert "S00.0" not in by_code
    assert "T99" not in by_code
    assert "V01" not in by_code
    assert by_code["V01.1"]["sex_selectable"] == "both"
    assert by_code["V01.1"]["age_group_selectable"] == "all"
    assert "V10" not in by_code
    assert by_code["V10.4"]["age_group_selectable"] == "all"
    assert by_code["P95"]["age_group_selectable"] == "neonate"
    assert by_code["R95"]["age_group_selectable"] == "infant"
    assert by_code["O00"]["sex_selectable"] == "female"
    assert by_code["O00"]["age_group_selectable"] == "adult"
    assert "O00.1" not in by_code
    assert by_code["C51"]["sex_selectable"] == "female"
    assert by_code["C51"]["age_group_selectable"] == "all"
    assert "C51.0" not in by_code
    assert by_code["C60"]["sex_selectable"] == "male"
    assert by_code["C60"]["age_group_selectable"] == "all"
    assert "C60.0" not in by_code
    assert by_code["C50"]["sex_selectable"] == "both"
    assert by_code["C50"]["age_group_selectable"] == "all"
    assert "C50.0" not in by_code
