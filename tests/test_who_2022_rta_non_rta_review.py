import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.generate_who_2022_cod_bucket_workbook import generate_workbook
from scripts.generate_who_2022_icd_policy import generate_policy
from scripts.generate_who_2022_rta_non_rta_review import generate_review_workbook


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    crosswalk = workbook.active
    crosswalk.title = "VA_2022_Crosswalk"
    crosswalk.append(["section", "va_code", "va_title", "icd10_codes_raw", "notes"])
    crosswalk.append(["External", "VAs-12.01", "Road traffic accident", "See RoadTraffic_Footnote sheet", None])
    crosswalk.append(["External", "VAs-12.02", "Other transport accident", "Transport codes not qualifying for VAs-12.01", None])
    road = workbook.create_sheet("RoadTraffic_Footnote")
    road.append(["va_code", "detail"])
    road.append(
        [
            "VAs-12.01 / VAs-12.02",
            "V10.4-V10.9; V81.1-V81.9; V82.1-V82.9; V90-V99; Y85.9",
        ]
    )
    workbook.save(path)


def _write_icd_csv(path: Path) -> None:
    fields = [
        "code",
        "title",
        "semantic_level",
        "sort_order",
        "chapter_code",
        "block_code",
        "three_character_code",
        "is_active",
    ]
    rows = [
        ("V10.4", "Driver injured in traffic accident", "detailed_code", 1),
        ("V81.1", "Occupant of railway train injured in collision with motor vehicle in traffic accident", "detailed_code", 2),
        ("V81.2", "Occupant of railway train injured in collision with or hit by rolling stock", "detailed_code", 3),
        ("V82.1", "Occupant of streetcar injured in collision with motor vehicle in traffic accident", "detailed_code", 4),
        ("V82.2", "Occupant of streetcar injured in collision with or hit by rolling stock", "detailed_code", 5),
        ("V82.5", "Occupant of streetcar injured by fall in streetcar", "detailed_code", 6),
        ("V90", "Accident to watercraft causing drowning and submersion", "three_character", 7),
        ("V90.0", "Accident to watercraft causing drowning and submersion : Merchant ship", "detailed_code", 8),
        ("Y85.9", "Sequelae of other and unspecified transport accidents", "detailed_code", 9),
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for code, title, semantic_level, sort_order in rows:
            writer.writerow(
                {
                    "code": code,
                    "title": title,
                    "semantic_level": semantic_level,
                    "sort_order": sort_order,
                    "chapter_code": code[0],
                    "block_code": "",
                    "three_character_code": code.split(".", 1)[0],
                    "is_active": "true",
                }
            )


def test_review_workbook_splits_road_and_non_road_transport_codes(tmp_path):
    workbook_path = tmp_path / "crosswalk.xlsx"
    icd_csv_path = tmp_path / "icd.csv"
    output_path = tmp_path / "review.xlsx"
    _write_workbook(workbook_path)
    _write_icd_csv(icd_csv_path)

    summary = generate_review_workbook(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        output_path=output_path,
    )

    assert summary["transport_codes_reviewed"] == 9
    assert output_path.exists()
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    sheet = workbook["RTA_NonRTA_Review"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = {
        row["icd_code"]: row
        for row in (
            dict(zip(headers, values))
            for values in sheet.iter_rows(min_row=2, values_only=True)
        )
    }
    assert rows["V10.4"]["proposed_va_code"] == "VAs-12.01"
    assert rows["V81.1"]["proposed_va_code"] == "VAs-12.01"
    assert rows["V81.2"]["proposed_va_code"] == "VAs-12.02"
    assert rows["V82.1"]["proposed_va_code"] == "VAs-12.01"
    assert rows["V82.2"]["proposed_va_code"] == "VAs-12.02"
    assert rows["V82.5"]["proposed_va_code"] == "VAs-12.02"
    assert rows["V90"]["proposed_va_code"] == "VAs-12.02"
    assert rows["V90.0"]["proposed_va_code"] == "VAs-12.02"
    assert rows["Y85.9"]["proposed_va_code"] == "VAs-12.02"


def test_cod_bucket_workbook_uses_reviewed_transport_decisions(tmp_path):
    workbook_path = tmp_path / "crosswalk.xlsx"
    icd_csv_path = tmp_path / "icd.csv"
    review_path = tmp_path / "review.xlsx"
    policy_path = tmp_path / "policy.json"
    output_path = tmp_path / "bucket.xlsx"
    _write_workbook(workbook_path)
    _write_icd_csv(icd_csv_path)
    generate_review_workbook(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        output_path=review_path,
    )
    policy_path.write_text(
        """
{
  "source_version": "test",
  "row_count": 3,
  "items": [
    {"code": "V10.4", "title": "Driver injured in traffic accident", "semantic_level": "detailed_code", "chapter_code": "XX", "chapter_title": "External", "block_code": "V01-X59", "block_title": "Accidents", "three_character_code": "V10", "three_character_title": "Pedal cyclist", "sex_selectable": "both", "age_group_selectable": "all"},
    {"code": "V81.2", "title": "Occupant of railway train injured in collision with or hit by rolling stock", "semantic_level": "detailed_code", "chapter_code": "XX", "chapter_title": "External", "block_code": "V01-X59", "block_title": "Accidents", "three_character_code": "V81", "three_character_title": "Railway train", "sex_selectable": "both", "age_group_selectable": "all"},
    {"code": "V90", "title": "Accident to watercraft causing drowning and submersion", "semantic_level": "three_character", "chapter_code": "XX", "chapter_title": "External", "block_code": "V01-X59", "block_title": "Accidents", "three_character_code": "V90", "three_character_title": "Watercraft", "sex_selectable": "both", "age_group_selectable": "all"}
  ]
}
""".strip()
        + "\n",
        encoding="utf-8",
    )

    generate_workbook(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        policy_path=policy_path,
        output_path=output_path,
        rta_review_path=review_path,
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    sheet = workbook["ICD_Mapped"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = {
        row["icd_code"]: row
        for row in (
            dict(zip(headers, values))
            for values in sheet.iter_rows(min_row=2, values_only=True)
        )
    }
    assert rows["V10.4"]["WHO_2022_VA_code"] == "VAs-12.01"
    assert rows["V81.2"]["WHO_2022_VA_code"] == "VAs-12.02"
    assert rows["V90"]["WHO_2022_VA_code"] == "VAs-12.02"


def test_policy_generator_includes_reviewed_non_road_transport_codes(tmp_path):
    workbook_path = tmp_path / "crosswalk.xlsx"
    icd_csv_path = tmp_path / "icd.csv"
    review_path = tmp_path / "review.xlsx"
    policy_path = tmp_path / "policy.json"
    _write_workbook(workbook_path)
    _write_icd_csv(icd_csv_path)
    generate_review_workbook(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        output_path=review_path,
    )

    payload = generate_policy(
        workbook_path=workbook_path,
        icd_csv_path=icd_csv_path,
        output_path=policy_path,
        rta_review_path=review_path,
    )

    codes = {item["code"] for item in payload["items"]}
    assert {"V10.4", "V81.2", "V82.5", "V90", "V90.0", "Y85.9"} <= codes
